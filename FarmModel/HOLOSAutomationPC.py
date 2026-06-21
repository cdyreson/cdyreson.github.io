from appium import webdriver
from selenium.webdriver import DesiredCapabilities
from selenium.webdriver.common.action_chains import ActionChains
import xlrd
import random
from openpyxl import *
import time

RUNS=1

desired_caps = {}
#desired_caps["app"] = "C://Program Files (x86)//AAFC/Holos 3.0.6//Holos 3.0.exe"
desired_caps["app"] = "C:\Program Files (x86)\AAFC\Holos 3.0.6\Holos 3.0.exe"
driver = webdriver.Remote(command_executor='http://127.0.0.1:4723',desired_capabilities= desired_caps)


# reading holos inpput range xlsx sheet

#loc = ("C:/Research/HolosInputs.xlsx")
loc = ("D:/Study Material/Research/HolosInputs.xlsx")
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(0)
sheet.cell_value(0, 0)

#Output excel
#wb1=load_workbook("C:/Research/Output.xlsx")
wb1=load_workbook("D:/Study Material/Research/Output.xlsx")
ws=wb1["Sheet1"]

# opening Holos
driver.find_element_by_name("OK").click()

driver.switch_to.window(driver.window_handles[0])
driver.find_element_by_name("Holos Research").click()
driver.switch_to.window(driver.window_handles[0])

desired_caps = {}
desired_caps["app"] = "Root"
rm = webdriver.Remote(command_executor='http://127.0.0.1:4723', desired_capabilities=desired_caps)

for r in range(RUNS+1):
    try:



        # choosing new file to create
        menu = driver.find_element_by_name("File")
        menu.click()
        item = menu.find_element_by_name("New")
        actions = ActionChains(driver)
        actions.click(item)
        actions.perform()
        driver.switch_to.window(driver.window_handles[0])
        driver.find_element_by_name("Ok").click()

        # Detail On
        radioButton = driver.find_element_by_xpath("//RadioButton[@Name=\"On\"]")
        radioButton.click()

        driver.find_element_by_xpath("//Button[@Name=\"Maximize\"][@AutomationId=\"Maximize-Restore\"]").click()

        # # Setting Farm Name
        # farmName = driver.find_element_by_xpath("//Edit[@AutomationId='FarmName']")
        # farmName.click()
        # farmName.clear()
        # name = "farm1"
        # farmName.send_keys(name)
        #
        # Setting year
        startYear = sheet.cell_value(2, 1)
        endYear = sheet.cell_value(2, 2)
        year = random.randrange(int(startYear), int(endYear) + 1, 1)
        print(year)
        farmYear = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FarmYear\"]")
        farmYear.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()
        #
        # farmYear.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10,10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # farmYear.send_keys(int(year))

        #Setting Ecodistrict
        startEd=sheet.cell_value(3, 1)
        endEd=sheet.cell_value(3, 2)
        ecoDistrictChoice = random.randrange(int(startEd), int(endEd)+1, 1)
        ecoDistrict=driver.find_element_by_xpath("//ComboBox[@AutomationId='Ecodistrict']")
        ecoDistrict.send_keys(ecoDistrictChoice)

        #Setting Province
        pVal=sheet.cell(4,3)
        provincesList= str(pVal).split(',')
        provinceName=random.choices(provincesList)
        print(provinceName)
        province=driver.find_element_by_xpath("//ComboBox[@AutomationId='Province']")
        province.send_keys(provinceName)

        print(farmYear.text)
        #Soil Texture
        # (0,30) Fine
        # (0,50) Medium
        # (0,70) Coarse

        soilTextureList = [30, 50, 70]
        soilTexturePosi = random.choices(soilTextureList)
        soilTexture=driver.find_element_by_xpath("//ComboBox[@AutomationId='SoilTexture']")
        soilTexture.click()
        time.sleep(2)
        actions = ActionChains(driver)
        actions.move_by_offset(0, soilTexturePosi[0]).click()
        actions.perform()

        #setting SoilType
        # (0,30) Black/gray
        # (0,50) brown
        # (0,70) dark brown
        # (0,90) eastern canada
        soilTypeList= [30, 50, 70, 90]
        soilTypePosi=random.choices(soilTypeList)
        soilType=driver.find_element_by_xpath("//ComboBox[@AutomationId='SoilType']")
        soilType.click()
        time.sleep(2)
        actions = ActionChains(driver)
        actions.move_by_offset(0, soilTypePosi[0]).click()
        actions.perform()

        #setting soilIntensity
        # (0,30) No-till
        # (0,50) Reduced
        # (0,70) Intensive
        presentIntensityList = [30, 50, 70]
        presentIntensityPosi=random.choices(presentIntensityList)
        presentIntensity=driver.find_element_by_xpath("//ComboBox[@AutomationId='TillagePresentIntensity']")
        presentIntensity.click()
        time.sleep(2)
        actions = ActionChains(driver)
        actions.move_by_offset(0, presentIntensityPosi[0]).click()
        actions.perform()

        #setting past intensity
        # (0,30) No-till
        # (0,50) Reduced
        # (0,70) Intensive
        pastIntensityList = [30, 50, 70]
        pastIntensityName=random.choices(pastIntensityList)
        pastIntensity=driver.find_element_by_xpath("//ComboBox[@AutomationId='TillagePastIntensity']")
        pastIntensity.click()
        time.sleep(2)
        actions = ActionChains(driver)
        actions.move_by_offset(0, presentIntensityPosi[0]).click()
        actions.perform()

        #setting LumC max
        startLCM=sheet.cell_value(10, 1)
        endLCM=sheet.cell_value(10, 2)
        LCMChoice = random.randrange(int(startLCM), int(endLCM)+1, 1)
        LCM=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"TillageLumCMax\"]")
        LCM.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        LCM.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        LCM.send_keys(LCMChoice)

        #setting k
        startK=sheet.cell_value(11, 1)
        endK=sheet.cell_value(11, 2)
        KChoice = round(random.uniform(float(startK), float(endK)+1),2)
        print(KChoice)
        K=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"TillageK\"]")
        K.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        K.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        K.send_keys(str(KChoice))

        #setting precipitaion
        startPreci=sheet.cell_value(12, 1)
        endPreci=sheet.cell_value(12, 2)
        PreciChoice = random.randrange(int(startPreci), int(endPreci)+1, 1)
        precipiation=driver.find_element_by_xpath("//ComboBox[@Name=\"EF Volatilization (kg N2O-N kg^-1 N)\"]")
        precipiation.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        precipiation.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        precipiation.send_keys(PreciChoice)

        # Setting Potential Evapotranspiration
        startEvapotranspiration=sheet.cell_value(13, 1)
        endEvapotranspiration=sheet.cell_value(13, 2)
        evapotranspirationChoice = random.randrange(int(startEvapotranspiration), int(endEvapotranspiration)+1, 1)
        evapotranspiration=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"PotentialEvapotranspiration\"]")
        evapotranspiration.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        evapotranspiration.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        evapotranspiration.send_keys(evapotranspirationChoice)

        #setting F topograpy
        startTopo=sheet.cell_value(14, 1)
        endTopo=sheet.cell_value(14, 2)
        topoChoice = round(random.uniform(float(startTopo), float(endTopo)+1),2)
        topo=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FTopography\"]")
        topo.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        topo.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        topo.send_keys(str(topoChoice))



        #setting Soil N2O Breakdown
        #january
        startBreakdown=sheet.cell_value(15, 1)
        endBreakdown=sheet.cell_value(15, 2)
        breakdownChoice = random.randrange(int(startBreakdown), int(endBreakdown)+1, 1)
        breakdown=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JanEmmisionPercentage\"]")
        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        breakdown.send_keys(breakdownChoice)

        #febuary
        startBreakdown=sheet.cell_value(15, 1)
        endBreakdown=sheet.cell_value(15, 2)
        breakdownChoice = random.randrange(int(startBreakdown), int(endBreakdown)+1, 1)
        breakdown=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FebEmmisionPercentage\"]")
        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        breakdown.send_keys(breakdownChoice)

        #march
        startBreakdown=sheet.cell_value(15, 1)
        endBreakdown=sheet.cell_value(15, 2)
        breakdownChoice = random.randrange(int(startBreakdown), int(endBreakdown)+1, 1)
        breakdown=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MarEmmisionPercentage\"]")
        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        breakdown.send_keys(breakdownChoice)

        #April
        startBreakdown=sheet.cell_value(15, 1)
        endBreakdown=sheet.cell_value(15, 2)
        breakdownChoice = random.randrange(int(startBreakdown), int(endBreakdown)+1, 1)
        breakdown=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AprEmmisionPercentage\"]")
        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        breakdown.send_keys(breakdownChoice)

        #April
        startBreakdown=sheet.cell_value(15, 1)
        endBreakdown=sheet.cell_value(15, 2)
        breakdownChoice = random.randrange(int(startBreakdown), int(endBreakdown)+1, 1)
        breakdown=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MayEmmisionPercentage\"]")
        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        breakdown.send_keys(breakdownChoice)

        #June
        startBreakdown=sheet.cell_value(15, 1)
        endBreakdown=sheet.cell_value(15, 2)
        breakdownChoice = random.randrange(int(startBreakdown), int(endBreakdown)+1, 1)
        breakdown=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JunEmmisionPercentage\"]")
        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        breakdown.send_keys(breakdownChoice)

        #July
        startBreakdown=sheet.cell_value(15, 1)
        endBreakdown=sheet.cell_value(15, 2)
        breakdownChoice = random.randrange(int(startBreakdown), int(endBreakdown)+1, 1)
        breakdown=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JulEmmisionPercentage\"]")
        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        breakdown.send_keys(breakdownChoice)

        #August
        startBreakdown=sheet.cell_value(15, 1)
        endBreakdown=sheet.cell_value(15, 2)
        breakdownChoice = random.randrange(int(startBreakdown), int(endBreakdown)+1, 1)
        breakdown=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AugEmmisionPercentage\"]")
        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        breakdown.send_keys(breakdownChoice)

        #September
        startBreakdown=sheet.cell_value(15, 1)
        endBreakdown=sheet.cell_value(15, 2)
        breakdownChoice = random.randrange(int(startBreakdown), int(endBreakdown)+1, 1)
        breakdown=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"SepEmmisionPercentage\"]")
        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        breakdown.send_keys(breakdownChoice)

        #October
        startBreakdown=sheet.cell_value(15, 1)
        endBreakdown=sheet.cell_value(15, 2)
        breakdownChoice = random.randrange(int(startBreakdown), int(endBreakdown)+1, 1)
        breakdown=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"OctEmmisionPercentage\"]")
        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        breakdown.send_keys(breakdownChoice)

        #november
        startBreakdown=sheet.cell_value(15, 1)
        endBreakdown=sheet.cell_value(15, 2)
        breakdownChoice = random.randrange(int(startBreakdown), int(endBreakdown)+1, 1)
        breakdown=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"NovEmmisionPercentage\"]")
        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        breakdown.send_keys(breakdownChoice)

        #December
        startBreakdown=sheet.cell_value(15, 1)
        endBreakdown=sheet.cell_value(15, 2)
        breakdownChoice = random.randrange(int(startBreakdown), int(endBreakdown)+1, 1)
        breakdown=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"DecEmmisionPercentage\"]")
        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        breakdown.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        breakdown.send_keys(breakdownChoice)



        #setting RF texture
        startRF=sheet.cell_value(17, 1)
        endRF=sheet.cell_value(17, 2)
        RFChoice = round(random.uniform(float(startRF), float(endRF)),1)
        RF=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"RFTexture\"]")
        RF.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        RF.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        RF.send_keys(str(RFChoice))

        #setting RF tillage
        startRFTillage=sheet.cell_value(18, 1)
        endRFTillage=sheet.cell_value(18, 2)
        RFTillageChoice = round(random.uniform(float(startRFTillage), float(endRFTillage)),1)
        RFTillage=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"RFTillage\"]")
        RFTillage.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        RFTillage.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        RFTillage.send_keys(str(RFTillageChoice))

        #setting EF ECO
        startEFEco=sheet.cell_value(19, 1)
        endEFEco=sheet.cell_value(19, 2)
        EFEcoChoice = round(random.uniform(float(startEFEco), float(endEFEco)),4)
        EFEco=driver.find_element_by_xpath("//ComboBox[@Name=\"k\"][@AutomationId=\"EFEco\"]")
        EFEco.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        EFEco.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        EFEco.send_keys(str(EFEcoChoice))


        #setting Leaching Fraction
        startLF=sheet.cell_value(20, 1)
        endLF=sheet.cell_value(20, 2)
        LFChoice = round(random.uniform(float(startLF), float(endLF)),2)
        LF=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FractionalLeaching\"]")
        LF.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        LF.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        LF.send_keys(str(LFChoice))

        #setting EF leaching
        startEFL=sheet.cell_value(21, 1)
        endEFL=sheet.cell_value(21, 2)
        EFLChoice = round(random.uniform(float(startEFL), float(endEFL)),4)
        EFL=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"EFLeaching\"]")
        EFL.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        EFL.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        EFL.send_keys(str(EFLChoice))

        #setting Volatization fraction
        startVF=sheet.cell_value(22, 1)
        endVF=sheet.cell_value(22, 2)
        VFChoice = round(random.uniform(float(startVF), float(endVF)),1)
        VF=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FractionalVolatilization\"]")
        VF.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        VF.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        VF.send_keys(str(VFChoice))

        #setting EF Volatization
        startEVF=sheet.cell_value(23, 1)
        endEVF=sheet.cell_value(23, 2)
        EVFChoice = round(random.uniform(float(startEVF), float(endEVF)),2)
        EVF=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"EFVolatilization\"]")
        EVF.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        EVF.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        EVF.send_keys(str(EVFChoice))

        #setting CO2
        startCO2=sheet.cell_value(24, 1)
        endCO2=sheet.cell_value(24, 2)
        CO2Choice = random.randrange(int(startCO2), int(endCO2)+1, 1)
        CO2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"GWPCO\")]")
        CO2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        CO2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        CO2.send_keys(CO2Choice)

        # Crops -Perennials
        menu = driver.find_element_by_name("Crops")
        menu.click()
        item = menu.find_element_by_name("Perennials")
        actions = ActionChains(driver)
        actions.click(item)
        actions.perform()

        # Detail On
        radioButton = driver.find_element_by_xpath("//RadioButton[@Name=\"On\"]")
        radioButton.click()

        driver.switch_to.window(driver.window_handles[0])

        # set Perreinnal crop

        # (0,30) None
        # (0,50) Hay-grass
        # (0,70) Hay-legum
        # (0,90) Hay-mixed
        # (0,110) hay and forage seed
        # (0,130) others
        cropListPositions=[50,70,90,110,130]
        perennialCropPosi=random.choices(cropListPositions)

        perennialType1=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"PerennialType1\"]")
        perennialType1.click()
        time.sleep(2)

        actions = ActionChains(driver)
        print(perennialCropPosi[0])
        actions.move_by_offset(0, perennialCropPosi[0]).click()
        actions.perform()

        cropListPositions.remove(perennialCropPosi[0])
        perennialCropPosi=random.choices(cropListPositions)

        perennialType2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialType2\")]")
        perennialType2.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, perennialCropPosi[0]).click()
        actions.perform()


        cropListPositions.remove(perennialCropPosi[0])
        perennialCropPosi=random.choices(cropListPositions)

        perennialType3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialType3\")]")
        perennialType3.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, perennialCropPosi[0]).click()
        actions.perform()

        cropListPositions.remove(perennialCropPosi[0])
        perennialCropPosi=random.choices(cropListPositions)
        perennialType4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialType4\")]")
        perennialType4.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, perennialCropPosi[0]).click()
        actions.perform()

        #set perinnalcrop area
        startArea=sheet.cell_value(29, 1)
        endArea=sheet.cell_value(29, 2)
        AreaChoice = random.randrange(int(startArea), int(endArea)+1, 100)
        perennialCropArea1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialArea1\")]")
        perennialCropArea1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropArea1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropArea1.send_keys(AreaChoice)


        AreaChoice = random.randrange(int(startArea), int(endArea)+1, 100)
        perennialCropArea2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialArea2\")]")
        perennialCropArea2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropArea2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropArea2.send_keys(AreaChoice)

        AreaChoice = random.randrange(int(startArea), int(endArea)+1, 100)
        perennialCropArea3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialArea3\")]")
        perennialCropArea3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropArea3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropArea3.send_keys(AreaChoice)

        AreaChoice = random.randrange(int(startArea), int(endArea)+1, 100)
        perennialCropArea4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialArea4\")]")
        perennialCropArea4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropArea4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropArea4.send_keys(AreaChoice)


        #setting year seeded
        startYear=sheet.cell_value(30, 1)
        endYear=sheet.cell_value(30, 2)
        YearChoice = random.randrange(int(startYear), int(endYear)+1, 1)
        perennialCropYear1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialYearSeeded1\")]")
        perennialCropYear1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropYear1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropYear1.send_keys(YearChoice)


        YearChoice = random.randrange(int(startYear), int(endYear)+1, 1)
        perennialCropYear2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialYearSeeded2\")]")
        perennialCropYear2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropYear2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropYear2.send_keys(YearChoice)

        YearChoice = random.randrange(int(startYear), int(endYear)+1, 1)
        perennialCropYear3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialYearSeeded3\")]")
        perennialCropYear3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropYear3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropYear3.send_keys(YearChoice)

        YearChoice = random.randrange(int(startYear), int(endYear)+1, 1)
        perennialCropYear4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialYearSeeded4\")]")
        perennialCropYear4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropYear4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropYear4.send_keys(YearChoice)



        #setting stand length
        startLength=sheet.cell_value(32, 1)
        endLength=sheet.cell_value(32, 2)
        lengthChoice = random.randrange(int(startLength), int(endLength)+1, 1)
        perennialCropStandLength1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialStandLength1\")]")
        perennialCropStandLength1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropStandLength1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropStandLength1.send_keys(lengthChoice)


        lengthChoice = random.randrange(int(startLength), int(endLength)+1, 1)
        perennialCropStandLength2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialStandLength2\")]")
        perennialCropStandLength2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropStandLength2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropStandLength2.send_keys(lengthChoice)

        lengthChoice = random.randrange(int(startLength), int(endLength)+1, 1)
        perennialCropStandLength3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialStandLength3\")]")
        perennialCropStandLength3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropYear3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropStandLength3.send_keys(lengthChoice)

        lengthChoice = random.randrange(int(startLength), int(endLength)+1, 1)
        perennialCropStandLength4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialStandLength4\")]")
        perennialCropStandLength4.click()

        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropStandLength4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropStandLength4.send_keys(lengthChoice)

        #setting yield
        startYield=sheet.cell_value(31, 1)
        endYield=sheet.cell_value(31, 2)
        yieldChoice = random.randrange(int(startYield), int(endYield)+1, 1)
        perennialCropYield1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialYield1\")]")
        perennialCropYield1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropYield1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropYield1.send_keys(yieldChoice)


        yieldChoice = random.randrange(int(startYield), int(endYield)+1, 1)
        perennialCropYield2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialYield2\")]")
        perennialCropYield2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropYield2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropYield2.send_keys(yieldChoice)

        yieldChoice = random.randrange(int(startYield), int(endYield)+1, 1)
        perennialCropYield3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialYield3\")]")
        perennialCropYield3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropYield3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropYield3.send_keys(yieldChoice)

        yieldChoice = random.randrange(int(startYield), int(endYield)+1, 1)
        perennialCropYield4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialYield4\")]")
        perennialCropYield4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropYield4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropYield4.send_keys(yieldChoice)

        #setting irrigated
        perennialCropIrrigated=sheet.cell(33,3)
        irrigatedOptionsList1=['Yes','No']
        irrigatedOptionsList= [str(y) for y in str(perennialCropIrrigated).split(',')]

        irrigatedChoice=random.choices(irrigatedOptionsList1)
        perennialCropIrrigated1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialIrrigated1\")]")
        perennialCropIrrigated1.send_keys(irrigatedChoice)

        irrigatedChoice=random.choices(irrigatedOptionsList1)
        perennialCropIrrigated2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialIrrigated2\")]")
        perennialCropIrrigated2.send_keys(irrigatedChoice)

        irrigatedChoice=random.choices(irrigatedOptionsList1)
        perennialCropIrrigated3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialIrrigated3\")]")
        perennialCropIrrigated3.send_keys(irrigatedChoice)

        irrigatedChoice=random.choices(irrigatedOptionsList1)
        perennialCropIrrigated4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialIrrigated4\")]")
        perennialCropIrrigated4.send_keys(irrigatedChoice)

        #setting Herbicide
        perennialCropHerbicide=sheet.cell(34,3)
        HerbicideOptionsList1=['Yes','No']
        HerbicideOptionsList= [str(y) for y in str(perennialCropHerbicide).split(',')]

        herbicideChoice=random.choices(irrigatedOptionsList1)
        perennialCropHerbicide1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialHerbicide1\")]")
        perennialCropHerbicide1.send_keys(herbicideChoice)

        herbicideChoice=random.choices(irrigatedOptionsList1)
        perennialCropHerbicide2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialHerbicide2\")]")
        perennialCropHerbicide2.send_keys(herbicideChoice)

        herbicideChoice=random.choices(irrigatedOptionsList1)
        perennialCropHerbicide3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialHerbicide3\")]")
        perennialCropHerbicide3.send_keys(herbicideChoice)

        herbicideChoice=random.choices(irrigatedOptionsList1)
        perennialCropHerbicide4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialHerbicide4\")]")
        perennialCropHerbicide4.send_keys(herbicideChoice)

        #setting  N fertilization rate
        startFertilization=sheet.cell_value(35, 1)
        endFertilization=sheet.cell_value(35, 2)
        fertilizationChoice = random.randrange(int(startFertilization), int(endFertilization)+1, 10)
        perennialCropFertilization1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialNFertilizerRate1\")]")
        perennialCropFertilization1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropFertilization1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropFertilization1.send_keys(fertilizationChoice)


        fertilizationChoice = random.randrange(int(startFertilization), int(endFertilization)+1, 10)
        perennialCropFertilization2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialNFertilizerRate2\")]")
        perennialCropFertilization2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropFertilization2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropFertilization2.send_keys(fertilizationChoice)

        fertilizationChoice = random.randrange(int(startFertilization), int(endFertilization)+1, 10)
        perennialCropFertilization3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialNFertilizerRate3\")]")
        perennialCropFertilization3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropFertilization3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropFertilization3.send_keys(fertilizationChoice)

        fertilizationChoice = random.randrange(int(startFertilization), int(endFertilization)+1, 10)
        perennialCropFertilization4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialNFertilizerRate4\")]")
        perennialCropFertilization4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropFertilization4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropFertilization4.send_keys(fertilizationChoice)

        #setting  P fertilization rate
        startPFertilization=sheet.cell_value(36, 1)
        endPFertilization=sheet.cell_value(36, 2)
        fertilizationChoice = random.randrange(int(startPFertilization), int(endPFertilization)+1, 10)
        perennialCropPFertilization1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialPFertilizerRate1\")]")
        perennialCropPFertilization1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropPFertilization1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropPFertilization1.send_keys(fertilizationChoice)


        fertilizationChoice = random.randrange(int(startPFertilization), int(endPFertilization)+1, 10)
        perennialCropPFertilization2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialPFertilizerRate2\")]")
        perennialCropPFertilization2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropPFertilization2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropPFertilization2.send_keys(fertilizationChoice)

        fertilizationChoice = random.randrange(int(startFertilization), int(endFertilization)+1, 10)
        perennialCropPFertilization3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialPFertilizerRate3\")]")
        perennialCropPFertilization3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropPFertilization3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropPFertilization3.send_keys(fertilizationChoice)

        fertilizationChoice = random.randrange(int(startFertilization), int(endFertilization)+1, 10)
        perennialCropPFertilization4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialPFertilizerRate4\")]")
        perennialCropPFertilization4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        perennialCropPFertilization4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        perennialCropPFertilization4.send_keys(fertilizationChoice)

        #setting Moisture Content
        startMoistureC=sheet.cell_value(37, 1)
        endMoistureC=sheet.cell_value(37, 2)
        print(startMoistureC)
        print(endMoistureC)
        moistureCChoice = round(random.uniform(float(startMoistureC), float(endMoistureC)),2)
        moistureContent1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialMoistureContent1\")]")
        moistureContent1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        moistureContent1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        print("ggg",moistureCChoice)
        moistureContent1.send_keys(str(moistureCChoice))

        moistureCChoice = round(random.uniform(float(startMoistureC), float(endMoistureC)),2)
        moistureContent2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialMoistureContent2\")]")
        moistureContent2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        moistureContent2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        moistureContent2.send_keys(str(moistureCChoice))

        moistureCChoice = round(random.uniform(float(startMoistureC), float(endMoistureC)),2)
        moistureContent3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialMoistureContent3\")]")
        moistureContent3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        moistureContent3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        moistureContent3.send_keys(str(moistureCChoice))

        moistureCChoice = round(random.uniform(float(startMoistureC), float(endMoistureC)),2)
        moistureContent4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialMoistureContent4\")]")
        moistureContent4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        moistureContent4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        moistureContent4.send_keys(str(moistureCChoice))


        #setting AGR N Concerntration
        startAGRN=sheet.cell_value(38, 1)
        endAGRN=sheet.cell_value(38, 2)
        AGRNChoice = round(random.uniform(float(startAGRN), float(endAGRN)),4)
        AGRN1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialAGRNConcentration1\")]")
        AGRN1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        AGRN1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        AGRN1.send_keys(str(AGRNChoice))

        AGRNChoice = round(random.uniform(float(startAGRN), float(endAGRN)),4)
        AGRN2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialAGRNConcentration2\")]")
        AGRN2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        AGRN2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        AGRN2.send_keys(str(AGRNChoice))

        AGRNChoice = round(random.uniform(float(startAGRN), float(endAGRN)),4)
        AGRN3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialAGRNConcentration3\")]")
        AGRN3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        AGRN3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        AGRN3.send_keys(str(AGRNChoice))

        AGRNChoice = round(random.uniform(float(startAGRN), float(endAGRN)),4)
        AGRN4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialAGRNConcentration4\")]")
        AGRN4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        AGRN4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        AGRN4.send_keys(str(AGRNChoice))

        #setting BGR N Concerntration
        startBGRN=sheet.cell_value(39, 1)
        endBGRN=sheet.cell_value(39, 2)
        BGRNChoice = round(random.uniform(float(startBGRN), float(endBGRN)),4)
        BGRN1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialBGRNConcentration1\")]")
        BGRN1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        BGRN1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        BGRN1.send_keys(str(BGRNChoice))

        BGRNChoice = round(random.uniform(float(startBGRN), float(endBGRN)),4)
        BGRN2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialBGRNConcentration2\")]")
        BGRN2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        BGRN2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        BGRN2.send_keys(str(BGRNChoice))

        BGRNChoice = round(random.uniform(float(startBGRN), float(endBGRN)),4)
        BGRN3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialBGRNConcentration3\")]")
        BGRN3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        BGRN3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        BGRN3.send_keys(str(BGRNChoice))

        BGRNChoice = round(random.uniform(float(startBGRN), float(endBGRN)),4)
        BGRN4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialBGRNConcentration4\")]")
        BGRN4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        BGRN4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        BGRN4.send_keys(str(BGRNChoice))


        #Setting YieldRatio
        startYieldRatio=sheet.cell_value(40, 1)
        endYieldRatio=sheet.cell_value(40, 2)
        yieldRatioChoice = round(random.uniform(float(startYieldRatio), float(endYieldRatio)),2)
        yieldRatio1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialYieldRatio1\")]")
        yieldRatio1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        yieldRatio1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        yieldRatio1.send_keys(str(yieldRatioChoice))

        yieldRatioChoice = round(random.uniform(float(startYieldRatio), float(endYieldRatio)),2)
        yieldRatio2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialYieldRatio2\")]")
        yieldRatio2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        yieldRatio2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        yieldRatio2.send_keys(str(yieldRatioChoice))

        yieldRatioChoice = round(random.uniform(float(startYieldRatio), float(endYieldRatio)),2)
        yieldRatio3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialYieldRatio3\")]")
        yieldRatio3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        yieldRatio3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        yieldRatio3.send_keys(str(yieldRatioChoice))

        yieldRatioChoice = round(random.uniform(float(startYieldRatio), float(endYieldRatio)),2)
        yieldRatio4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialYieldRatio4\")]")
        yieldRatio4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        yieldRatio4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        yieldRatio4.send_keys(str(yieldRatioChoice))

        #Setting AGR Ratio
        startAGRRatio=sheet.cell_value(41, 1)
        endAGRRatio=sheet.cell_value(41, 2)
        AGRRatioChoice = round(random.uniform(float(startAGRRatio), float(endAGRRatio)),2)
        AGRRatio1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialAGRRatio1\")]")
        AGRRatio1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        AGRRatio1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        AGRRatio1.send_keys(str(AGRRatioChoice))


        AGRRatioChoice = round(random.uniform(float(startAGRRatio), float(endAGRRatio)),2)
        AGRRatio2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialAGRRatio2\")]")
        AGRRatio2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        AGRRatio2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        AGRRatio2.send_keys(str(AGRRatioChoice))


        AGRRatioChoice = round(random.uniform(float(startAGRRatio), float(endAGRRatio)),2)
        AGRRatio3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialAGRRatio3\")]")
        AGRRatio3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        AGRRatio3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        AGRRatio3.send_keys(str(AGRRatioChoice))


        AGRRatioChoice = round(random.uniform(float(startAGRRatio), float(endAGRRatio)),2)
        AGRRatio4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialAGRRatio4\")]")
        AGRRatio4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        AGRRatio4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        AGRRatio4.send_keys(str(AGRRatioChoice))

        #Stteing BGR Ratio
        startBGRRatio=sheet.cell_value(42, 1)
        endBGRRatio=sheet.cell_value(42, 2)
        BGRRatioChoice = round(random.uniform(float(startBGRRatio), float(endBGRRatio)),2)
        BGRRatio1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialBGRRatio1\")]")
        BGRRatio1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        BGRRatio1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        BGRRatio1.send_keys(str(BGRRatioChoice))

        BGRRatioChoice = round(random.uniform(float(startBGRRatio), float(endBGRRatio)),2)
        BGRRatio2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialBGRRatio2\")]")
        BGRRatio2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        BGRRatio2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        BGRRatio2.send_keys(str(BGRRatioChoice))

        BGRRatioChoice = round(random.uniform(float(startBGRRatio), float(endBGRRatio)),2)
        BGRRatio3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialBGRRatio3\")]")
        BGRRatio3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        BGRRatio3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        BGRRatio3.send_keys(str(BGRRatioChoice))


        BGRRatioChoice = round(random.uniform(float(startBGRRatio), float(endBGRRatio)),2)
        BGRRatio4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialBGRRatio4\")]")
        BGRRatio4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        BGRRatio4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        BGRRatio4.send_keys(str(BGRRatioChoice))

        # Setting Fuel Energy
        startFuelEnergy=sheet.cell_value(43, 1)
        endFuelEnergy=sheet.cell_value(43, 2)
        fuelEnergyChoice = round(random.uniform(float(startFuelEnergy), float(endFuelEnergy)),2)
        fuelEnergy1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialEFuel1\")]")
        fuelEnergy1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        fuelEnergy1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        fuelEnergy1.send_keys(str(fuelEnergyChoice))

        fuelEnergyChoice = round(random.uniform(float(startFuelEnergy), float(endFuelEnergy)),2)
        fuelEnergy2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialEFuel2\")]")
        fuelEnergy2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        fuelEnergy2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        fuelEnergy2.send_keys(str(fuelEnergyChoice))

        fuelEnergyChoice = round(random.uniform(float(startFuelEnergy), float(endFuelEnergy)),2)
        fuelEnergy3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialEFuel3\")]")
        fuelEnergy3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        fuelEnergy3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        fuelEnergy3.send_keys(str(fuelEnergyChoice))

        fuelEnergyChoice = round(random.uniform(float(startFuelEnergy), float(endFuelEnergy)),2)
        fuelEnergy4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialEFuel4\")]")
        fuelEnergy4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        fuelEnergy4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        fuelEnergy4.send_keys(str(fuelEnergyChoice))

        #setting Herbicide energy
        startHerbicideEnergy=sheet.cell_value(43, 1)
        endHerbicideEnergy=sheet.cell_value(43, 2)
        herbicideEnergyChoice = round(random.uniform(float(startHerbicideEnergy), float(endHerbicideEnergy)),2)
        herbicideEnergy1=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialEHerbicide1\")]")
        herbicideEnergy1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        herbicideEnergy1.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        herbicideEnergy1.send_keys(str(herbicideEnergyChoice))

        herbicideEnergyChoice = round(random.uniform(float(startHerbicideEnergy), float(endHerbicideEnergy)),2)
        herbicideEnergy2=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialEHerbicide2\")]")
        herbicideEnergy2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        herbicideEnergy2.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        herbicideEnergy2.send_keys(str(herbicideEnergyChoice))

        herbicideEnergyChoice = round(random.uniform(float(startHerbicideEnergy), float(endHerbicideEnergy)),2)
        herbicideEnergy3=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialEHerbicide3\")]")
        herbicideEnergy3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        herbicideEnergy3.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        herbicideEnergy3.send_keys(str(herbicideEnergyChoice))

        herbicideEnergyChoice = round(random.uniform(float(startHerbicideEnergy), float(endHerbicideEnergy)),2)
        herbicideEnergy4=driver.find_element_by_xpath("//ComboBox[starts-with(@AutomationId,\"PerennialEHerbicide4\")]")
        herbicideEnergy4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        herbicideEnergy4.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        herbicideEnergy4.send_keys(str(herbicideEnergyChoice))

        #Setting LumCMax
        startLumCMAX=sheet.cell_value(45, 1)
        endLumCMAX=sheet.cell_value(45, 2)
        luCMAXChoice = random.randrange(int(startLumCMAX), int(endLumCMAX)+1, 1)
        luCMAX=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"PerennialPastLUMCMax\"]")
        luCMAX.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        luCMAX.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        luCMAX.send_keys(luCMAXChoice)

        #setting K
        startK=sheet.cell_value(43, 1)
        endK=sheet.cell_value(43, 2)
        kChoice = round(random.uniform(float(startK), float(endK)),2)
        k=driver.find_element_by_xpath("//ComboBox[@AutomationId=\"PerennialPastK\"]")
        k.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        k.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10,10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        k.send_keys(str(kChoice))


        # beef Finisher group 1
        menu = driver.find_element_by_name("Beef")
        menu.click()
        item = menu.find_element_by_name("Finishing Group 1")
        actions = ActionChains(driver)
        actions.click(item)
        actions.perform()

        # Detail On
        radioButton = driver.find_element_by_xpath("//RadioButton[@Name=\"On\"]")
        radioButton.click()

        driver.switch_to.window(driver.window_handles[0])
        #
        # # Setting Heifers
        # # January
        # startHeifers = sheet.cell_value(54, 1)
        # endHeifers = sheet.cell_value(54, 2)
        # heifersChoice = random.randrange(int(startHeifers), int(endHeifers) + 1, 1)
        # heifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JanNumberOfHeifers\"]")
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heifers.send_keys(heifersChoice)
        #
        # # Febuary
        # heifersChoice = random.randrange(int(startHeifers), int(endHeifers) + 1, 1)
        # heifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FebNumberOfHeifers\"]")
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heifers.send_keys(heifersChoice)
        #
        # # March
        # heifersChoice = random.randrange(int(startHeifers), int(endHeifers) + 1, 1)
        # heifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MarNumberOfHeifers\"]")
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heifers.send_keys(heifersChoice)
        #
        # # April
        # heifersChoice = random.randrange(int(startHeifers), int(endHeifers) + 1, 1)
        # heifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AprNumberOfHeifers\"]")
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heifers.send_keys(heifersChoice)
        #
        # # May
        # heifersChoice = random.randrange(int(startHeifers), int(endHeifers) + 1, 1)
        # heifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MayNumberOfHeifers\"]")
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heifers.send_keys(heifersChoice)
        #
        # # June
        # heifersChoice = random.randrange(int(startHeifers), int(endHeifers) + 1, 1)
        # heifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JunNumberOfHeifers\"]")
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heifers.send_keys(heifersChoice)
        #
        # # July
        # heifersChoice = random.randrange(int(startHeifers), int(endHeifers) + 1, 1)
        # heifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JulNumberOfHeifers\"]")
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heifers.send_keys(heifersChoice)
        #
        # # August
        # heifersChoice = random.randrange(int(startHeifers), int(endHeifers) + 1, 1)
        # heifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AugNumberOfHeifers\"]")
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heifers.send_keys(heifersChoice)
        #
        # # September
        # heifersChoice = random.randrange(int(startHeifers), int(endHeifers) + 1, 1)
        # heifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"SepNumberOfHeifers\"]")
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heifers.send_keys(heifersChoice)
        #
        # # October
        # heifersChoice = random.randrange(int(startHeifers), int(endHeifers) + 1, 1)
        # heifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"OctNumberOfHeifers\"]")
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heifers.send_keys(heifersChoice)
        #
        # # November
        # heifersChoice = random.randrange(int(startHeifers), int(endHeifers) + 1, 1)
        # heifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"NovNumberOfHeifers\"]")
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heifers.send_keys(heifersChoice)
        #
        # # December
        # heifersChoice = random.randrange(int(startHeifers), int(endHeifers) + 1, 1)
        # heifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"DecNumberOfHeifers\"]")
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heifers.send_keys(heifersChoice)
        #
        # # setting Initial Heifer weight
        # # January
        # startHeifersInit = sheet.cell_value(55, 1)
        # endHeifersInit = sheet.cell_value(55, 2)
        # initHeifersChoice = random.randrange(int(startHeifersInit), int(endHeifersInit) + 1, 1)
        # initHeifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JanHeiferInitialWeight\"]")
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # initHeifers.send_keys(initHeifersChoice)
        #
        # # Febuary
        # initHeifersChoice = random.randrange(int(startHeifersInit), int(endHeifersInit) + 1, 1)
        # initHeifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FebHeiferInitialWeight\"]")
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # initHeifers.send_keys(initHeifersChoice)
        #
        # # March
        # initHeifersChoice = random.randrange(int(startHeifersInit), int(endHeifersInit) + 1, 1)
        # initHeifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MarHeiferInitialWeight\"]")
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # initHeifers.send_keys(initHeifersChoice)
        #
        # # April
        # initHeifersChoice = random.randrange(int(startHeifersInit), int(endHeifersInit) + 1, 1)
        # initHeifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AprHeiferInitialWeight\"]")
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # initHeifers.send_keys(initHeifersChoice)
        #
        # # May
        # initHeifersChoice = random.randrange(int(startHeifersInit), int(endHeifersInit) + 1, 1)
        # initHeifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MayHeiferInitialWeight\"]")
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # initHeifers.send_keys(initHeifersChoice)
        #
        # # June
        # initHeifersChoice = random.randrange(int(startHeifersInit), int(endHeifersInit) + 1, 1)
        # initHeifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JunHeiferInitialWeight\"]")
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # initHeifers.send_keys(initHeifersChoice)
        #
        # # July
        # initHeifersChoice = random.randrange(int(startHeifersInit), int(endHeifersInit) + 1, 1)
        # initHeifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JulHeiferInitialWeight\"]")
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # initHeifers.send_keys(initHeifersChoice)
        #
        # # August
        # initHeifersChoice = random.randrange(int(startHeifersInit), int(endHeifersInit) + 1, 1)
        # initHeifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AugHeiferInitialWeight\"]")
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # initHeifers.send_keys(initHeifersChoice)
        #
        # # September
        # initHeifersChoice = random.randrange(int(startHeifersInit), int(endHeifersInit) + 1, 1)
        # initHeifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"SepHeiferInitialWeight\"]")
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # initHeifers.send_keys(initHeifersChoice)
        #
        # # October
        # initHeifersChoice = random.randrange(int(startHeifersInit), int(endHeifersInit) + 1, 1)
        # initHeifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"OctHeiferInitialWeight\"]")
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # initHeifers.send_keys(initHeifersChoice)
        #
        # # November
        # initHeifersChoice = random.randrange(int(startHeifersInit), int(endHeifersInit) + 1, 1)
        # initHeifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"NovHeiferInitialWeight\"]")
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # initHeifers.send_keys(initHeifersChoice)
        #
        # # December
        # initHeifersChoice = random.randrange(int(startHeifersInit), int(endHeifersInit) + 1, 1)
        # initHeifers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"DecHeiferInitialWeight\"]")
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # initHeifers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # initHeifers.send_keys(initHeifersChoice)
        #
        # # Heifer ADG
        # # January
        # startHeiferADG = sheet.cell_value(57, 1)
        # endHeiferADG = sheet.cell_value(57, 2)
        # heiferADGChoice = round(random.uniform(float(startHeiferADG), float(endHeiferADG)), 2)
        # heiferADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JanHeiferADG\"]")
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heiferADG.send_keys(str(heiferADGChoice))
        #
        # # Febuary
        # heiferADGChoice = round(random.uniform(float(startHeiferADG), float(endHeiferADG)), 2)
        # heiferADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FebHeiferADG\"]")
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heiferADG.send_keys(str(heiferADGChoice))
        #
        # # March
        # heiferADGChoice = round(random.uniform(float(startHeiferADG), float(endHeiferADG)), 2)
        # heiferADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FebHeiferADG\"]")
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heiferADG.send_keys(str(heiferADGChoice))
        #
        # # April
        # heiferADGChoice = round(random.uniform(float(startHeiferADG), float(endHeiferADG)), 2)
        # heiferADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AprHeiferADG\"]")
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heiferADG.send_keys(str(heiferADGChoice))
        #
        # # May
        # heiferADGChoice = round(random.uniform(float(startHeiferADG), float(endHeiferADG)), 2)
        # heiferADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MayHeiferADG\"]")
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heiferADG.send_keys(str(heiferADGChoice))
        #
        # # June
        # heiferADGChoice = round(random.uniform(float(startHeiferADG), float(endHeiferADG)), 2)
        # heiferADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JunHeiferADG\"]")
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heiferADG.send_keys(str(heiferADGChoice))
        #
        # # July
        # heiferADGChoice = round(random.uniform(float(startHeiferADG), float(endHeiferADG)), 2)
        # heiferADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JulHeiferADG\"]")
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heiferADG.send_keys(str(heiferADGChoice))
        #
        # # August
        # heiferADGChoice = round(random.uniform(float(startHeiferADG), float(endHeiferADG)), 2)
        # heiferADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AugHeiferADG\"]")
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heiferADG.send_keys(str(heiferADGChoice))
        #
        # # September
        # heiferADGChoice = round(random.uniform(float(startHeiferADG), float(endHeiferADG)), 2)
        # heiferADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"SepHeiferADG\"]")
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heiferADG.send_keys(str(heiferADGChoice))
        #
        # # October
        # heiferADGChoice = round(random.uniform(float(startHeiferADG), float(endHeiferADG)), 2)
        # heiferADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"OctHeiferADG\"]")
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heiferADG.send_keys(str(heiferADGChoice))
        #
        # # November
        # heiferADGChoice = round(random.uniform(float(startHeiferADG), float(endHeiferADG)), 2)
        # heiferADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"NovHeiferADG\"]")
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heiferADG.send_keys(str(heiferADGChoice))
        #
        # # December
        # heiferADGChoice = round(random.uniform(float(startHeiferADG), float(endHeiferADG)), 2)
        # heiferADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"DecHeiferADG\"]")
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # heiferADG.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # heiferADG.send_keys(str(heiferADGChoice))

        # # setting #steers
        # startSteers = sheet.cell_value(58, 1)
        # endSteers = sheet.cell_value(58, 2)
        #
        # # January
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JanNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #
        # # Febuary
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FebNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #
        # # March
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FebNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #
        # # April
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AprNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #
        # # March
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MarNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #
        # # April
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AprNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #
        # # May
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MayNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #
        # # June
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JunNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #
        # # July
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JulNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #
        # # August
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AugNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #
        # # September
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"SepNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #
        # # October
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"OctNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #
        # # November
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"NovNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #
        # # December
        # steersChoice = random.randrange(int(startSteers), int(endSteers) + 1, 1)
        # steers = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"DecNumberOfSteers\"]")
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Select All").click()
        #
        # steers.click()
        # actions = ActionChains(driver)
        # actions.move_by_offset(10, 10)
        # actions.context_click()
        # actions.perform()
        # rm.find_element_by_name("Cut").click()
        # steers.send_keys(steersChoice)
        #

        # setting Steer Initial Weigth
        startSteersInit = sheet.cell_value(59, 1)
        endSteersInit = sheet.cell_value(59, 2)

        # January
        steersInitChoice = random.randrange(int(startSteersInit), int(endSteersInit) + 1, 1)
        steersInit = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JanSteerInitialWeight\"]")
        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersInit.send_keys(steersInitChoice)

        # Febuary
        steersInitChoice = random.randrange(int(startSteersInit), int(endSteersInit) + 1, 1)
        steersInit = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FebSteerInitialWeight\"]")
        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersInit.send_keys(steersInitChoice)

        # March
        steersInitChoice = random.randrange(int(startSteersInit), int(endSteersInit) + 1, 1)
        steersInit = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MarSteerInitialWeight\"]")
        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersInit.send_keys(steersInitChoice)

        # April
        steersInitChoice = random.randrange(int(startSteersInit), int(endSteersInit) + 1, 1)
        steersInit = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AprSteerInitialWeight\"]")
        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersInit.send_keys(steersInitChoice)

        # May
        steersInitChoice = random.randrange(int(startSteersInit), int(endSteersInit) + 1, 1)
        steersInit = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MaySteerInitialWeight\"]")
        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersInit.send_keys(steersInitChoice)

        # June
        steersInitChoice = random.randrange(int(startSteersInit), int(endSteersInit) + 1, 1)
        steersInit = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JunSteerInitialWeight\"]")
        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersInit.send_keys(steersInitChoice)

        # July
        steersInitChoice = random.randrange(int(startSteersInit), int(endSteersInit) + 1, 1)
        steersInit = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JulSteerInitialWeight\"]")
        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersInit.send_keys(steersInitChoice)

        # August
        steersInitChoice = random.randrange(int(startSteersInit), int(endSteersInit) + 1, 1)
        steersInit = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AugSteerInitialWeight\"]")
        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersInit.send_keys(steersInitChoice)

        # September
        steersInitChoice = random.randrange(int(startSteersInit), int(endSteersInit) + 1, 1)
        steersInit = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"SepSteerInitialWeight\"]")
        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersInit.send_keys(steersInitChoice)

        # October
        steersInitChoice = random.randrange(int(startSteersInit), int(endSteersInit) + 1, 1)
        steersInit = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"OctSteerInitialWeight\"]")
        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersInit.send_keys(steersInitChoice)

        # November
        steersInitChoice = random.randrange(int(startSteersInit), int(endSteersInit) + 1, 1)
        steersInit = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"NovSteerInitialWeight\"]")
        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersInit.send_keys(steersInitChoice)

        # December
        steersInitChoice = random.randrange(int(startSteersInit), int(endSteersInit) + 1, 1)
        steersInit = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"DecSteerInitialWeight\"]")
        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersInit.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersInit.send_keys(steersInitChoice)

        # setting Steer ADG
        startSteersADG = sheet.cell_value(61, 1)
        endSteersADG = sheet.cell_value(61, 2)

        # January
        steersADGChoice = round(random.uniform(float(startSteersADG), float(endSteersADG)), 2)
        steersADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JanSteerADG\"]")
        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersADG.send_keys(str(steersADGChoice))

        # Febuary
        steersADGChoice = round(random.uniform(float(startSteersADG), float(endSteersADG)), 2)
        steersADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FebSteerADG\"]")
        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersADG.send_keys(str(steersADGChoice))

        # March
        steersADGChoice = round(random.uniform(float(startSteersADG), float(endSteersADG)), 2)
        steersADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MarSteerADG\"]")
        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersADG.send_keys(str(steersADGChoice))

        # April
        steersADGChoice = round(random.uniform(float(startSteersADG), float(endSteersADG)), 2)
        steersADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AprSteerADG\"]")
        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersADG.send_keys(str(steersADGChoice))

        # May
        steersADGChoice = round(random.uniform(float(startSteersADG), float(endSteersADG)), 2)
        steersADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MaySteerADG\"]")
        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersADG.send_keys(str(steersADGChoice))

        # June
        steersADGChoice = round(random.uniform(float(startSteersADG), float(endSteersADG)), 2)
        steersADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JunSteerADG\"]")
        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersADG.send_keys(str(steersADGChoice))

        # July
        steersADGChoice = round(random.uniform(float(startSteersADG), float(endSteersADG)), 2)
        steersADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JulSteerADG\"]")
        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersADG.send_keys(str(steersADGChoice))

        # August
        steersADGChoice = round(random.uniform(float(startSteersADG), float(endSteersADG)), 2)
        steersADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AugSteerADG\"]")
        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersADG.send_keys(str(steersADGChoice))

        # September
        steersADGChoice = round(random.uniform(float(startSteersADG), float(endSteersADG)), 2)
        steersADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"SepSteerADG\"]")
        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersADG.send_keys(str(steersADGChoice))

        # October
        steersADGChoice = round(random.uniform(float(startSteersADG), float(endSteersADG)), 2)
        steersADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"OctSteerADG\"]")
        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersADG.send_keys(str(steersADGChoice))

        # November
        steersADGChoice = round(random.uniform(float(startSteersADG), float(endSteersADG)), 2)
        steersADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"NovSteerADG\"]")
        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersADG.send_keys(str(steersADGChoice))

        # December
        steersADGChoice = round(random.uniform(float(startSteersADG), float(endSteersADG)), 2)
        steersADG = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"DecSteerADG\"]")
        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Select All").click()

        steersADG.click()
        actions = ActionChains(driver)
        actions.move_by_offset(10, 10)
        actions.context_click()
        actions.perform()
        rm.find_element_by_name("Cut").click()
        steersADG.send_keys(str(steersADGChoice))

        # Setting Housing
        # (0,30) Confined No Barn
        # (0,50) Housed In Barn
        # (0,70) Enclosed Pasture
        # (0,90) Open Range
        # (0,110) Custom
        housingListPositions = [30, 50, 70, 90, 110]

        # January
        housingPosi = random.choices(housingListPositions)
        housing = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JanHousing\"]")
        housing.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, housingPosi[0]).click()
        actions.perform()

        # if housingPosi[0] == 110:
        #     # setting CF Temp
        #     startCFTemp = sheet.cell_value(66, 1)
        #     endCFTemp = sheet.cell_value(66, 2)
        #
        #     # January
        #     CFTempChoice = round(random.uniform(float(startCFTemp), float(endCFTemp)), 3)
        #     CFTemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JanCF\"]")
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CFTemp.send_keys(str(CFTempChoice))
        #
        #     # setting CA Temp
        #     startCATemp = sheet.cell_value(66, 1)
        #     endCATemp = sheet.cell_value(66, 2)
        #
        #     # January
        #     CATempChoice = round(random.uniform(float(startCATemp), float(endCATemp)), 2)
        #     CATemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JanCA\"]")
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CATemp.send_keys(str(CATempChoice))

        # Febuary
        housingPosi = random.choices(housingListPositions)
        housing = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FebHousing\"]")
        housing.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, housingPosi[0]).click()
        actions.perform()

        # if housingPosi[0] == 110:
        #     # setting CF Temp
        #     startCFTemp = sheet.cell_value(66, 1)
        #     endCFTemp = sheet.cell_value(66, 2)
        #     CFTempChoice = round(random.uniform(float(startCFTemp), float(endCFTemp)), 3)
        #     CFTemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FebCF\"]")
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CFTemp.send_keys(str(CFTempChoice))
        #
        #     # setting CA Temp
        #     startCATemp = sheet.cell_value(66, 1)
        #     endCATemp = sheet.cell_value(66, 2)
        #     CATempChoice = round(random.uniform(float(startCATemp), float(endCATemp)), 2)
        #     CATemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"FebCA\"]")
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CATemp.send_keys(str(CATempChoice))

        # March
        housingPosi = random.choices(housingListPositions)
        housing = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MarHousing\"]")
        housing.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, housingPosi[0]).click()
        actions.perform()

        # if housingPosi[0] == 110:
        #     # setting CF Temp
        #     startCFTemp = sheet.cell_value(66, 1)
        #     endCFTemp = sheet.cell_value(66, 2)
        #     CFTempChoice = round(random.uniform(float(startCFTemp), float(endCFTemp)), 3)
        #     CFTemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MarCF\"]")
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CFTemp.send_keys(str(CFTempChoice))
        #
        #     # setting CA Temp
        #     startCATemp = sheet.cell_value(66, 1)
        #     endCATemp = sheet.cell_value(66, 2)
        #     CATempChoice = round(random.uniform(float(startCATemp), float(endCATemp)), 2)
        #     CATemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MarCA\"]")
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CATemp.send_keys(str(CATempChoice))

        # April
        housingPosi = random.choices(housingListPositions)
        housing = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AprHousing\"]")
        housing.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, housingPosi[0]).click()
        actions.perform()

        # if housingPosi[0] == 110:
        #     # setting CF Temp
        #     startCFTemp = sheet.cell_value(66, 1)
        #     endCFTemp = sheet.cell_value(66, 2)
        #     CFTempChoice = round(random.uniform(float(startCFTemp), float(endCFTemp)), 3)
        #     CFTemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AprCF\"]")
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CFTemp.send_keys(str(CFTempChoice))
        #
        #     # setting CA Temp
        #     startCATemp = sheet.cell_value(66, 1)
        #     endCATemp = sheet.cell_value(66, 2)
        #     CATempChoice = round(random.uniform(float(startCATemp), float(endCATemp)), 2)
        #     CATemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AprCA\"]")
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CATemp.send_keys(str(CATempChoice))

        # May
        housingPosi = random.choices(housingListPositions)
        housing = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MayHousing\"]")
        housing.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, housingPosi[0]).click()
        actions.perform()

        # if housingPosi[0] == 110:
        #     # setting CF Temp
        #     startCFTemp = sheet.cell_value(66, 1)
        #     endCFTemp = sheet.cell_value(66, 2)
        #     CFTempChoice = round(random.uniform(float(startCFTemp), float(endCFTemp)), 3)
        #     CFTemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MayCF\"]")
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CFTemp.send_keys(str(CFTempChoice))
        #
        #     # setting CA Temp
        #     startCATemp = sheet.cell_value(66, 1)
        #     endCATemp = sheet.cell_value(66, 2)
        #     CATempChoice = round(random.uniform(float(startCATemp), float(endCATemp)), 2)
        #     CATemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"MayCA\"]")
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CATemp.send_keys(str(CATempChoice))

        # June
        housingPosi = random.choices(housingListPositions)
        housing = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JunHousing\"]")
        housing.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, housingPosi[0]).click()
        actions.perform()

        # if housingPosi[0] == 110:
        #     # setting CF Temp
        #     startCFTemp = sheet.cell_value(66, 1)
        #     endCFTemp = sheet.cell_value(66, 2)
        #     CFTempChoice = round(random.uniform(float(startCFTemp), float(endCFTemp)), 3)
        #     CFTemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JunCF\"]")
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CFTemp.send_keys(str(CFTempChoice))
        #
        #     # setting CA Temp
        #     startCATemp = sheet.cell_value(66, 1)
        #     endCATemp = sheet.cell_value(66, 2)
        #     CATempChoice = round(random.uniform(float(startCATemp), float(endCATemp)), 2)
        #     CATemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JunCA\"]")
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CATemp.send_keys(str(CATempChoice))

        # July
        housingPosi = random.choices(housingListPositions)
        housing = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JulHousing\"]")
        housing.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, housingPosi[0]).click()
        actions.perform()

        # if housingPosi[0] == 110:
        #     # setting CF Temp
        #     startCFTemp = sheet.cell_value(66, 1)
        #     endCFTemp = sheet.cell_value(66, 2)
        #     CFTempChoice = round(random.uniform(float(startCFTemp), float(endCFTemp)), 3)
        #     CFTemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JulCF\"]")
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CFTemp.send_keys(str(CFTempChoice))
        #
        #     # setting CA Temp
        #     startCATemp = sheet.cell_value(66, 1)
        #     endCATemp = sheet.cell_value(66, 2)
        #     CATempChoice = round(random.uniform(float(startCATemp), float(endCATemp)), 2)
        #     CATemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"JulCA\"]")
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CATemp.send_keys(str(CATempChoice))

        # August
        housingPosi = random.choices(housingListPositions)
        housing = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AugHousing\"]")
        housing.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, housingPosi[0]).click()
        actions.perform()

        # if housingPosi[0] == 110:
        #     # setting CF Temp
        #     startCFTemp = sheet.cell_value(66, 1)
        #     endCFTemp = sheet.cell_value(66, 2)
        #     CFTempChoice = round(random.uniform(float(startCFTemp), float(endCFTemp)), 3)
        #     CFTemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AugCF\"]")
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CFTemp.send_keys(str(CFTempChoice))
        #
        #     # setting CA Temp
        #     startCATemp = sheet.cell_value(66, 1)
        #     endCATemp = sheet.cell_value(66, 2)
        #     CATempChoice = round(random.uniform(float(startCATemp), float(endCATemp)), 2)
        #     CATemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"AugCA\"]")
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CATemp.send_keys(str(CATempChoice))

        # September
        housingPosi = random.choices(housingListPositions)
        housing = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"SepHousing\"]")
        housing.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, housingPosi[0]).click()
        actions.perform()

        # if housingPosi[0] == 110:
        #     # setting CF Temp
        #     startCFTemp = sheet.cell_value(66, 1)
        #     endCFTemp = sheet.cell_value(66, 2)
        #     CFTempChoice = round(random.uniform(float(startCFTemp), float(endCFTemp)), 3)
        #     CFTemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"SepCF\"]")
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CFTemp.send_keys(str(CFTempChoice))
        #
        #     # setting CA Temp
        #     startCATemp = sheet.cell_value(66, 1)
        #     endCATemp = sheet.cell_value(66, 2)
        #     CATempChoice = round(random.uniform(float(startCATemp), float(endCATemp)), 2)
        #     CATemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"SepCA\"]")
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CATemp.send_keys(str(CATempChoice))

        # October
        housingPosi = random.choices(housingListPositions)
        housing = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"OctHousing\"]")
        housing.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, housingPosi[0]).click()
        actions.perform()

        # if housingPosi[0] == 110:
        #     # setting CF Temp
        #     startCFTemp = sheet.cell_value(66, 1)
        #     endCFTemp = sheet.cell_value(66, 2)
        #     CFTempChoice = round(random.uniform(float(startCFTemp), float(endCFTemp)), 3)
        #     CFTemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"OctCF\"]")
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CFTemp.send_keys(str(CFTempChoice))
        #
        #     # setting CA Temp
        #     startCATemp = sheet.cell_value(66, 1)
        #     endCATemp = sheet.cell_value(66, 2)
        #     CATempChoice = round(random.uniform(float(startCATemp), float(endCATemp)), 2)
        #     CATemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"OctCA\"]")
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CATemp.send_keys(str(CATempChoice))

        # November
        housingPosi = random.choices(housingListPositions)
        housing = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"NovHousing\"]")
        housing.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, housingPosi[0]).click()
        actions.perform()

        # if housingPosi[0] == 110:
        #     # setting CF Temp
        #     startCFTemp = sheet.cell_value(66, 1)
        #     endCFTemp = sheet.cell_value(66, 2)
        #     CFTempChoice = round(random.uniform(float(startCFTemp), float(endCFTemp)), 3)
        #     CFTemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"NovCF\"]")
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CFTemp.send_keys(str(CFTempChoice))
        #
        #     # setting CA Temp
        #     startCATemp = sheet.cell_value(66, 1)
        #     endCATemp = sheet.cell_value(66, 2)
        #     CATempChoice = round(random.uniform(float(startCATemp), float(endCATemp)), 2)
        #     CATemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"NovCA\"]")
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CATemp.send_keys(str(CATempChoice))

        # Decemeber
        housingPosi = random.choices(housingListPositions)
        housing = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"DecHousing\"]")
        housing.click()
        time.sleep(2)

        actions = ActionChains(driver)
        actions.move_by_offset(0, housingPosi[0]).click()
        actions.perform()

        # if housingPosi[0] == 110:
        #     # setting CF Temp
        #     startCFTemp = sheet.cell_value(66, 1)
        #     endCFTemp = sheet.cell_value(66, 2)
        #     CFTempChoice = round(random.uniform(float(startCFTemp), float(endCFTemp)), 3)
        #     CFTemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"DecCF\"]")
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CFTemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CFTemp.send_keys(str(CFTempChoice))
        #
        #     # setting CA Temp
        #     startCATemp = sheet.cell_value(66, 1)
        #     endCATemp = sheet.cell_value(66, 2)
        #     CATempChoice = round(random.uniform(float(startCATemp), float(endCATemp)), 2)
        #     CATemp = driver.find_element_by_xpath("//ComboBox[@AutomationId=\"DecCA\"]")
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Select All").click()
        #
        #     CATemp.click()
        #     actions = ActionChains(driver)
        #     actions.move_by_offset(10, 10)
        #     actions.context_click()
        #     actions.perform()
        #     rm.find_element_by_name("Cut").click()
        #     CATemp.send_keys(str(CATempChoice))


        # setting Diet
        # (0,30) Barley
        # (0,50) Corn
        # (0,70) Custom

        def settingDiet(dietPath, TDNPath, CPPath, YMPath):
            dietListPositions = [30, 50, 70]
            dietPosi = random.choices(dietListPositions)
            diet = driver.find_element_by_xpath(dietPath)
            diet.click()
            time.sleep(2)
            actions = ActionChains(driver)
            actions.move_by_offset(0, dietPosi[0]).click()
            actions.perform()

            # if dietPosi[0] == 70:
            #     # setting TDN
            #     startTDN = sheet.cell_value(68, 1)
            #     endTDN = sheet.cell_value(68, 2)
            #
            #     TDNChoice = random.randrange(int(startTDN), int(endTDN))
            #     TDN = driver.find_element_by_xpath(TDNPath)
            #     TDN.click()
            #     actions = ActionChains(driver)
            #     actions.move_by_offset(10, 10)
            #     actions.context_click()
            #     actions.perform()
            #     rm.find_element_by_name("Select All").click()
            #
            #     TDN.click()
            #     actions = ActionChains(driver)
            #     actions.move_by_offset(10, 10)
            #     actions.context_click()
            #     actions.perform()
            #     rm.find_element_by_name("Cut").click()
            #     TDN.send_keys(TDNChoice)
            #
            #     # setting CP
            #     startCP = sheet.cell_value(69, 1)
            #     endCP = sheet.cell_value(69, 2)
            #     CPChoice = round(random.uniform(float(startCP), float(endCP)), 3)
            #     CP = driver.find_element_by_xpath(CPPath)
            #     CP.click()
            #     actions = ActionChains(driver)
            #     actions.move_by_offset(10, 10)
            #     actions.context_click()
            #     actions.perform()
            #     rm.find_element_by_name("Select All").click()
            #
            #     CP.click()
            #     actions = ActionChains(driver)
            #     actions.move_by_offset(10, 10)
            #     actions.context_click()
            #     actions.perform()
            #     rm.find_element_by_name("Cut").click()
            #     CP.send_keys(str(CPChoice))
            #
            #     # setting YM Unadjysted
            #     startYMUnadjusted = sheet.cell_value(70, 1)
            #     endYMUnadjusted = sheet.cell_value(70, 2)
            #     YMUnadjustedChoice = round(random.uniform(float(startYMUnadjusted), float(endYMUnadjusted)), 3)
            #     YMUnadjusted = driver.find_element_by_xpath(YMPath)
            #     YMUnadjusted.click()
            #     actions = ActionChains(driver)
            #     actions.move_by_offset(10, 10)
            #     actions.context_click()
            #     actions.perform()
            #     rm.find_element_by_name("Select All").click()
            #
            #     YMUnadjusted.click()
            #     actions = ActionChains(driver)
            #     actions.move_by_offset(10, 10)
            #     actions.context_click()
            #     actions.perform()
            #     rm.find_element_by_name("Cut").click()
            #     YMUnadjusted.send_keys(str(YMUnadjustedChoice))


        # January
        Dietpath = "//ComboBox[@AutomationId=\"JanDiet\"]"
        DietTDN = "//ComboBox[@AutomationId=\"JanTDN\"]"
        DietCP = "//ComboBox[@AutomationId=\"JanCP\"]"
        DietYM = "//ComboBox[@AutomationId=\"JanYM\"]"
        settingDiet(Dietpath, DietTDN, DietCP, DietYM)

        # Febuary
        Dietpath = "//ComboBox[@AutomationId=\"FebDiet\"]"
        DietTDN = "//ComboBox[@AutomationId=\"FebTDN\"]"
        DietCP = "//ComboBox[@AutomationId=\"FebCP\"]"
        DietYM = "//ComboBox[@AutomationId=\"FebYM\"]"
        settingDiet(Dietpath, DietTDN, DietCP, DietYM)

        # March
        Dietpath = "//ComboBox[@AutomationId=\"MarDiet\"]"
        DietTDN = "//ComboBox[@AutomationId=\"MarTDN\"]"
        DietCP = "//ComboBox[@AutomationId=\"MarCP\"]"
        DietYM = "//ComboBox[@AutomationId=\"MarYM\"]"
        settingDiet(Dietpath, DietTDN, DietCP, DietYM)

        # April
        Dietpath = "//ComboBox[@AutomationId=\"AprDiet\"]"
        DietTDN = "//ComboBox[@AutomationId=\"AprTDN\"]"
        DietCP = "//ComboBox[@AutomationId=\"AprCP\"]"
        DietYM = "//ComboBox[@AutomationId=\"AprYM\"]"
        settingDiet(Dietpath, DietTDN, DietCP, DietYM)

        # May
        Dietpath = "//ComboBox[@AutomationId=\"MayDiet\"]"
        DietTDN = "//ComboBox[@AutomationId=\"MayTDN\"]"
        DietCP = "//ComboBox[@AutomationId=\"MayCP\"]"
        DietYM = "//ComboBox[@AutomationId=\"MayYM\"]"
        settingDiet(Dietpath, DietTDN, DietCP, DietYM)

        # June
        Dietpath = "//ComboBox[@AutomationId=\"JunDiet\"]"
        DietTDN = "//ComboBox[@AutomationId=\"JunTDN\"]"
        DietCP = "//ComboBox[@AutomationId=\"JunCP\"]"
        DietYM = "//ComboBox[@AutomationId=\"JunYM\"]"
        settingDiet(Dietpath, DietTDN, DietCP, DietYM)

        # July
        Dietpath = "//ComboBox[@AutomationId=\"JulDiet\"]"
        DietTDN = "//ComboBox[@AutomationId=\"JulTDN\"]"
        DietCP = "//ComboBox[@AutomationId=\"JulCP\"]"
        DietYM = "//ComboBox[@AutomationId=\"JulYM\"]"
        settingDiet(Dietpath, DietTDN, DietCP, DietYM)

        # August
        Dietpath = "//ComboBox[@AutomationId=\"AugDiet\"]"
        DietTDN = "//ComboBox[@AutomationId=\"AugTDN\"]"
        DietCP = "//ComboBox[@AutomationId=\"AugCP\"]"
        DietYM = "//ComboBox[@AutomationId=\"AugYM\"]"
        settingDiet(Dietpath, DietTDN, DietCP, DietYM)

        # September
        Dietpath = "//ComboBox[@AutomationId=\"SepDiet\"]"
        DietTDN = "//ComboBox[@AutomationId=\"SepTDN\"]"
        DietCP = "//ComboBox[@AutomationId=\"SepCP\"]"
        DietYM = "//ComboBox[@AutomationId=\"SepYM\"]"
        settingDiet(Dietpath, DietTDN, DietCP, DietYM)

        # October
        Dietpath = "//ComboBox[@AutomationId=\"OctDiet\"]"
        DietTDN = "//ComboBox[@AutomationId=\"OctTDN\"]"
        DietCP = "//ComboBox[@AutomationId=\"OctCP\"]"
        DietYM = "//ComboBox[@AutomationId=\"OctYM\"]"
        settingDiet(Dietpath, DietTDN, DietCP, DietYM)

        # November
        Dietpath = "//ComboBox[@AutomationId=\"NovDiet\"]"
        DietTDN = "//ComboBox[@AutomationId=\"NovTDN\"]"
        DietCP = "//ComboBox[@AutomationId=\"NovCP\"]"
        DietYM = "//ComboBox[@AutomationId=\"NovYM\"]"
        settingDiet(Dietpath, DietTDN, DietCP, DietYM)

        # December
        Dietpath = "//ComboBox[@AutomationId=\"DecDiet\"]"
        DietTDN = "//ComboBox[@AutomationId=\"DecTDN\"]"
        DietCP = "//ComboBox[@AutomationId=\"DecCP\"]"
        DietYM = "//ComboBox[@AutomationId=\"DecYM\"]"
        settingDiet(Dietpath, DietTDN, DietCP, DietYM)


        # (0,30) None
        # (0,50) 2%
        # (0,70) 4%
        # (0,90) Custom
        def settingDietAdditive(dietAdditivePath):
            dietAddListPositions = [30, 50, 70, 90]
            dietAddPosi = random.choices(dietAddListPositions)
            dietAdditive = driver.find_element_by_xpath(dietAdditivePath)
            dietAdditive.click()
            time.sleep(2)
            actions = ActionChains(driver)
            actions.move_by_offset(0, dietAddPosi[0]).click()
            actions.perform()


        # January
        path = "//ComboBox[@AutomationId=\"JanDietAdditive\"]"
        settingDietAdditive(path)

        # Febuary
        path = "//ComboBox[@AutomationId=\"FebDietAdditive\"]"
        settingDietAdditive(path)

        # March
        path = "//ComboBox[@AutomationId=\"MarDietAdditive\"]"
        settingDietAdditive(path)

        # April
        path = "//ComboBox[@AutomationId=\"AprDietAdditive\"]"
        settingDietAdditive(path)

        # May
        path = "//ComboBox[@AutomationId=\"MayDietAdditive\"]"
        settingDietAdditive(path)

        # June
        path = "//ComboBox[@AutomationId=\"JunDietAdditive\"]"
        settingDietAdditive(path)

        # July
        path = "//ComboBox[@AutomationId=\"JulDietAdditive\"]"
        settingDietAdditive(path)

        # August
        path = "//ComboBox[@AutomationId=\"AugDietAdditive\"]"
        settingDietAdditive(path)

        # September
        path = "//ComboBox[@AutomationId=\"SepDietAdditive\"]"
        settingDietAdditive(path)

        # October
        path = "//ComboBox[@AutomationId=\"OctDietAdditive\"]"
        settingDietAdditive(path)

        # November
        path = "//ComboBox[@AutomationId=\"NovDietAdditive\"]"
        settingDietAdditive(path)

        # December
        path = "//ComboBox[@AutomationId=\"DecDietAdditive\"]"
        settingDietAdditive(path)


        # setting Manure
        # (0,30) Pature
        # (0,50) Solid Storage, ,
        # (0,70) Compost Intensive
        # (0,90) Compost Passive,
        # (0,110) Deep Bedding
        # (0,130) Custom

        def settingManure(manurePath, MCFPath, EFPath, VFPath):
            manureListPositions = [30, 50, 70, 90, 110, 130]
            manurePosi = random.choices(manureListPositions)
            diet = driver.find_element_by_xpath(manurePath)
            diet.click()
            time.sleep(2)
            actions = ActionChains(driver)
            actions.move_by_offset(0, manurePosi[0]).click()
            actions.perform()

            # if manurePosi[0] == 130:
            #     # setting MCF
            #     startMCF = sheet.cell_value(71, 1)
            #     endMCF = sheet.cell_value(71, 2)
            #
            #     MCFChoice = round(random.uniform(float(startMCF), float(endMCF)), 3)
            #     MCF = driver.find_element_by_xpath(MCFPath)
            #     MCF.click()
            #     actions = ActionChains(driver)
            #     actions.move_by_offset(10, 10)
            #     actions.context_click()
            #     actions.perform()
            #     rm.find_element_by_name("Select All").click()
            #
            #     MCF.click()
            #     actions = ActionChains(driver)
            #     actions.move_by_offset(10, 10)
            #     actions.context_click()
            #     actions.perform()
            #     rm.find_element_by_name("Cut").click()
            #     MCF.send_keys(str(MCFChoice))
            #
            #     # setting EF Direct
            #     startEF = sheet.cell_value(72, 1)
            #     endEF = sheet.cell_value(72, 2)
            #     EFChoice = round(random.uniform(float(startEF), float(endEF)), 3)
            #     EF = driver.find_element_by_xpath(EFPath)
            #     EF.click()
            #     actions = ActionChains(driver)
            #     actions.move_by_offset(10, 10)
            #     actions.context_click()
            #     actions.perform()
            #     rm.find_element_by_name("Select All").click()
            #
            #     EF.click()
            #     actions = ActionChains(driver)
            #     actions.move_by_offset(10, 10)
            #     actions.context_click()
            #     actions.perform()
            #     rm.find_element_by_name("Cut").click()
            #     EF.send_keys(str(EFChoice))
            #
            #     # setting volatization factor
            #     startVF = sheet.cell_value(73, 1)
            #     endVF = sheet.cell_value(73, 2)
            #     VFChoice = round(random.uniform(float(startVF), float(endVF)), 2)
            #     VF = driver.find_element_by_xpath(VFPath)
            #     VF.click()
            #     actions = ActionChains(driver)
            #     actions.move_by_offset(10, 10)
            #     actions.context_click()
            #     actions.perform()
            #     rm.find_element_by_name("Select All").click()
            #
            #     VF.click()
            #     actions = ActionChains(driver)
            #     actions.move_by_offset(10, 10)
            #     actions.context_click()
            #     actions.perform()
            #     rm.find_element_by_name("Cut").click()
            #     VF.send_keys(str(VFChoice))


        # January
        manurePath = "//ComboBox[@AutomationId=\"JanManure\"]"
        MCFPath = "//ComboBox[@AutomationId=\"JanMCF\"]"
        EFPath = "//ComboBox[@AutomationId=\"JanEFDirect\"]"
        VFPath = "//ComboBox[@AutomationId=\"JanFractionalVolatilization\"]"
        settingManure(manurePath, MCFPath, EFPath, VFPath)

        # Febuary
        manurePath = "//ComboBox[@AutomationId=\"FebManure\"]"
        MCFPath = "//ComboBox[@AutomationId=\"FebMCF\"]"
        EFPath = "//ComboBox[@AutomationId=\"FebEFDirect\"]"
        VFPath = "//ComboBox[@AutomationId=\"FebFractionalVolatilization\"]"
        settingManure(manurePath, MCFPath, EFPath, VFPath)

        # March
        manurePath = "//ComboBox[@AutomationId=\"MarManure\"]"
        MCFPath = "//ComboBox[@AutomationId=\"MarMCF\"]"
        EFPath = "//ComboBox[@AutomationId=\"MarEFDirect\"]"
        VFPath = "//ComboBox[@AutomationId=\"MarFractionalVolatilization\"]"
        settingManure(manurePath, MCFPath, EFPath, VFPath)

        # April
        Dietpath = "//ComboBox[@AutomationId=\"AprDiet\"]"
        DietTDN = "//ComboBox[@AutomationId=\"AprTDN\"]"
        DietCP = "//ComboBox[@AutomationId=\"AprCP\"]"
        DietYM = "//ComboBox[@AutomationId=\"AprYM\"]"
        settingDiet(Dietpath, DietTDN, DietCP, DietYM)

        # May
        manurePath = "//ComboBox[@AutomationId=\"MayManure\"]"
        MCFPath = "//ComboBox[@AutomationId=\"MayMCF\"]"
        EFPath = "//ComboBox[@AutomationId=\"MayEFDirect\"]"
        VFPath = "//ComboBox[@AutomationId=\"MayFractionalVolatilization\"]"
        settingManure(manurePath, MCFPath, EFPath, VFPath)

        # June
        manurePath = "//ComboBox[@AutomationId=\"JunManure\"]"
        MCFPath = "//ComboBox[@AutomationId=\"JunMCF\"]"
        EFPath = "//ComboBox[@AutomationId=\"JunEFDirect\"]"
        VFPath = "//ComboBox[@AutomationId=\"JunFractionalVolatilization\"]"
        settingManure(manurePath, MCFPath, EFPath, VFPath)

        # July
        manurePath = "//ComboBox[@AutomationId=\"JulManure\"]"
        MCFPath = "//ComboBox[@AutomationId=\"JulMCF\"]"
        EFPath = "//ComboBox[@AutomationId=\"JulEFDirect\"]"
        VFPath = "//ComboBox[@AutomationId=\"JulFractionalVolatilization\"]"
        settingManure(manurePath, MCFPath, EFPath, VFPath)

        # August
        manurePath = "//ComboBox[@AutomationId=\"AugManure\"]"
        MCFPath = "//ComboBox[@AutomationId=\"AugMCF\"]"
        EFPath = "//ComboBox[@AutomationId=\"AugEFDirect\"]"
        VFPath = "//ComboBox[@AutomationId=\"AugFractionalVolatilization\"]"
        settingManure(manurePath, MCFPath, EFPath, VFPath)

        # September
        manurePath = "//ComboBox[@AutomationId=\"SepManure\"]"
        MCFPath = "//ComboBox[@AutomationId=\"SepMCF\"]"
        EFPath = "//ComboBox[@AutomationId=\"SepEFDirect\"]"
        VFPath = "//ComboBox[@AutomationId=\"SepFractionalVolatilization\"]"
        settingManure(manurePath, MCFPath, EFPath, VFPath)

        # October
        manurePath = "//ComboBox[@AutomationId=\"OctManure\"]"
        MCFPath = "//ComboBox[@AutomationId=\"OctMCF\"]"
        EFPath = "//ComboBox[@AutomationId=\"OctEFDirect\"]"
        VFPath = "//ComboBox[@AutomationId=\"OctFractionalVolatilization\"]"
        settingManure(manurePath, MCFPath, EFPath, VFPath)

        # November
        manurePath = "//ComboBox[@AutomationId=\"NovManure\"]"
        MCFPath = "//ComboBox[@AutomationId=\"NovMCF\"]"
        EFPath = "//ComboBox[@AutomationId=\"NovEFDirect\"]"
        VFPath = "//ComboBox[@AutomationId=\"NovFractionalVolatilization\"]"
        settingManure(manurePath, MCFPath, EFPath, VFPath)

        # December
        manurePath = "//ComboBox[@AutomationId=\"DecManure\"]"
        MCFPath = "//ComboBox[@AutomationId=\"DecMCF\"]"
        EFPath = "//ComboBox[@AutomationId=\"DecEFDirect\"]"
        VFPath = "//ComboBox[@AutomationId=\"DecFractionalVolatilization\"]"
        settingManure(manurePath, MCFPath, EFPath, VFPath)

        #Saving HOLOS File
        driver.switch_to.window(driver.window_handles[0])
        # driver.find_element_by_name("File").click()
        menu = driver.find_element_by_name("File")
        menu.click()
        item = menu.find_element_by_name("Save As")
        actions = ActionChains(driver)
        actions.click(item)
        actions.perform()
        driver.find_element_by_xpath("//Edit[@Name=\"File name:\"]").send_keys("farm1")
        driver.switch_to.window(driver.window_handles[0])
        driver.find_element_by_name("Save").click()
        driver.find_element_by_name("Yes").click()


        driver.switch_to.window(driver.window_handles[0])
        # driver.find_element_by_name("File").click()
        menu = driver.find_element_by_name("Results")
        menu.click()
        item = menu.find_element_by_name("Emission Details CO2e Report")
        actions = ActionChains(driver)
        actions.click(item)
        actions.perform()

        print(driver.window_handles)
        driver.switch_to.window(driver.window_handles[0])
        p = driver.find_elements_by_xpath("//Pane[@Name=\"Holos Emissions Report\"]")
        print(p)
        table = driver.page_source
        print(table)

        actions = ActionChains(driver)
        actions.move_by_offset(146, 600)
        actions.context_click()
        actions.perform()

        desired_caps = {}
        desired_caps["app"] = "Root"
        ds = webdriver.Remote(command_executor='http://127.0.0.1:4723', desired_capabilities=desired_caps)

        ds.find_element_by_name("Export to Microsoft Excel").click()
        ds.find_element_by_name("Excel - 1 running window").click()
        ds.find_element_by_name("File Tab").click()
        ds.find_element_by_xpath("//ListItem[@Name=\"Save As\"]").click()
        ds.find_element_by_name("Browse").click()
        ds.find_element_by_name("Documents").click()
        ds.find_element_by_xpath("//Edit[@Name=\"File name:\"]").send_keys("1")
        ds.find_element_by_name("Save").click()

        # menu=ds.find_element_by_cla
        # item = menu.find_element_by_name("Save")
        # actions = ActionChains(ds)
        # actions.key_down(KeyboardInterrupt.C)
        # actions.click(item)
        # actions.perform()
        #
        # Actions actionObj = new Actions(driver);
        # actionObj.keyDown(Keys.CONTROL)
        #          .sendKeys(Keys.chord("A"))
        #          .keyUp(Keys.CONTROL)
        #          .perform();

        # table=driver.find_element_by_xpath("//Table[@Name=""]")
        # cell=driver.find_element_by_xpath("//Table/DataItem[@Name=\"Land Use Change \"]")
        # table = driver.find_element_by_class_name("UIATableView")

        # table = driver.find_element_by_class_name("table")
        # cells_in_table = table.find_elements_by_class_name("UIATableCell")
        # cell = cells_in_table[0]






    #except (RuntimeError, TypeError, NameError, ValueError):
    except(TypeError):
        print("Error Occured in the run ",r)



