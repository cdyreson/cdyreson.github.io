import selenium.webdriver as webdriver
import selenium.webdriver.support.ui as ui
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import *
import csv


import os


from time import sleep, time
import random

newProject = False
newLocation = False
browser = webdriver.Chrome("C:/Program Files (x86)/Google/chromedriver_win32/chromedriver.exe")  # Optional argument, if not specified will search path.
actions = ActionChains(browser)
browser.get('https://cometfarm.nrel.colostate.edu/Account/LogOn?ReturnUrl=%2fActivityType')
sleep(5)

#browser.maximize_window()
#sleep(5)
browser.find_element_by_id('ext-gen10').click()
sleep(2)
browser.find_element_by_id('getstartedbutton').click()
#browser.find_element_by_name('Start Using COMET-Farm').click()
username = browser.find_element_by_id("username")
password = browser.find_element_by_id("password")

username.send_keys("aditimaheshwari93@gmail.com")
password.send_keys("Aditi93@")
browser.find_element_by_id('signInButton').click()

browser.maximize_window()
sleep(5)
wb=load_workbook("D:/Study Material/Research/CometOutput.xlsx")
ws=wb["Sheet1"]

for i in range(3,100):
    if i != 3:
        browser.get('https://cometfarm.nrel.colostate.edu/Account/LogOn?ReturnUrl=%2fActivityType')
        sleep(5)
        browser.find_element_by_id('ext-gen10').click()
        sleep(2)
        browser.find_element_by_id('getstartedbutton').click()
        sleep(2)
    # delete farm1
    try:
        browser.find_element_by_id('farm1Project')
        browser.find_element_by_id('farm1Project').click()
        browser.find_element_by_id('deleteProjectLink').click()
        sleep(2)
        browser.find_element_by_xpath('//span[text()=\'Delete\']').click()

    except:
        pass

    if browser.find_element_by_id('10').get_attribute("checked") == False:
        browser.find_element_by_id('10').click()

    if browser.find_element_by_id('212320').get_attribute("checked") == False:
        browser.find_element_by_id('212320').click()
    sleep(5)
    browser.find_element_by_id('newprojectbutton').click()

    # Create new project and click button

    browser.find_element_by_xpath('//input[@placeholder="for example, Southwestern Region Crops"]').send_keys(
        "farm1")
    browser.find_element_by_xpath('//span[text()=\'Create\']').click()

    sleep(2)
    # if browser.find_element_by_id('10').get_attribute("checked")==False:
    browser.find_element_by_id('10').click()
    # if browser.find_element_by_id('212320').get_attribute("checked")==False:
    browser.find_element_by_id('212320').click()

    browser.find_element_by_id("definebutton").click()

    # # select and go to location
    sleep(5)
    location = browser.find_element_by_xpath('//*[@id="locationName"]')
    location.send_keys('84321')
    browser.find_element_by_xpath('//*[@id="locationGo"]').click()

    sleep(2)
    browser.find_element_by_xpath('//*[@id="uploadButton"]').click()
    sleep(2)
    if i % 3 == 0:
        num = random.randrange(1, 21, 1)
        n = str(num)
        print(n)
        browser.find_element_by_id("file").send_keys(
            "D://Study Material//Research//COMET//Orchards_Subset_20-20200320T221734Z-001//Orchards_Subset_20//Orchard" + str(
                n) + ".zip")
        ws.cell(i, 1).value = "Orchard" + str(n) + ".zip"
    else:
        num = random.randrange(1, 21, 1)
        n = str(num)
        print(n)
        browser.find_element_by_id("file").send_keys(
            "D://Study Material//Research//COMET//Forage_Subset_20-20200320T210731Z-001//Forage_Subset_20//Forage" + str(
                n) + ".zip")
        ws.cell(i, 1).value = "Forage" + str(n) + ".zip"

    browser.find_element_by_id("shapeFileUploadParcelIdent").send_keys("Split")
    browser.find_element_by_id("shapeFileUploadButton").click()
    sleep(2)
    browser.find_element_by_xpath('//button[text()="OK"]').click()

    # Click to save location

    sleep(2)
    browser.find_element_by_id("doneParcelsButton").click()

    sleep(2)
    # Select drop down
    managementList = ['Irrigation (Pre 1980s)', 'Upland Non-Irrigated (Pre 1980s)',
                      'Lowland Non-Irrigated (Pre 1980s)',
                      'Livestock Grazing']
    managementChoice = random.choice(managementList)
    management = browser.find_element_by_xpath(
        '//select[@id="management_pre_1980"]/option[text()="' + managementChoice + '"]')
    management.click()
    sleep(2)
    ws.cell(i, 2).value = managementChoice

    # 1980-2000 management
    afterMangementList = ['Irrigated: Annual Crops in Rotation',
                          'Irrigated: Annual Crops with Hay/Pasture in Rotation',
                          'Irrigated: Continuous Hay', 'Irrigated: Orchard or Vineyard',
                          'Non-Irrigated: Annual Crops in Rotation', 'Non-Irrigated: Continuous Hay',
                          'Non-Irrigated: Livestock Grazing', 'Non-Irrigated: Fallow-Grain',
                          'Non-Irrigated: Orchard or Vineyard']
    afterMangementChoice = random.choice(afterMangementList)
    browser.find_element_by_xpath(
        '//select[@id="management_1980"]/option[text()="' + afterMangementChoice + '"]').click()
    sleep(2)

    ws.cell(i, 3).value = afterMangementChoice
    # Tillage
    tillList = ['Intensive Tillage', 'Reduced Tillage', 'No Till']
    tillChoice = random.choice(tillList)
    browser.find_element_by_xpath('//select[@id="tillage_1980"]/option[text()="' + tillChoice + '"]').click()
    sleep(2)
    sleep(5)
    ws.cell(i, 4).value = tillChoice
    browser.find_element_by_xpath('//*[@id="nexthistoricbutton"]').click()
    # browser.find_element_by_xpath('/html/body/div[2]/form/div[2]/div[11]/span/a/span[2]').click()
    sleep(5)
    sleep(5)
    # clicking No Thanks continue button
    # browser.find_element_by_id("getBurnButton").click()
    browser.find_element_by_xpath('//span[text()="No Thanks, Continue >>"]').click()

    # choosing type of crop
    # 1151 Annual crop
    # 1155 Orchard/vinyard
    # 1153 Seasonal crop cover
    cropTypeList = ['ext-gen1153', 'ext-gen1155', 'ext-gen1151']
    if i % 3 == 0:
        cropChoice = 'ext-gen1155'
        ws.cell(i, 5).value = "Orchard/vinyard"
    else:
        cropChoice = random.choice(cropTypeList)
        if cropChoice == 'ext-gen1153':
            ws.cell(i, 5).value = "Seasonal crop cover"
        elif cropChoice == 'ext-gen1151':
            ws.cell(i, 5).value = "Annual crop"

    if cropChoice == "ext-gen1155":
        browser.find_element_by_xpath('//input[@id="ext-gen1155"]').click()
        sleep(2)
        # choose crop
        cropList = ['Cherries', 'Grape, Raisin', 'Grape, Raisin', 'Grape, Wine (GT1950 GDD)',
                    'Peaches and Nectarines',
                    'Pistachio']
        cropChoice = random.choice(cropList)
        ws.cell(i, 6).value = cropChoice
        browser.find_element_by_xpath('//div[@id="ext-gen1164"]').click()
        browser.find_element_by_xpath('//li[text()="' + cropChoice + '"]').click()

        # did you prune
        pruneChoice = random.randrange(1, 3, 1)
        if pruneChoice == 1:
            browser.find_element_by_xpath('//input[@id="ext-gen1052"]').click()
            ws.cell(i, 9).value = "Yes"
        else:
            ws.cell(i, 9).value = "No"

        # renew or clear vineyard
        renewChoice = random.randrange(1, 3, 1)
        if renewChoice == 1:
            browser.find_element_by_id("ext-gen1051").click()
            ws.cell(i, 10).value = "Yes"
        else:
            ws.cell(i, 10).value = "No"

        # click next button
        sleep(5)
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

        # add tillage
        browser.find_element_by_id('addTillage').click()
        sleep(2)
        browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
        browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
        browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
        sleep(2)
        browser.find_element_by_xpath('//span[text()="Mow"]').click()
        ws.cell(i, 17).value = "Mow"

        tillDate = browser.find_element_by_xpath('//div[text()="1/1/2000"]')
        tillDate.click()
        browser.find_element_by_xpath('//div[text()="1/1/2000"]').click()
        browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').click()
        browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
            Keys.ARROW_LEFT)
        browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
            Keys.ARROW_UP)
        browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
            Keys.ARROW_UP)
        browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
            Keys.ARROW_UP)
        browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
            Keys.ARROW_UP)
        ws.cell(i, 18).value = "05/01/2000"


        def setTill():
            browser.find_element_by_id('addTillage').click()
            sleep(2)
            browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
            browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
            browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
            sleep(2)
            browser.find_element_by_xpath('//span[text()="Mow"]').click()


        def setOrchardTillDate():
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(
                browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridTillageTillage"), -100,
                5).double_click().perform()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
                Keys.ARROW_LEFT)
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
                Keys.ARROW_UP)


        setTill()
        setOrchardTillDate()
        ws.cell(i, 19).value = "Mow"
        ws.cell(i, 20).value = "06/02/2000"

        setTill()
        setOrchardTillDate()
        ws.cell(i, 21).value = "Mow"
        ws.cell(i, 22).value = "07/03/2000"
        setTill()
        setOrchardTillDate()
        ws.cell(i, 23).value = "Mow"
        ws.cell(i, 24).value = "08/04/2000"
        setTill()
        setOrchardTillDate()
        ws.cell(i, 25).value = "Mow"
        ws.cell(i, 26).value = "09/05/2000"
        setTill()
        setOrchardTillDate()
        ws.cell(i, 27).value = "Mow"
        ws.cell(i, 28).value = "10/06/2000"
        setTill()
        setOrchardTillDate()
        ws.cell(i, 29).value = "Mow"
        ws.cell(i, 30).value = "11/07/2000"
        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()


        # add fertilizer
        def scrollFertilizer():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id(
                "jqxScrollThumbverticalScrollBarinnerListBoxdropdownlisteditorjqxGridNitrogenFertilizerType")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, 50).perform()
            actionChains.release()
            sleep(5)


        orchardFertilizerList = ['Ammonium Nitrate (34-0-0)', 'Ammonium Nitrate Phosphate (23-23-00)',
                                 'Ammonium Nitrate Phosphate (27-14-00)', 'Ammonium Phosphate Sulphate (16-20-00)',
                                 'Ammonium Polyphosphate Solution (10-34-00)', 'Ammonium Sulphate (21-00-00)',
                                 'Ammonium Thiosulphate Solution (12-00-00)', 'Anhydrous Ammonia (gas) (82-00-00)',
                                 'Calcium Ammonium Nitrate', 'Calcium Nitrate', 'Diammonium Phosphate (18-46-00)',
                                 'Element-N (N)', 'Element-P (P)', 'Mixed Blends',
                                 'Monoammonium Phosphate (11-55-00)',
                                 'Monoammonium Phosphate (12-51-00)', 'Potassium Nitrate', 'Urea (46-00-00)',
                                 'Urea Ammonium Nitrate (30-00-00)', 'Urea Ammonium Phosphate (27-27-00)',
                                 'Urea Ammonium Phosphate (34-17-00)']
        orchardFertilizerChoice = random.choice(orchardFertilizerList)
        ws.cell(i, 31).value = orchardFertilizerChoice

        browser.find_element_by_id("addNitrogen").click()
        sleep(2)
        browser.find_element_by_xpath('//div[text()="Ammonium Nitrate (34-0-0)"]').click()
        browser.find_element_by_xpath('//div[text()="Ammonium Nitrate (34-0-0)"]').click()
        sleep(2)
        browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridNitrogenFertilizerType").click()
        click = False
        print(orchardFertilizerChoice)
        while click == False:
            try:
                sleep(2)
                browser.find_element_by_xpath('//span[text()="' + orchardFertilizerChoice + '"]').click()
                click = True
            except:
                sleep(5)
                scrollFertilizer()
        # Set Total N Applied
        totalNChoice = random.randrange(19, 151 + 1, 1)
        actionChains = ActionChains(browser)
        actionChains.move_to_element_with_offset(
            browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridNitrogenFertilizerType"), 450,
            5).double_click().perform()
        browser.find_element_by_id("textboxeditorjqxGridNitrogenTotalApplied").send_keys(totalNChoice)
        ws.cell(i, 33).value = totalNChoice


        def setOrchardFertilizerDate(month, day):
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(
                browser.find_element_by_id("textboxeditorjqxGridNitrogenTotalApplied"),
                -500, 5).double_click().perform()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').send_keys(
                Keys.ARROW_LEFT)

            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').send_keys(
                Keys.ARROW_LEFT)
            for m in range(month):
                browser.find_element_by_xpath(
                    '//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').send_keys(
                    Keys.ARROW_UP)


        setOrchardFertilizerDate(4, 0)
        ws.cell(i, 32).value = "05/01/2000"

        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

        # add Manure
        orchardManureList = ['Alfalfa Meal', 'Beef Manure, Solid', 'Beef Slurry', 'Bone Meal',
                             'Chicken - Broiler (litter), Solid', 'Chicken - Broiler Slurry',
                             'Chicken - Layer Slurry',
                             'Chicken - Layer, Solid', 'Compost or Composted Manure, Solid', 'Dairy Manure, Solid',
                             'Dairy Slurry', 'Farmyard Manure, Solid', 'Feather Meal', 'Fish Emulsion',
                             'Fish Scrap',
                             'Guano', 'Horse Manure, Solid', 'Sheep Manure, Solid', 'Soybean Meal',
                             'Swine Manure, Slurry',
                             'Swine Manure, Solid']
        orchardManureChoice = random.choice(orchardManureList)
        ws.cell(i, 35).value = orchardManureChoice
        print(orchardManureChoice)
        browser.find_element_by_id("addManure").click()
        sleep(2)
        browser.find_element_by_xpath('//div[text()="Farmyard Manure, Solid"]').click()
        browser.find_element_by_xpath('//div[text()="Farmyard Manure, Solid"]').click()
        browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridManureManureType").click()
        sleep(5)


        def scrollManureDown():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id(
                "jqxScrollThumbverticalScrollBarinnerListBoxdropdownlisteditorjqxGridManureManureType")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, 50).perform()
            actionChains.release()
            sleep(2)


        def scrollManureUp():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id(
                "jqxScrollThumbverticalScrollBarinnerListBoxdropdownlisteditorjqxGridManureManureType")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, -50).perform()
            actionChains.release()
            sleep(2)


        click = False
        count = 0


        def chechManureOption():
            try:
                browser.find_element_by_xpath('//span[text()="' + orchardManureChoice + '"]').click()
                return True
            except:
                return False


        while click == False:
            if chechManureOption() == True:
                click = True
                break
            count = count + 1
            if (count != 1):
                scrollManureUp()
                if chechManureOption() == True:
                    click = True
                    break
                scrollManureUp()
                if chechManureOption() == True:
                    click = True
                    break
                scrollManureUp()
                if chechManureOption() == True:
                    click = True
                    break
            else:
                scrollManureUp()
                if chechManureOption() == True:
                    click = True
                    break
            scrollManureDown()
            if chechManureOption() == True:
                click = True
                break
            scrollManureDown()
            if chechManureOption() == True:
                click = True
                break
            scrollManureDown()
            if chechManureOption() == True:
                click = True
                break
            sleep(2)

        # amount applied
        browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridManureManureType").click()
        el = browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridManureManureType")
        e2 = browser.find_element_by_xpath('//div[text()="1.00"]')
        actionChains = ActionChains(browser)
        actionChains.move_to_element_with_offset(el, 230, 5).double_click().perform()

        seasonalAmountAppliedChoice = random.randrange(5, 25 + 1, 1)
        sleep(2)
        browser.find_element_by_id("textboxeditorjqxGridManureActualTotalApplied").send_keys(
            seasonalAmountAppliedChoice)
        ws.cell(i, 36).value = seasonalAmountAppliedChoice


        def setOrchardManureDate(month):
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(
                browser.find_element_by_id("textboxeditorjqxGridManureActualTotalApplied"), -300,
                5).double_click().perform()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridManureShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridManureShowDate"]').send_keys(
                Keys.ARROW_LEFT)
            for d in range(month):
                browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridManureShowDate"]').send_keys(
                    Keys.ARROW_UP)


        setOrchardManureDate(4)
        ws.cell(i, 34).value = "05/01/2000"

        # click next button
        sleep(5)
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()
        sleep(5)


        # add Irrigation
        def scrollIrrBar():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id("jqxScrollThumbverticalScrollBarjqxGridIrrigation")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, 200).perform()
            actionChains.release()
            sleep(2)


        def setIrr():
            browser.find_element_by_id("addIrrigation").click()
            try:
                browser.find_element_by_xpath('//div[text()="1"]').click()
            except:
                scrollIrrBar()
                browser.find_element_by_xpath('//div[text()="1"]').click()
            browser.find_element_by_xpath('//div[text()="1"]').click()
            irrChoice = round(random.uniform(float(1.80), float(2.25)), 2)
            browser.find_element_by_xpath('//input[@id="textboxeditorjqxGridIrrigationTotalApplied"]').send_keys(
                str(irrChoice))
            sleep(2)
            return irrChoice


        def setOrchardIrrDate(month, day):
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(
                browser.find_element_by_xpath('//input[@id="textboxeditorjqxGridIrrigationTotalApplied"]'), -100,
                5).double_click().perform()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                Keys.ARROW_LEFT)
            for i in range(day):
                browser.find_element_by_xpath(
                    '//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                    Keys.ARROW_UP)
            if month != 0:
                browser.find_element_by_xpath(
                    '//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                    Keys.ARROW_LEFT)
                for m in range(month):
                    browser.find_element_by_xpath(
                        '//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                        Keys.ARROW_UP)


        irrAmount = setIrr()
        setOrchardIrrDate(4, 0)
        ws.cell(i, 37).value = "05/01/2000"
        ws.cell(i, 38).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 39).value = "05/08/2000"
        ws.cell(i, 40).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 41).value = "05/15/2000"
        ws.cell(i, 42).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 43).value = "05/22/2000"
        ws.cell(i, 44).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 45).value = "05/29/2000"
        ws.cell(i, 46).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(1, 2)
        ws.cell(i, 47).value = "06/01/2000"
        ws.cell(i, 48).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 49).value = "06/08/2000"
        ws.cell(i, 50).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 51).value = "06/15/2000"
        ws.cell(i, 52).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 53).value = "06/22/2000"
        ws.cell(i, 54).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 55).value = "06/29/2000"
        ws.cell(i, 56).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(1, 1)
        ws.cell(i, 57).value = "07/01/2000"
        ws.cell(i, 58).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 59).value = "07/08/2000"
        ws.cell(i, 60).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 61).value = "07/15/2000"
        ws.cell(i, 62).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 63).value = "07/22/2000"
        ws.cell(i, 64).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 65).value = "07/29/2000"
        ws.cell(i, 66).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(1, 2)
        ws.cell(i, 67).value = "08/01/2000"
        ws.cell(i, 68).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 69).value = "08/08/2000"
        ws.cell(i, 70).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 71).value = "08/15/2000"
        ws.cell(i, 72).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 73).value = "08/22/2000"
        ws.cell(i, 74).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 75).value = "08/29/2000"
        ws.cell(i, 76).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(1, 2)
        ws.cell(i, 77).value = "09/01/2000"
        ws.cell(i, 78).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 79).value = "09/08/2000"
        ws.cell(i, 80).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 81).value = "09/15/2000"
        ws.cell(i, 82).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 83).value = "09/22/2000"
        ws.cell(i, 84).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 85).value = "09/29/2000"
        ws.cell(i, 86).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(1, 1)
        ws.cell(i, 87).value = "10/01/2000"
        ws.cell(i, 88).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 89).value = "10/08/2000"
        ws.cell(i, 90).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 91).value = "10/15/2000"
        ws.cell(i, 92).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 93).value = "10/22/2000"
        ws.cell(i, 94).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(0, 6)
        ws.cell(i, 95).value = "10/29/2000"
        ws.cell(i, 96).value = irrAmount
        irrAmount = setIrr()
        setOrchardIrrDate(1, 2)
        ws.cell(i, 97).value = "11/01/2000"
        ws.cell(i, 98).value = irrAmount

        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

        # add lime
        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

        # burning
        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

    elif cropChoice == "ext-gen1153":
        # Seasonal Crop
        # choosing type of crop
        browser.find_element_by_xpath('//input[@id="ext-gen1153"]').click()
        sleep(2)
        # choose crop
        SeasonalCropList = ['Annual Rye - Legume - Radish', 'Annual Rye - Legume', 'Annual Rye',
                            'Austrian Winter Pea',
                            'Cereal Rye', 'Clover', 'Corn', 'Forage Radish', 'Millet', 'Oilseed Radish',
                            'Winter Grain-Other', 'Sorghum', 'Vetch', 'Winter Wheat']
        SeasonalCropChoice = random.choice(SeasonalCropList)
        ws.cell(i, 6).value = SeasonalCropChoice

        browser.find_element_by_xpath('//input[@id="ext-gen1047"]').click()
        browser.find_element_by_xpath('//li[text()="' + SeasonalCropChoice + '"]').click()

        # choose planiting date
        browser.find_element_by_id("ext-gen1050").clear()
        plantingDateList = ['09/01/2000', '10/01/2000', '11/01/2000']
        SeasonalPlantingDateChoice = random.choice(plantingDateList)
        browser.find_element_by_id("ext-gen1050").send_keys(SeasonalPlantingDateChoice)
        ws.cell(i, 7).value = SeasonalPlantingDateChoice

        # add Harvesting
        browser.find_element_by_id("addHarvest").click()
        browser.find_element_by_xpath('//div[text()="0"]').click()
        browser.find_element_by_xpath('//div[text()="0"]').click()
        strawChoice = random.randrange(12, 72 + 1, 1)
        browser.find_element_by_id("textboxeditorjqxGridHarvestStrawStoverHayPct").send_keys(strawChoice)

        if SeasonalPlantingDateChoice == "09/01/2000":
            dateLoopChoice = random.randrange(4, 7, 1)
            try:
                browser.find_element_by_xpath('//div[text()="11/01/2000"]').click()
                browser.find_element_by_xpath('//div[text()="11/01/2000"]').click()
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
                for i in range(dateLoopChoice + 1):
                    browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
                if dateLoopChoice == 4:
                    harvestDate = "04/01/2001"
                elif dateLoopChoice == 5:
                    harvestDate = "05/01/2001"
                elif dateLoopChoice == 6:
                    harvestDate = "06/01/2001"
            except:
                browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
                browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
                harvestDate = "04/01/2001"
            ws.cell(i, 11).value = harvestDate

        if SeasonalPlantingDateChoice == "10/01/2000":
            dateLoopChoice = random.randrange(3, 6, 1)
            try:
                browser.find_element_by_xpath('//div[text()="12/01/2000"]').click()
                browser.find_element_by_xpath('//div[text()="12/01/2000"]').click()
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
                for i in range(dateLoopChoice + 1):
                    browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
                if dateLoopChoice == 3:
                    harvestDate = "04/01/2001"
                elif dateLoopChoice == 4:
                    harvestDate = "05/01/2001"
                elif dateLoopChoice == 5:
                    harvestDate = "06/01/2001"
            except:
                browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
                browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
                harvestDate = "04/01/2001"
            ws.cell(i, 11).value = harvestDate

        if SeasonalPlantingDateChoice == "11/01/2000":
            dateLoopChoice = random.randrange(2, 5, 1)
            try:
                browser.find_element_by_xpath('//div[text()="01/01/2001"]').click()
                browser.find_element_by_xpath('//div[text()="01/01/2001"]').click()
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
                for i in range(dateLoopChoice + 1):
                    browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
                if dateLoopChoice == 2:
                    harvestDate = "04/01/2001"
                elif dateLoopChoice == 3:
                    harvestDate = "05/01/2001"
                elif dateLoopChoice == 4:
                    harvestDate = "06/01/2001"
            except:
                browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
                browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
                harvestDate = "06/01/2001"
            ws.cell(i, 11).value = harvestDate
        print(harvestDate)

        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

        try:
            browser.find_element_by_xpath('//span[text()="Ok >>"]').click()
        except:
            pass


        # add tillage

        def scrollTill():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id(
                "jqxScrollThumbverticalScrollBarinnerListBoxdropdownlisteditorjqxGridTillageTillage")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, 50).perform()
            actionChains.release()
            sleep(2)


        tillageList = ['Intensive Tillage', 'Reduced Tillage', 'Mulch Tillage', 'Ridge Tillage', 'Strip Tillage',
                       'No Tillage', 'Crimp', 'Broadcast Seed', 'Aerial Seed', 'Zero Soil Disturbance',
                       'No Implement Passes']
        seasonalTillageChoice = random.choice(tillageList)
        ws.cell(i, 17).value = seasonalTillageChoice

        browser.find_element_by_id('addTillage').click()
        sleep(2)
        browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
        browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
        browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
        sleep(2)
        clickTill = False
        while clickTill == False:
            try:
                browser.find_element_by_xpath('//span[text()="' + seasonalTillageChoice + '"]').click()
                clickTill = True
            except:
                scrollTill()


        def setSeasonallTillDate(year, month, day):
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(
                browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridTillageTillage"), -100,
                5).double_click().perform()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
                Keys.ARROW_LEFT)
            for m in range(month):
                browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
                    Keys.ARROW_UP)
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
                Keys.ARROW_RIGHT)
            for d in range(day):
                browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
                    Keys.ARROW_DOWN)
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
                Keys.ARROW_RIGHT)
            for y in range(year):
                browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
                    Keys.ARROW_UP)


        if seasonalTillageChoice == 'Intensive Tillage' or seasonalTillageChoice == 'Reduced Tillage' or seasonalTillageChoice == 'Mulch Tillage' or seasonalTillageChoice == 'Ridge Tillage' or seasonalTillageChoice == 'Strip Tillage' or seasonalTillageChoice == 'No Tillage':
            if SeasonalPlantingDateChoice == "09/01/2000":
                setSeasonallTillDate(0, 7, 7)
                ws.cell(i, 18).value = "08/24/2000"
            if SeasonalPlantingDateChoice == "10/01/2000":
                setSeasonallTillDate(0, 8, 7)
                ws.cell(i, 18).value = "09/24/2000"
            if SeasonalPlantingDateChoice == "11/01/2000":
                setSeasonallTillDate(0, 9, 7)
                ws.cell(i, 18).value = "10/24/2000"
        elif seasonalTillageChoice == 'Crimp':
            if harvestDate == "05/01/2001":
                setSeasonallTillDate(1, 4, 0)
                ws.cell(i, 18).value = "05/01/2001"
            elif harvestDate == "06/01/2001":
                setSeasonallTillDate(1, 5, 0)
                ws.cell(i, 18).value = "06/01/2001"
            elif harvestDate == "07/01/2001":
                setSeasonallTillDate(1, 6, 0)
                ws.cell(i, 18).value = "07/01/2001"
        elif seasonalTillageChoice == 'Broadcast Seed' or seasonalTillageChoice == 'Aerial Seed':
            if SeasonalPlantingDateChoice == "09/01/2000":
                setSeasonallTillDate(0, 8, 0)
                ws.cell(i, 18).value = "09/01/2000"
            elif SeasonalPlantingDateChoice == "10/01/2000":
                setSeasonallTillDate(0, 9, 0)
                ws.cell(i, 18).value = "10/01/2000"
            elif SeasonalPlantingDateChoice == "11/01/2000":
                setSeasonallTillDate(0, 10, 0)
                ws.cell(i, 18).value = "11/01/2000"

        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()


        def scrollFertilizer():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id(
                "jqxScrollThumbverticalScrollBarinnerListBoxdropdownlisteditorjqxGridNitrogenFertilizerType")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, 50).perform()
            actionChains.release()
            sleep(3)


        # add fertilizer
        seasonalFertilizerList = ['Ammonium Nitrate (34-0-0)', 'Ammonium Nitrate Phosphate (23-23-00)',
                                  'Ammonium Nitrate Phosphate (27-14-00)', 'Ammonium Phosphate Sulphate (16-20-00)',
                                  'Ammonium Polyphosphate Solution (10-34-00)', 'Ammonium Sulphate (21-00-00)',
                                  'Ammonium Thiosulphate Solution (12-00-00)', 'Anhydrous Ammonia (gas) (82-00-00)',
                                  'Calcium Ammonium Nitrate', 'Calcium Nitrate', 'Diammonium Phosphate (18-46-00)',
                                  'Element-N (N)', 'Element-P (P)', 'Mixed Blends',
                                  'Monoammonium Phosphate (11-55-00)',
                                  'Monoammonium Phosphate (12-51-00)', 'Potassium Nitrate', 'Urea (46-00-00)',
                                  'Urea Ammonium Nitrate (30-00-00)', 'Urea Ammonium Phosphate (27-27-00)',
                                  'Urea Ammonium Phosphate (34-17-00)']
        seasonalFertilizerChoice = random.choice(seasonalFertilizerList)
        ws.cell(i, 31).value = seasonalFertilizerChoice

        browser.find_element_by_id("addNitrogen").click()
        sleep(2)
        browser.find_element_by_xpath('//div[text()="Ammonium Nitrate (34-0-0)"]').click()
        browser.find_element_by_xpath('//div[text()="Ammonium Nitrate (34-0-0)"]').click()
        sleep(2)
        browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridNitrogenFertilizerType").click()
        click = False
        print(seasonalFertilizerChoice)
        while click == False:
            try:
                sleep(2)
                browser.find_element_by_xpath('//span[text()="' + seasonalFertilizerChoice + '"]').click()
                click = True
            except:
                sleep(5)
                scrollFertilizer()

        # Set Total N Applied
        totalNChoice = random.randrange(19, 151 + 1, 1)
        actionChains = ActionChains(browser)
        actionChains.move_to_element_with_offset(
            browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridNitrogenFertilizerType"), 450,
            5).double_click().perform()
        browser.find_element_by_id("textboxeditorjqxGridNitrogenTotalApplied").send_keys(totalNChoice)
        ws.cell(i, 33).value = totalNChoice


        def setSeasonallFertilizerDate(month, day):
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(
                browser.find_element_by_id("textboxeditorjqxGridNitrogenTotalApplied"),
                -500, 5).double_click().perform()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').send_keys(
                Keys.ARROW_LEFT)
            for d in range(day):
                browser.find_element_by_xpath(
                    '//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').send_keys(
                    Keys.ARROW_UP)
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').send_keys(
                Keys.ARROW_LEFT)
            for m in range(month):
                browser.find_element_by_xpath(
                    '//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').send_keys(
                    Keys.ARROW_UP)


        # Setting date
        if SeasonalPlantingDateChoice == "09/01/2000":
            setSeasonallFertilizerDate(8, 13)
            ws.cell(i, 32).value = "09/16/2000"
        if SeasonalPlantingDateChoice == "10/01/2000":
            setSeasonallFertilizerDate(9, 13)
            ws.cell(i, 32).value = "10/16/2000"
        if SeasonalPlantingDateChoice == "11/01/2000":
            setSeasonallFertilizerDate(10, 13)
            ws.cell(i, 32).value = "11/16/2000"
        else:
            ws.cell(i, 32).value = "01/01/2000"

        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

        # add Manure
        seasonalManureList = ['Alfalfa Meal', 'Beef Manure, Solid', 'Beef Slurry', 'Bone Meal',
                              'Chicken - Broiler (litter), Solid', 'Chicken - Broiler Slurry',
                              'Chicken - Layer Slurry',
                              'Chicken - Layer, Solid', 'Compost or Composted Manure, Solid', 'Dairy Manure, Solid',
                              'Dairy Slurry', 'Farmyard Manure, Solid', 'Feather Meal', 'Fish Emulsion',
                              'Fish Scrap',
                              'Guano', 'Horse Manure, Solid', 'Sheep Manure, Solid', 'Soybean Meal',
                              'Swine Manure, Slurry',
                              'Swine Manure, Solid']
        seasonalManureChoice = random.choice(seasonalManureList)
        print(seasonalManureChoice)
        ws.cell(i, 35).value = seasonalManureChoice
        browser.find_element_by_id("addManure").click()
        sleep(2)
        browser.find_element_by_xpath('//div[text()="Farmyard Manure, Solid"]').click()
        browser.find_element_by_xpath('//div[text()="Farmyard Manure, Solid"]').click()
        browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridManureManureType").click()
        sleep(5)


        def scrollManureDown():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id(
                "jqxScrollThumbverticalScrollBarinnerListBoxdropdownlisteditorjqxGridManureManureType")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, 50).perform()
            actionChains.release()
            sleep(2)


        def scrollManureUp():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id(
                "jqxScrollThumbverticalScrollBarinnerListBoxdropdownlisteditorjqxGridManureManureType")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, -50).perform()
            actionChains.release()
            sleep(2)


        click = False
        count = 0


        def chechManureOption():
            try:
                browser.find_element_by_xpath('//span[text()="' + seasonalManureChoice + '"]').click()
                # browser.find_element_by_id('listitem0innerListBoxdropdownlisteditorjqxGridManureManureType').click()
                return True
            except:
                return False


        while click == False:
            # browser.find_element_by_id('listitem0innerListBoxdropdownlisteditorjqxGridManureManureType').click()
            if chechManureOption() == True:
                click = True
                break
            count = count + 1
            if (count != 1):
                scrollManureUp()
                if chechManureOption() == True:
                    click = True
                    break
                scrollManureUp()
                if chechManureOption() == True:
                    click = True
                    break
                scrollManureUp()
                if chechManureOption() == True:
                    click = True
                    break
            else:
                scrollManureUp()
                if chechManureOption() == True:
                    click = True
                    break
            scrollManureDown()
            if chechManureOption() == True:
                click = True
                break
            scrollManureDown()
            if chechManureOption() == True:
                click = True
                break
            scrollManureDown()
            if chechManureOption() == True:
                click = True
                break
            sleep(2)

        # amount applied
        browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridManureManureType").click()
        el = browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridManureManureType")
        e2 = browser.find_element_by_xpath('//div[text()="1.00"]')
        actionChains = ActionChains(browser)
        actionChains.move_to_element_with_offset(el, 230, 5).double_click().perform()

        seasonalAmountAppliedChoice = random.randrange(5, 25 + 1, 1)
        sleep(2)
        ws.cell(i, 36).value = seasonalAmountAppliedChoice
        browser.find_element_by_id("textboxeditorjqxGridManureActualTotalApplied").send_keys(
            seasonalAmountAppliedChoice)


        def setSeasonallManureDate(month, day):
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(
                browser.find_element_by_id("textboxeditorjqxGridManureActualTotalApplied"), -300,
                5).double_click().perform()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridManureShowDate"]').click()
            for m in range(day):
                browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridManureShowDate"]').send_keys(
                    Keys.ARROW_UP)
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridManureShowDate"]').send_keys(
                Keys.ARROW_LEFT)
            for d in range(month):
                browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridManureShowDate"]').send_keys(
                    Keys.ARROW_UP)


        if SeasonalPlantingDateChoice == "09/01/2000":
            setSeasonallManureDate(7, 13)
            ws.cell(i, 34).value = "08/14/2000"
        if SeasonalPlantingDateChoice == "10/01/2000":
            setSeasonallManureDate(8, 13)
            ws.cell(i, 34).value = "09/15/2000"
        if SeasonalPlantingDateChoice == "11/01/2000":
            setSeasonallManureDate(9, 13)
            ws.cell(i, 34).value = "10/14/2000"

        # click next button
        sleep(5)
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()
        sleep(5)


        # add Irrigation
        def scrollIrrBar():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id("jqxScrollThumbverticalScrollBarjqxGridIrrigation")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, 200).perform()
            actionChains.release()
            sleep(2)


        print(SeasonalPlantingDateChoice)
        print(harvestDate)


        def setIrr(scroll):
            browser.find_element_by_id("addIrrigation").click()
            sleep(2)
            try:
                browser.find_element_by_xpath('//div[text()="1"]').click()
            except:
                scrollIrrBar()
                sleep(2)
                browser.find_element_by_xpath('//div[text()="1"]').click()
            browser.find_element_by_xpath('//div[text()="1"]').click()
            sleep(2)
            irrChoice = round(random.uniform(float(1), float(2)), 2)
            browser.find_element_by_xpath('//input[@id="textboxeditorjqxGridIrrigationTotalApplied"]').send_keys(
                str(irrChoice))
            sleep(2)
            return irrChoice


        def setSeasonalIrrDate(date, month, day, year):
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(
                browser.find_element_by_xpath('//input[@id="textboxeditorjqxGridIrrigationTotalApplied"]'), -100,
                5).double_click().perform()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').click()
            if year == 1:
                browser.find_element_by_xpath(
                    '//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                    Keys.ARROW_UP)
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                Keys.ARROW_LEFT)
            for i in range(day):
                browser.find_element_by_xpath(
                    '//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                    Keys.ARROW_UP)
            if month != 0:
                browser.find_element_by_xpath(
                    '//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                    Keys.ARROW_LEFT)
                for m in range(month):
                    browser.find_element_by_xpath(
                        '//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                        Keys.ARROW_UP)


        def setIrrForSep():
            irrAmount = setIrr(False)
            setSeasonalIrrDate("9/2/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 39).value = "09/08/2000"
                ws.cell(i, 40).value = irrAmount
            irrAmount = setIrr(False)
            setSeasonalIrrDate("9/9/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 41).value = "09/15/2000"
                ws.cell(i, 42).value = irrAmount
            irrAmount = setIrr(False)
            setSeasonalIrrDate("9/16/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 43).value = "09/22/2000"
                ws.cell(i, 44).value = irrAmount
            irrAmount = setIrr(False)
            setSeasonalIrrDate("9/23/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 45).value = "09/29/2000"
                ws.cell(i, 46).value = irrAmount
            irrAmount = setIrr(False)
            setSeasonalIrrDate("9/30/2000", 1, 1, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 47).value = "10/01/2000"
                ws.cell(i, 48).value = irrAmount


        def setIrrForOct():
            irrAmount = setIrr(False)
            setSeasonalIrrDate("10/2/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 49).value = "10/08/2000"
                ws.cell(i, 50).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 39).value = "10/08/2000"
                ws.cell(i, 40).value = irrAmount
            irrAmount = setIrr(False)
            setSeasonalIrrDate("10/9/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 51).value = "10/15/2000"
                ws.cell(i, 52).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 41).value = "10/15/2000"
                ws.cell(i, 42).value = irrAmount
            irrAmount = setIrr(False)
            setSeasonalIrrDate("10/16/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 53).value = "10/22/2000"
                ws.cell(i, 54).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 43).value = "10/22/2000"
                ws.cell(i, 44).value = irrAmount
            irrAmount = setIrr(False)
            setSeasonalIrrDate("10/23/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 55).value = "10/29/2000"
                ws.cell(i, 56).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 45).value = "10/29/2000"
                ws.cell(i, 46).value = irrAmount
            irrAmount = setIrr(False)
            setSeasonalIrrDate("10/30/2000", 1, 2, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 57).value = "11/01/2000"
                ws.cell(i, 58).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 47).value = "11/01/2000"
                ws.cell(i, 48).value = irrAmount


        def setIrrNovTillApr():
            irrAmount = setIrr(False)
            setSeasonalIrrDate("11/2/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 59).value = "11/08/2000"
                ws.cell(i, 60).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 49).value = "11/08/2000"
                ws.cell(i, 50).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 39).value = "11/08/2000"
                ws.cell(i, 40).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("11/9/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 61).value = "11/15/2000"
                ws.cell(i, 62).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 51).value = "11/15/2000"
                ws.cell(i, 52).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 41).value = "11/15/2000"
                ws.cell(i, 42).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("11/16/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 63).value = "11/22/2000"
                ws.cell(i, 64).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 53).value = "11/22/2000"
                ws.cell(i, 54).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 43).value = "11/22/2000"
                ws.cell(i, 44).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("11/23/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 65).value = "11/29/2000"
                ws.cell(i, 66).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 55).value = "11/29/2000"
                ws.cell(i, 56).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 46).value = "11/29/2000"
                ws.cell(i, 46).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("11/30/2000", 1, 1, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 67).value = "12/01/2000"
                ws.cell(i, 68).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 57).value = "12/01/2000"
                ws.cell(i, 58).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 47).value = "12/01/2000"
                ws.cell(i, 48).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("12/2/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 69).value = "12/08/2000"
                ws.cell(i, 70).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 59).value = "12/08/2000"
                ws.cell(i, 60).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 49).value = "12/08/2000"
                ws.cell(i, 50).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("12/9/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 71).value = "12/15/2000"
                ws.cell(i, 72).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 61).value = "12/15/2000"
                ws.cell(i, 62).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 51).value = "12/15/2000"
                ws.cell(i, 52).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("12/16/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 73).value = "12/22/2000"
                ws.cell(i, 74).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 63).value = "12/22/2000"
                ws.cell(i, 64).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 53).value = "12/22/2000"
                ws.cell(i, 54).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("12/23/2000", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 75).value = "12/29/2000"
                ws.cell(i, 76).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 65).value = "12/29/2000"
                ws.cell(i, 66).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 55).value = "12/29/2000"
                ws.cell(i, 56).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("12/30/2000", 1, 2, 1)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 77).value = "01/01/2001"
                ws.cell(i, 78).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 67).value = "01/01/2001"
                ws.cell(i, 68).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 57).value = "01/01/2001"
                ws.cell(i, 58).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("1/2/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 79).value = "01/08/2001"
                ws.cell(i, 80).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 69).value = "01/08/2001"
                ws.cell(i, 70).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 59).value = "01/08/2001"
                ws.cell(i, 60).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("1/9/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 81).value = "01/15/2001"
                ws.cell(i, 82).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 71).value = "01/15/2001"
                ws.cell(i, 72).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 61).value = "01/15/2001"
                ws.cell(i, 62).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("1/16/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 83).value = "01/22/2001"
                ws.cell(i, 84).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 73).value = "01/22/2001"
                ws.cell(i, 74).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 63).value = "01/22/2001"
                ws.cell(i, 64).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("1/23/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 85).value = "01/29/2001"
                ws.cell(i, 86).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 75).value = "01/29/2001"
                ws.cell(i, 76).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 65).value = "01/29/2001"
                ws.cell(i, 66).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("1/30/2001", 1, 2, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 87).value = "02/01/2001"
                ws.cell(i, 88).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 77).value = "02/01/2001"
                ws.cell(i, 78).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 67).value = "02/01/2001"
                ws.cell(i, 68).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("2/2/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 89).value = "02/08/2001"
                ws.cell(i, 90).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 79).value = "02/08/2001"
                ws.cell(i, 80).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 69).value = "02/08/2001"
                ws.cell(i, 70).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("2/9/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 91).value = "02/15/2001"
                ws.cell(i, 92).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 81).value = "02/15/2001"
                ws.cell(i, 82).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 71).value = "02/15/2001"
                ws.cell(i, 72).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("2/16/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 93).value = "02/22/2001"
                ws.cell(i, 94).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 83).value = "02/22/2001"
                ws.cell(i, 84).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 73).value = "02/22/2001"
                ws.cell(i, 74).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("2/23/2001", 0, 4, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 95).value = "02/29/2001"
                ws.cell(i, 96).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 85).value = "02/29/2001"
                ws.cell(i, 86).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 75).value = "02/29/2001"
                ws.cell(i, 76).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("2/28/2001", 1, 1, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 97).value = "03/01/2001"
                ws.cell(i, 98).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 87).value = "03/01/2001"
                ws.cell(i, 88).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 77).value = "03/01/2001"
                ws.cell(i, 78).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("3/2/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 99).value = "03/08/2001"
                ws.cell(i, 100).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 89).value = "03/08/2001"
                ws.cell(i, 90).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 79).value = "03/08/2001"
                ws.cell(i, 80).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("3/9/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 101).value = "03/15/2001"
                ws.cell(i, 102).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 91).value = "03/15/2001"
                ws.cell(i, 92).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 81).value = "03/15/2001"
                ws.cell(i, 82).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("3/16/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 103).value = "03/22/2001"
                ws.cell(i, 104).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 93).value = "03/22/2001"
                ws.cell(i, 94).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 83).value = "03/22/2001"
                ws.cell(i, 84).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("3/23/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 105).value = "03/29/2001"
                ws.cell(i, 106).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 95).value = "03/29/2001"
                ws.cell(i, 96).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 85).value = "03/29/2001"
                ws.cell(i, 86).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("3/30/2001", 1, 2, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 107).value = "04/01/2001"
                ws.cell(i, 108).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 97).value = "04/01/2001"
                ws.cell(i, 98).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 87).value = "04/01/2001"
                ws.cell(i, 88).value = irrAmount


        def setIrrForApr():
            irrAmount = setIrr(True)
            setSeasonalIrrDate("4/2/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 109).value = "04/08/2001"
                ws.cell(i, 110).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 99).value = "04/08/2001"
                ws.cell(i, 100).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 89).value = "04/08/2001"
                ws.cell(i, 90).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("4/9/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 111).value = "04/15/2001"
                ws.cell(i, 112).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 101).value = "04/15/2001"
                ws.cell(i, 102).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 91).value = "04/15/2001"
                ws.cell(i, 92).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("4/16/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 113).value = "04/22/2001"
                ws.cell(i, 114).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 103).value = "04/22/2001"
                ws.cell(i, 104).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 93).value = "04/22/2001"
                ws.cell(i, 94).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("4/23/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 115).value = "04/29/2001"
                ws.cell(i, 116).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 105).value = "04/29/2001"
                ws.cell(i, 106).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 95).value = "04/29/2001"
                ws.cell(i, 96).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("4/30/2001", 1, 1, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 117).value = "05/01/2001"
                ws.cell(i, 118).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 107).value = "05/01/2001"
                ws.cell(i, 108).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 97).value = "05/01/2001"
                ws.cell(i, 98).value = irrAmount


        def setIrrForMay():
            irrAmount = setIrr(True)
            setSeasonalIrrDate("5/2/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 119).value = "05/08/2001"
                ws.cell(i, 120).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 109).value = "05/08/2001"
                ws.cell(i, 110).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 99).value = "05/08/2001"
                ws.cell(i, 100).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("5/9/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 121).value = "05/15/2001"
                ws.cell(i, 122).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 111).value = "05/15/2001"
                ws.cell(i, 112).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 101).value = "05/15/2001"
                ws.cell(i, 102).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("5/16/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 123).value = "05/22/2001"
                ws.cell(i, 124).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 113).value = "05/22/2001"
                ws.cell(i, 114).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 103).value = "05/22/2001"
                ws.cell(i, 104).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("5/23/2001", 0, 6, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 125).value = "05/29/2001"
                ws.cell(i, 126).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 115).value = "05/29/2001"
                ws.cell(i, 116).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 105).value = "05/29/2001"
                ws.cell(i, 106).value = irrAmount
            irrAmount = setIrr(True)
            setSeasonalIrrDate("5/30/2001", 1, 2, 0)
            if SeasonalPlantingDateChoice == "09/01/2000":
                ws.cell(i, 127).value = "06/01/2001"
                ws.cell(i, 128).value = irrAmount
            elif SeasonalPlantingDateChoice == '10/01/2000':
                ws.cell(i, 117).value = "06/01/2001"
                ws.cell(i, 118).value = irrAmount
            elif SeasonalPlantingDateChoice == '11/01/2000':
                ws.cell(i, 107).value = "06/01/2001"
                ws.cell(i, 108).value = irrAmount


        if SeasonalPlantingDateChoice == "09/01/2000":
            irrAmount = setIrr(False)
            setSeasonalIrrDate("", 8, 0, 0)
            ws.cell(i, 37).value = "09/01/2000"
            ws.cell(i, 38).value = irrAmount

            if harvestDate == "04/01/2001":
                setIrrForSep()
                setIrrForOct()
                setIrrNovTillApr()

            if harvestDate == "05/01/2001":
                setIrrForSep()
                setIrrForOct()
                setIrrNovTillApr()
                setIrrForApr()
            if harvestDate == "06/01/2001":
                setIrrForSep()
                setIrrForOct()
                setIrrNovTillApr()
                setIrrForApr()
                setIrrForMay()

        elif SeasonalPlantingDateChoice == '10/01/2000':
            irrAmount = setIrr(False)
            setSeasonalIrrDate("", 9, 0, 0)
            ws.cell(i, 37).value = "10/01/2000"
            ws.cell(i, 38).value = irrAmount
            if harvestDate == "04/01/2001":
                setIrrForOct()
                setIrrNovTillApr()
            elif harvestDate == "05/01/2001":
                setIrrForOct()
                setIrrNovTillApr()
                setIrrForApr()
            elif harvestDate == "06/01/2001":
                setIrrForOct()
                setIrrNovTillApr()
                setIrrForApr()
                setIrrForMay()

        elif SeasonalPlantingDateChoice == "11/01/2000":
            irrAmount = setIrr(False)
            setSeasonalIrrDate("", 10, 0, 0)
            ws.cell(i, 37).value = "11/01/2000"
            ws.cell(i, 38).value = irrAmount
            if harvestDate == "04/01/2001":
                setIrrNovTillApr()
            elif harvestDate == "05/01/2001":
                setIrrNovTillApr()
                setIrrForApr()
            elif harvestDate == "06/01/2001":
                setIrrNovTillApr()
                setIrrForApr()
                setIrrForMay()

        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

        # add lime
        ws.cell(i, 129).value = "No Liming"
        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

        # burning
        ws.cell(i, 130).value = "No Burning"
        # click next button
        sleep(5)
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

    elif cropChoice == "ext-gen1151":
        # Annual crop
        # choosing type of crop
        browser.find_element_by_xpath('//input[@id="ext-gen1151"]').click()
        sleep(2)
        # choose crop#
        annualCropList = ['Alfalfa', 'Barley', 'Corn', 'Corn Silage', 'Grass', 'Grass-Legume Mix', 'Rye', 'Sorghum',
                          'Spring Wheat']
        annualCropChoice = random.choice(annualCropList)
        ws.cell(i, 6).value = annualCropChoice
        browser.find_element_by_xpath('//input[@id="ext-gen1046"]').click()
        browser.find_element_by_xpath('//li[text()="' + annualCropChoice + '"]').click()
        plantingDateChoice = "04/01/2000"
        if annualCropChoice == "Alfalfa" or annualCropChoice == "Grass" or annualCropChoice == "Grass-Legume Mix":
            browser.find_element_by_xpath('//input[@class="x-form-field x-form-checkbox"]').click()
            ws.cell(i, 8).value = "Yes"

        else:
            # choose planiting date
            browser.find_element_by_id("ext-gen1050").clear()
            plantingDateList = ['05/01/2000', '06/01/2000', '07/01/2000']
            plantingDateChoice = random.choice(plantingDateList)
            browser.find_element_by_id("ext-gen1050").send_keys(plantingDateChoice)
            ws.cell(i, 8).value = "No"
            ws.cell(i, 7).value = plantingDateChoice

        # add Harvesting
        browser.find_element_by_id("addHarvest").click()
        browser.find_element_by_xpath('//div[text()="0"]').click()
        browser.find_element_by_xpath('//div[text()="0"]').click()

        # add Straw
        annualstrawChoice = random.randrange(12, 72 + 1, 1)
        actionChains = ActionChains(browser)
        actionChains.move_to_element_with_offset(browser.find_element_by_xpath(
            '//span[@class="jqx-checkbox-check-checked jqx-checkbox-check-checked-energyblue"]'), 200,
            5).double_click().perform()
        ws.cell(i, 13).value = annualstrawChoice
        browser.find_element_by_id("textboxeditorjqxGridHarvestStrawStoverHayPct").send_keys(annualstrawChoice)

        # add yield
        if annualCropChoice == "Alfalfa":
            annualYield = 4.3
        elif annualCropChoice == "Barley":
            annualYield = 93
        elif annualCropChoice == "Corn":
            annualYield = 143
        elif annualCropChoice == "Corn Silage":
            annualYield = 24
        elif annualCropChoice == "Grass":
            annualYield = 3.85
        elif annualCropChoice == "Grass-Legume Mix":
            annualYield = 4.3
        elif annualCropChoice == "Rye":
            annualYield = 54
        elif annualCropChoice == "Sorghum":
            annualYield = 54
        elif annualCropChoice == "Spring Wheat":
            annualYield = 54
        elif annualCropChoice == "Winter Wheat":
            annualYield = 54
        ws.cell(i, 12).value = annualstrawChoice

        actionChains = ActionChains(browser)
        actionChains.move_to_element_with_offset(browser.find_element_by_xpath(
            '//span[@class="jqx-checkbox-check-checked jqx-checkbox-check-checked-energyblue"]'), 100,
            5).double_click().perform()
        browser.find_element_by_id("textboxeditorjqxGridHarvestHarvestYield").send_keys(str(annualYield))
        browser.find_element_by_id("addGrazing").click()


        def addGrazing(start, end):
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(browser.find_element_by_xpath('//div[text()="1"]'), -100,
                                                     5).double_click().perform()
            browser.find_element_by_id("inputdatetimeeditorjqxGridGrazingShowDateStart").send_keys(Keys.ARROW_LEFT)
            browser.find_element_by_id("inputdatetimeeditorjqxGridGrazingShowDateStart").send_keys(Keys.ARROW_LEFT)
            for s in range(start):
                browser.find_element_by_id("inputdatetimeeditorjqxGridGrazingShowDateStart").send_keys(
                    Keys.ARROW_DOWN)
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(browser.find_element_by_xpath('//div[text()="1"]'), -50,
                                                     5).double_click().perform()
            browser.find_element_by_id("inputdatetimeeditorjqxGridGrazingShowDateEnd").send_keys(Keys.ARROW_LEFT)
            browser.find_element_by_id("inputdatetimeeditorjqxGridGrazingShowDateEnd").send_keys(Keys.ARROW_LEFT)
            for e in range(end):
                browser.find_element_by_id("inputdatetimeeditorjqxGridGrazingShowDateEnd").send_keys(Keys.ARROW_UP)
            browser.find_element_by_xpath('//div[text()="1"]').click()
            browser.find_element_by_xpath('//div[text()="1"]').click()
            browser.find_element_by_id("textboxeditorjqxGridGrazingRestPeriod").send_keys("21")


        harvestDate = "04/01/2000"
        if plantingDateChoice == "04/01/2000":
            dateLoopChoice = random.randrange(3, 7, 1)
            browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
            browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
            browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
            browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
            for i in range(dateLoopChoice):
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
            print(dateLoopChoice)
            if dateLoopChoice == 3:
                harvestDate = "07/01/2000"
            elif dateLoopChoice == 4:
                harvestDate = "08/01/2000"
            elif dateLoopChoice == 5:
                harvestDate = "09/01/2000"
            elif dateLoopChoice == 6:
                harvestDate = "10/01/2000"
            elif dateLoopChoice == 7:
                harvestDate = "11/01/2000"
            addGrazing(0, 6)

        if plantingDateChoice == "05/01/2000":
            dateLoopChoice = random.randrange(0, 5, 1)
            try:
                browser.find_element_by_xpath('//div[text()="07/01/2000"]').click()
                browser.find_element_by_xpath('//div[text()="07/01/2000"]').click()
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
                for i in range(dateLoopChoice):
                    browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
                print(dateLoopChoice)
                if dateLoopChoice == 2:
                    harvestDate = "07/01/2000"
                elif dateLoopChoice == 3:
                    harvestDate = "08/01/2000"
                elif dateLoopChoice == 4:
                    harvestDate = "09/01/2000"
                elif dateLoopChoice == 5:
                    harvestDate = "10/01/2000"
                elif dateLoopChoice == 6:
                    harvestDate = "11/01/2000"

            except:
                browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
                browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
            # add grazing
            addGrazing(3, 3)

        if plantingDateChoice == "06/01/2000":
            dateLoopChoice = random.randrange(0, 4, 1)
            try:
                browser.find_element_by_xpath('//div[text()="08/01/2000"]').click()
                browser.find_element_by_xpath('//div[text()="08/01/2000"]').click()
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
                for i in range(dateLoopChoice):
                    browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
                print(dateLoopChoice)
                if dateLoopChoice == 0:
                    harvestDate = "08/01/2000"
                elif dateLoopChoice == 1:
                    harvestDate = "09/01/2000"
                elif dateLoopChoice == 2:
                    harvestDate = "10/01/2000"
                elif dateLoopChoice == 3:
                    harvestDate = "11/01/2000"

            except:
                browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
                browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
            # add  grazing
            addGrazing(4, 2)

        if plantingDateChoice == "07/01/2000":
            dateLoopChoice = random.randrange(0, 3, 1)
            try:
                browser.find_element_by_xpath('//div[text()="09/01/2000"]').click()
                browser.find_element_by_xpath('//div[text()="09/01/2000"]').click()
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_LEFT)
                for i in range(dateLoopChoice):
                    browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
                print(dateLoopChoice)
                if dateLoopChoice == 0:
                    harvestDate = "09/01/2000"
                elif dateLoopChoice == 1:
                    harvestDate = "10/01/2000"
                elif dateLoopChoice == 2:
                    harvestDate = "11/01/2000"
            except:
                browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
                browser.find_element_by_xpath('//div[text()="04/01/2000"]').click()
                browser.find_element_by_id("inputdatetimeeditorjqxGridHarvestShowDate").send_keys(Keys.ARROW_UP)
                harvestDate = "04/01/2001"
            # add Grazing
            addGrazing(5, 1)
        ws.cell(i, 11).value = harvestDate
        ws.cell(i, 14).value = "04/01/2000"
        ws.cell(i, 15).value = "10/01/2000"
        ws.cell(i, 16).value = "21"

        print(dateLoopChoice)
        print(plantingDateChoice)
        print(harvestDate)
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()


        # add tillage

        def scrollTill():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id(
                "jqxScrollThumbverticalScrollBarinnerListBoxdropdownlisteditorjqxGridTillageTillage")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, 50).perform()
            actionChains.release()
            sleep(2)


        annualTillageList = ['No Tillage', 'Mow', 'Zero Soil Disturbance', 'No Implement Passes']
        annualTillageChoice = random.choice(annualTillageList)
        print(annualTillageChoice)
        browser.find_element_by_id('addTillage').click()
        sleep(2)
        browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
        browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
        browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
        sleep(2)
        clickTill = False
        while clickTill == False:
            try:
                browser.find_element_by_xpath('//span[text()="' + annualTillageChoice + '"]').click()
                clickTill = True
                break
            except:
                scrollTill()

        actionChains = ActionChains(browser)
        actionChains.move_to_element_with_offset(
            browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridTillageTillage"), -100,
            5).double_click().perform()


        def setAnnualTillDate(month):
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
                Keys.ARROW_LEFT)
            for m in range(month):
                browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridTillageShowDate"]').send_keys(
                    Keys.ARROW_UP)


        def setTill():
            browser.find_element_by_id('addTillage').click()
            browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
            browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
            browser.find_element_by_xpath('//div[text()="Intensive Tillage"]').click()
            sleep(2)
            browser.find_element_by_xpath('//span[text()="Mow"]').click()
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(
                browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridTillageTillage"), -100,
                5).double_click().perform()


        if annualTillageChoice == "Mow":
            setAnnualTillDate(4)
            ws.cell(i, 17).value = "Mow"
            ws.cell(i, 18).value = "05/01/2000"
            setTill()
            setAnnualTillDate(1)
            ws.cell(i, 19).value = "Mow"
            ws.cell(i, 20).value = "06/02/2000"
            setTill()
            setAnnualTillDate(2)
            ws.cell(i, 21).value = "Mow"
            ws.cell(i, 22).value = "08/03/2000"
            setTill()
            setAnnualTillDate(2)
            ws.cell(i, 23).value = "Mow"
            ws.cell(i, 24).value = "10/04/2000"
        else:
            setAnnualTillDate(4)
            ws.cell(i, 17).value = annualTillageChoice
            ws.cell(i, 18).value = "05/01/2000"

        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()


        def scrollFertilizer():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id(
                "jqxScrollThumbverticalScrollBarinnerListBoxdropdownlisteditorjqxGridNitrogenFertilizerType")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, 50).perform()
            actionChains.release()


        # add fertilizer
        seasonalFertilizerList = ['Ammonium Nitrate (34-0-0)', 'Ammonium Nitrate Phosphate (23-23-00)',
                                  'Ammonium Nitrate Phosphate (27-14-00)', 'Ammonium Phosphate Sulphate (16-20-00)',
                                  'Ammonium Polyphosphate Solution (10-34-00)', 'Ammonium Sulphate (21-00-00)',
                                  'Ammonium Thiosulphate Solution (12-00-00)', 'Anhydrous Ammonia (gas) (82-00-00)',
                                  'Calcium Ammonium Nitrate', 'Calcium Nitrate', 'Diammonium Phosphate (18-46-00)',
                                  'Element-N (N)', 'Element-P (P)', 'Mixed Blends',
                                  'Monoammonium Phosphate (11-55-00)',
                                  'Monoammonium Phosphate (12-51-00)', 'Potassium Nitrate', 'Urea (46-00-00)',
                                  'Urea Ammonium Nitrate (30-00-00)', 'Urea Ammonium Phosphate (27-27-00)',
                                  'Urea Ammonium Phosphate (34-17-00)']
        seasonalFertilizerChoice = random.choice(seasonalFertilizerList)
        ws.cell(i, 31).value = seasonalFertilizerChoice
        browser.find_element_by_id("addNitrogen").click()
        sleep(2)
        browser.find_element_by_xpath('//div[text()="Ammonium Nitrate (34-0-0)"]').click()
        browser.find_element_by_xpath('//div[text()="Ammonium Nitrate (34-0-0)"]').click()
        browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridNitrogenFertilizerType").click()
        click = False
        print(seasonalFertilizerChoice)
        while click == False:
            try:
                sleep(2)
                browser.find_element_by_xpath('//span[text()="' + seasonalFertilizerChoice + '"]').click()
                click = True
            except:
                scrollFertilizer()

        # Set Total N Applied
        totalNChoice = random.randrange(19, 151 + 1, 1)
        actionChains = ActionChains(browser)
        actionChains.move_to_element_with_offset(
            browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridNitrogenFertilizerType"), 450,
            5).double_click().perform()
        browser.find_element_by_id("textboxeditorjqxGridNitrogenTotalApplied").send_keys(totalNChoice)
        ws.cell(i, 33).value = totalNChoice


        def setAnnualFertilizerDate(month, day):
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(
                browser.find_element_by_id("textboxeditorjqxGridNitrogenTotalApplied"),
                -500, 5).double_click().perform()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').send_keys(
                Keys.ARROW_LEFT)
            for d in range(day):
                browser.find_element_by_xpath(
                    '//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').send_keys(
                    Keys.ARROW_UP)
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').send_keys(
                Keys.ARROW_LEFT)
            for m in range(month):
                browser.find_element_by_xpath(
                    '//input[@id="inputdatetimeeditorjqxGridNitrogenShowDate"]').send_keys(
                    Keys.ARROW_UP)


        # Setting date
        if plantingDateChoice == "05/01/2000":
            setAnnualFertilizerDate(4, 13)
            ws.cell(i, 32).value = "05/14/2000"
        if plantingDateChoice == "06/01/2000":
            setAnnualFertilizerDate(5, 13)
            ws.cell(i, 32).value = "06/15/2000"
        if plantingDateChoice == "07/01/2000":
            setAnnualFertilizerDate(6, 13)
            ws.cell(i, 32).value = "07/14/2000"

        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

        # add Manure
        seasonalManureList = ['Alfalfa Meal', 'Beef Manure, Solid', 'Beef Slurry', 'Bone Meal',
                              'Chicken - Broiler (litter), Solid', 'Chicken - Broiler Slurry',
                              'Chicken - Layer Slurry',
                              'Chicken - Layer, Solid', 'Compost or Composted Manure, Solid', 'Dairy Manure, Solid',
                              'Dairy Slurry', 'Farmyard Manure, Solid', 'Feather Meal', 'Feather Meal',
                              'Fish Emulsion',
                              'Fish Scrap', 'Guano', 'Horse Manure, Solid', 'Sheep Manure, Solid', 'Soybean Meal',
                              'Swine Manure, Slurry', 'Swine Manure, Solid']
        seasonalManureChoice = random.choice(seasonalManureList)
        ws.cell(i, 35).value = seasonalManureChoice

        print(seasonalManureChoice)
        browser.find_element_by_id("addManure").click()
        sleep(2)
        browser.find_element_by_xpath('//div[text()="Farmyard Manure, Solid"]').click()
        browser.find_element_by_xpath('//div[text()="Farmyard Manure, Solid"]').click()
        browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridManureManureType").click()
        sleep(2)


        def scrollManureDown():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id(
                "jqxScrollThumbverticalScrollBarinnerListBoxdropdownlisteditorjqxGridManureManureType")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, 50).perform()
            actionChains.release()
            sleep(2)


        def scrollManureUp():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id(
                "jqxScrollThumbverticalScrollBarinnerListBoxdropdownlisteditorjqxGridManureManureType")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, -50).perform()
            actionChains.release()
            sleep(2)


        click = False
        count = 0


        def chechManureOption():
            try:
                browser.find_element_by_xpath('//span[text()="' + seasonalManureChoice + '"]').click()
                return True
            except:
                return False


        while click == False:
            scrollManureUp()

            if chechManureOption() == True:
                break
            scrollManureDown()

            if chechManureOption() == True:
                break
            scrollManureDown()
            if chechManureOption() == True:
                break
            scrollManureDown()
            if chechManureOption() == True:
                break
            sleep(2)

        # amount applied
        browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridManureManureType").click()
        el = browser.find_element_by_id("dropdownlistContentdropdownlisteditorjqxGridManureManureType")
        actionChains = ActionChains(browser)
        actionChains.move_to_element_with_offset(el, 230, 5).double_click().perform()

        seasonalAmountAppliedChoice = random.randrange(5, 25 + 1, 1)
        browser.find_element_by_id("textboxeditorjqxGridManureActualTotalApplied").send_keys(
            seasonalAmountAppliedChoice)
        ws.cell(i, 36).value = seasonalAmountAppliedChoice


        def setAnnualManureDate(month, day):
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(
                browser.find_element_by_id("textboxeditorjqxGridManureActualTotalApplied"), -300,
                5).double_click().perform()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridManureShowDate"]').click()
            for m in range(day):
                browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridManureShowDate"]').send_keys(
                    Keys.ARROW_UP)
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridManureShowDate"]').send_keys(
                Keys.ARROW_LEFT)
            for d in range(month):
                browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridManureShowDate"]').send_keys(
                    Keys.ARROW_UP)


        if plantingDateChoice == "05/01/2000":
            setAnnualManureDate(3, 13)
            ws.cell(i, 34).value = "04/14/2000"
        if plantingDateChoice == "06/01/2000":
            setAnnualManureDate(4, 13)
            ws.cell(i, 34).value = "05/14/2000"
        if plantingDateChoice == "07/01/2000":
            setAnnualManureDate(5, 13)
            ws.cell(i, 34).value = "06/14/2000"

        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()


        # add Irrigation
        def scrollIrrBar():
            actionChains = ActionChains(browser)
            option = browser.find_element_by_id("jqxScrollThumbverticalScrollBarjqxGridIrrigation")
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, 200).perform()
            actionChains.release()
            sleep(2)


        def setIrr(scroll):
            browser.find_element_by_id("addIrrigation").click()
            try:
                sleep(2)
                browser.find_element_by_xpath('//div[text()="1"]').click()
            except:
                scrollIrrBar()
                sleep(1)
                browser.find_element_by_xpath('//div[text()="1"]').click()
            browser.find_element_by_xpath('//div[text()="1"]').click()
            irrChoice = 2.5
            browser.find_element_by_xpath('//input[@id="textboxeditorjqxGridIrrigationTotalApplied"]').send_keys(
                str(irrChoice))
            sleep(2)


        def setIrrDate(date, month, day, year):
            actionChains = ActionChains(browser)
            actionChains.move_to_element_with_offset(
                browser.find_element_by_xpath('//input[@id="textboxeditorjqxGridIrrigationTotalApplied"]'), -100,
                5).double_click().perform()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').click()
            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').click()
            if year == 1:
                browser.find_element_by_xpath(
                    '//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                    Keys.ARROW_UP)

            browser.find_element_by_xpath('//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                Keys.ARROW_LEFT)

            for i in range(day):
                browser.find_element_by_xpath(
                    '//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                    Keys.ARROW_UP)
            if month != 0:
                browser.find_element_by_xpath(
                    '//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                    Keys.ARROW_LEFT)
                for n in range(month):
                    browser.find_element_by_xpath(
                        '//input[@id="inputdatetimeeditorjqxGridIrrigationShowDate"]').send_keys(
                        Keys.ARROW_UP)
            sleep(2)


        def annualIrrForMay():
            setIrr(False)
            setIrrDate("5/2/2000", 0, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 39).value = "05/12/2000"
                ws.cell(i, 40).value = "2.5"
            setIrr(False)
            setIrrDate("5/13/2000", 0, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 41).value = "05/23/2000"
                ws.cell(i, 42).value = "2.5"
            setIrr(False)
            setIrrDate("5/24/2000", 1, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 43).value = "06/01/2000"
                ws.cell(i, 44).value = "2.5"


        def annualIrrForJune():
            setIrr(False)
            setIrrDate("6/4/2000", 0, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 45).value = "06/14/2000"
                ws.cell(i, 46).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                ws.cell(i, 39).value = "06/12/2000"
                ws.cell(i, 40).value = "2.5"
            setIrr(False)
            setIrrDate("6/15/2000", 0, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 47).value = "06/25/2000"
                ws.cell(i, 48).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                ws.cell(i, 41).value = "06/23/2000"
                ws.cell(i, 42).value = "2.5"
            setIrr(False)
            setIrrDate("6/26/2000", 1, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 49).value = "07/06/2000"
                ws.cell(i, 50).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                ws.cell(i, 43).value = "07/03/2000"
                ws.cell(i, 44).value = "2.5"


        def annualIrrForJuly():
            setIrr(False)
            setIrrDate("7/7/2000", 0, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 51).value = "07/17/2000"
                ws.cell(i, 52).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                ws.cell(i, 45).value = "07/14/2000"
                ws.cell(i, 46).value = "2.5"
            elif plantingDateChoice == '07/01/2000':
                ws.cell(i, 39).value = "07/12/2000"
                ws.cell(i, 40).value = "2.5"
            setIrr(False)
            setIrrDate("7/18/2000", 0, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 53).value = "07/28/2000"
                ws.cell(i, 54).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                ws.cell(i, 47).value = "07/25/2000"
                ws.cell(i, 48).value = "2.5"
            elif plantingDateChoice == '07/01/2000':
                ws.cell(i, 41).value = "07/23/2000"
                ws.cell(i, 42).value = "2.5"
            setIrr(False)
            setIrrDate("7/29/2000", 1, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 55).value = "08/08/2000"
                ws.cell(i, 56).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                ws.cell(i, 49).value = "08/05/2000"
                ws.cell(i, 50).value = "2.5"
            elif plantingDateChoice == '07/01/2000':
                ws.cell(i, 43).value = "08/03/2000"
                ws.cell(i, 44).value = "2.5"


        def annualIrrForAug():
            setIrr(False)
            setIrrDate("8/9/2000", 0, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 57).value = "08/19/2000"
                ws.cell(i, 58).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                ws.cell(i, 51).value = "08/16/2000"
                ws.cell(i, 52).value = "2.5"
            elif plantingDateChoice == '07/01/2000':
                ws.cell(i, 45).value = "08/14/2000"
                ws.cell(i, 46).value = "2.5"
            setIrr(False)
            setIrrDate("8/20/2000", 0, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 59).value = "08/30/2000"
                ws.cell(i, 60).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                ws.cell(i, 53).value = "08/27/2000"
                ws.cell(i, 54).value = "2.5"
            elif plantingDateChoice == '07/01/2000':
                ws.cell(i, 47).value = "08/25/2000"
                ws.cell(i, 48).value = "2.5"
            setIrr(True)
            setIrrDate("8/31/2000", 1, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 61).value = "09/10/2000"
                ws.cell(i, 62).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                ws.cell(i, 55).value = "09/07/2000"
                ws.cell(i, 56).value = "2.5"
            elif plantingDateChoice == '07/01/2000':
                ws.cell(i, 49).value = "09/05/2000"
                ws.cell(i, 50).value = "2.5"


        def annualIrrSep():
            setIrr(True)
            setIrrDate("6/4/2000", 0, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 63).value = "09/21/2000"
                ws.cell(i, 64).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                ws.cell(i, 57).value = "09/18/2000"
                ws.cell(i, 58).value = "2.5"
            elif plantingDateChoice == '07/01/2000':
                ws.cell(i, 51).value = "09/16/2000"
                ws.cell(i, 52).value = "2.5"

            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                setIrr(True)
                setIrrDate("6/15/2000", 1, 10, 0)
                ws.cell(i, 65).value = "10/2/2000"
                ws.cell(i, 66).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                setIrr(True)
                setIrrDate("6/15/2000", 0, 10, 0)
                ws.cell(i, 59).value = "09/29/2000"
                ws.cell(i, 60).value = "2.5"
            elif plantingDateChoice == '07/01/2000':
                setIrr(True)
                setIrrDate("6/15/2000", 0, 10, 0)
                ws.cell(i, 53).value = "09/27/2000"
                ws.cell(i, 54).value = "2.5"

            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                pass
            elif plantingDateChoice == '06/01/2000':
                setIrr(True)
                setIrrDate("6/26/2000", 1, 10, 0)
                ws.cell(i, 61).value = "10/10/2000"
                ws.cell(i, 62).value = "2.5"
            elif plantingDateChoice == '07/01/2000':
                setIrr(True)
                setIrrDate("6/26/2000", 1, 10, 0)
                ws.cell(i, 55).value = "10/08/2000"
                ws.cell(i, 56).value = "2.5"


        def annualIrrOct():
            setIrr(True)
            setIrrDate("6/4/2000", 0, 10, 0)
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                ws.cell(i, 67).value = "10/12/2000"
                ws.cell(i, 68).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                ws.cell(i, 63).value = "10/21/2000"
                ws.cell(i, 64).value = "2.5"
            elif plantingDateChoice == '07/01/2000':
                ws.cell(i, 57).value = "10/19/2000"
                ws.cell(i, 58).value = "2.5"

            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                setIrr(True)
                setIrrDate("6/15/2000", 0, 10, 0)
                ws.cell(i, 69).value = "10/23/2000"
                ws.cell(i, 70).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                setIrr(True)
                setIrrDate("6/15/2000", 1, 10, 0)
                ws.cell(i, 65).value = "11/01/2000"
                ws.cell(i, 66).value = "2.5"
            elif plantingDateChoice == '07/01/2000':
                setIrr(True)
                setIrrDate("6/15/2000", 0, 10, 0)
                ws.cell(i, 59).value = "10/30/2000"
                ws.cell(i, 60).value = "2.5"
            if plantingDateChoice == "05/01/2000" or plantingDateChoice == "04/01/2000":
                setIrr(True)
                setIrrDate("6/15/2000", 1, 10, 0)
                ws.cell(i, 69).value = "11/03/2000"
                ws.cell(i, 70).value = "2.5"
            elif plantingDateChoice == '06/01/2000':
                pass
            elif plantingDateChoice == '07/01/2000':
                setIrr(True)
                setIrrDate("6/15/2000", 1, 10, 0)
                ws.cell(i, 59).value = "11/10/2000"
                ws.cell(i, 60).value = "2.5"


        if plantingDateChoice == "05/01/2000":
            setIrr(False)
            setIrrDate("", 4, 0, 0)
            ws.cell(i, 37).value = "05/01/2000"
            ws.cell(i, 38).value = "2.5"
            if harvestDate == "07/01/2000":
                annualIrrForMay()
                annualIrrForJune()

            if harvestDate == "08/01/2000":
                annualIrrForMay()
                annualIrrForJune()
                annualIrrForJuly()
            if harvestDate == "09/01/2000":
                annualIrrForMay()
                annualIrrForJune()
                annualIrrForJuly()
                annualIrrForAug()
            if harvestDate == "10/01/2000":
                annualIrrForMay()
                annualIrrForJune()
                annualIrrForJuly()
                annualIrrForAug()
                annualIrrSep()
            elif harvestDate == "11/01/2000":
                annualIrrForMay()
                annualIrrForJune()
                annualIrrForJuly()
                annualIrrForAug()
                annualIrrSep()
                annualIrrOct()

        elif plantingDateChoice == '06/01/2000':
            setIrr(False)
            setIrrDate("", 5, 0, 0)
            ws.cell(i, 37).value = "06/01/2000"
            ws.cell(i, 38).value = "2.5"
            if harvestDate == "07/01/2000":
                annualIrrForJune()

            elif harvestDate == "08/01/2000":
                annualIrrForJune()
                annualIrrForJuly()
            elif harvestDate == "09/01/2000":
                annualIrrForJune()
                annualIrrForJuly()
                annualIrrForAug()
            elif harvestDate == "10/01/2000":
                annualIrrForJune()
                annualIrrForJuly()
                annualIrrForAug()
                annualIrrSep()
            elif harvestDate == "11/01/2000":
                annualIrrForJune()
                annualIrrForJuly()
                annualIrrForAug()
                annualIrrSep()
                annualIrrOct()

        elif plantingDateChoice == "07/01/2000":
            setIrr(False)
            setIrrDate("", 6, 0, 0)
            ws.cell(i, 37).value = "07/01/2000"
            ws.cell(i, 38).value = "2.5"

            if harvestDate == "08/01/2000":
                annualIrrForJuly()
            elif harvestDate == "09/01/2000":
                annualIrrForJuly()
                annualIrrForAug()
            elif harvestDate == "10/01/2000":
                annualIrrForJuly()
                annualIrrForAug()
                annualIrrSep()
            elif harvestDate == "11/01/2000":
                annualIrrForJuly()
                annualIrrForAug()
                annualIrrSep()
                annualIrrOct()

        else:
            setIrr(False)
            setIrrDate("", 4, 0, 0)
            ws.cell(i, 37).value = "05/01/2000"
            ws.cell(i, 38).value = "2.5"
            if harvestDate == "07/01/2000":
                annualIrrForMay()
                annualIrrForJune()
            if harvestDate == "08/01/2000":
                annualIrrForMay()
                annualIrrForJune()
                annualIrrForJuly()
            if harvestDate == "09/01/2000":
                annualIrrForMay()
                annualIrrForJune()
                annualIrrForJuly()
                annualIrrForAug()
            if harvestDate == "10/01/2000":
                annualIrrForMay()
                annualIrrForJune()
                annualIrrForJuly()
                annualIrrForAug()
                annualIrrSep()
            if harvestDate == "11/01/2000":
                annualIrrForMay()
                annualIrrForJune()
                annualIrrForJuly()
                annualIrrForAug()
                annualIrrSep()
                annualIrrOct()

        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

        # add lime
        ws.cell(i, 129).value = "No Liming"
        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

        # burning
        ws.cell(i, 130).value = "No Burning"
        # click next button
        browser.find_element_by_xpath('//span[text()="Next >>"]').click()

    browser.find_element_by_xpath('//span[text()="Skip Ahead >>"]').click()
    sleep(5)
    try:
        browser.find_element_by_xpath('//a[@class="nextStep"]').click()
    except:
        pass
    browser.find_element_by_xpath('//span[text()="Skip Ahead >>"]').click()
    try:
        browser.find_element_by_xpath('//a[@class="nextStep"]').click()
    except:
        pass

    browser.find_element_by_xpath('//span[text()="No Thanks, Continue >>"]').click()
    sleep(5)
    actions = ActionChains(browser)
    actions.move_to_element_with_offset(browser.find_element_by_xpath('//div[text()="select"]'), 5,
                                        5).click().perform()
    # add crop rotation for particular year\

    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2001"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2002"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2003"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2004"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2005"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2006"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2007"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2008"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2009"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2010"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2011"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2012"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2013"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2014"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2015"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2016"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2017"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2018"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2019"]').click()

    browser.find_element_by_xpath('//span[text()="Copy & Continue >>"]').click()
    sleep(10)

    # try:
    #     browser.find_element_by_xpath('//span[text()="Continue to Next Parcel >>"]').click()
    # except:
    #     pass

    # Future Management
    browser.find_element_by_xpath('//span[text()="Continue to Future Management >>"]').click()

    sleep(5)
    # name future management
    browser.find_element_by_xpath(
        '//input[@placeholder="for example, Reduced Tillage Practices Scenario"]').send_keys(
        "Management1")
    browser.find_element_by_xpath('//input[@aria-describedby="newScenarioCheck-errorEl"]').click()
    sleep(2)
    browser.find_element_by_xpath('//span[text()="Start >>"]').click()

    browser.find_element_by_xpath('//span[text()="Next >>"]').click()

    try:
        browser.find_element_by_xpath('//span[text()="Okay"]').click()
    except:
        print("aleerrt")
    try:
        browser.find_element_by_xpath('//a[@class="nextStep"]').click()
    except:
        print("error")
    browser.find_element_by_xpath('//span[text()="Next >>"]').click()

    browser.find_element_by_xpath('//span[text()="Skip Ahead >>"]').click()
    try:
        browser.find_element_by_xpath('//a[@class="nextStep"]').click()
    except:
        print("errrrr")
    sleep(3)
    # browser.find_element_by_xpath('//a[@class="nextStep"]').click()
    sleep(5)
    browser.find_element_by_xpath('//span[text()="Skip Ahead >>"]').click()
    sleep(2)
    browser.find_element_by_xpath('//span[text()="No Thanks, Continue >>"]').click()
    sleep(5)
    # add crop rotation for particular year\

    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2021"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2022"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2023"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2024"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2025"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2026"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2027"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2028"]').click()
    browser.find_element_by_xpath('//input[@id="copyRotation' + n + '2029"]').click()

    browser.find_element_by_xpath('//span[text()="Copy & Continue >>"]').click()
    sleep(5)
    sleep(5)
    sleep(5)

    browser.find_element_by_xpath('//span[text()="Continue to Animal Ag >>"]').click()

    # Enter ZipCode
    browser.find_element_by_id("ZipCode").send_keys("84321")
    ws.cell(i, 131).value = "84321"

    # select metric
    browser.find_element_by_xpath('//select[@id="UserUnits"]/option[text()="Metric"]').click()
    sleep(5)
    browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()

    # select animal type
    animalTypeList = ["Feedlot CattleCheck-bodyEl", "Beef-Heifer StockersCheck-bodyEl",
                      "Beef-Steer StockersCheck-bodyEl", "Dairy-Lactating CowsCheck-bodyEl"]


    def selectAnimalType(type):
        el = browser.find_element_by_id(type)
        actionChains = ActionChains(browser)
        actionChains.move_to_element_with_offset(el, 5, 5).click().perform()
        return type


    animalTypeCount = random.randrange(1, 3, 1)
    animalTypes = []
    col = 131
    for n in range(animalTypeCount):
        choice = random.choice(animalTypeList)
        selectAnimalType(choice)
        animalTypes.append(choice)
        if choice == "Feedlot CattleCheck-bodyEl":
            ws.cell(i, col).value = "Feedlot Cattle"
        if choice == "Beef-Heifer StockersCheck-bodyEl":
            ws.cell(i, col).value = "Beef-Heifer Stockers"
        if choice == "Beef-Steer StockersCheck-bodyEl":
            ws.cell(i, col).value = "Beef-Steer Stockers"
        if choice == "Dairy-Lactating CowsCheck-bodyEl":
            ws.cell(i, col).value = "Dairy-Lactating Cows"
        col = col + 1
        print(choice)
        animalTypeList.remove(choice)

    sleep(5)
    browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
    sleep(10)
    try:
        browser.find_element_by_xpath('//span[text()="Ok"]').click()
    except:
        pass


    # Feedolt and  steer

    def setAnimalCount(id, animalCount):
        animalOther = random.randrange(0, 10, 1)
        # either we want to deduct or add animal for next month
        animalOperationChoice = random.randrange(1, 3, 1)
        if animalOperationChoice == 1:
            animalCount = animalCount - animalOther
        else:
            animalCount = animalCount + animalOther
        el = browser.find_element_by_id(id)
        actionChains = ActionChains(browser)
        actionChains.move_to_element_with_offset(el, 10, 10).click().perform()
        actionChains.move_to_element_with_offset(el, 10, 10).send_keys(Keys.BACKSPACE).perform()
        actionChains.move_to_element_with_offset(el, 10, 10).send_keys(animalCount).perform()
        return animalCount


    def setAnimalWeight(id, animalWeight):
        animalOther = random.randrange(0, 10, 1)
        # either we want to deduct or add animal for next month
        animalOperationChoice = random.randrange(1, 3, 1)
        if animalOperationChoice == 1:
            animalWeight = animalWeight - animalOther
        else:
            animalWeight = animalWeight + animalOther
        el = browser.find_element_by_id(id)
        actionChains = ActionChains(browser)
        actionChains.move_to_element_with_offset(el, 10, 10).click().perform()
        actionChains.move_to_element_with_offset(el, 10, 10).send_keys(Keys.BACKSPACE).perform()
        actionChains.move_to_element_with_offset(el, 10, 10).send_keys(animalWeight).perform()
        return animalWeight


    def setAnimalCharacterstics(id, value):
        el = browser.find_element_by_id(id)
        actionChains = ActionChains(browser)
        actionChains.move_to_element_with_offset(el, 5, 5).click().perform()
        actionChains.move_to_element_with_offset(el, 5, 5).send_keys(Keys.BACKSPACE).perform()
        actionChains.move_to_element_with_offset(el, 5, 5).send_keys(value).perform()


    def setTyoesOfFeed(AnimalCategory):
        # step 1
        try:
            browser.find_element_by_id("gridcolumn-1089").click()
        except:
            try:
                browser.find_element_by_id("gridcolumn-1088").click()
            except:
                pass

        # step 2
        def scrollFeedDown():
            try:
                option = browser.find_element_by_id("gridscroller-1049_ct")
            except:
                option = browser.find_element_by_id("gridscroller-1050_ct")
            actionChains = ActionChains(browser)
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, 150).perform()
            actionChains.release().perform()

        def scrollFeedUp():
            try:
                option = browser.find_element_by_id("gridscroller-1049_ct")
            except:
                option = browser.find_element_by_id("gridscroller-1050_ct")
            actionChains = ActionChains(browser)
            actionChains.click_and_hold(option).perform()
            actionChains.move_by_offset(0, -50).perform()
            actionChains.release()
            sleep(5)

        find = False

        def findFeed(name, find):
            count = 0
            while find == False:
                try:
                    sleep(2)
                    browser.find_element_by_xpath('//div[text()="' + name + '"]').click()
                    find = True
                    break
                except:
                    if count % 10 == 0:
                        scrollFeedDown()
                    else:
                        scrollFeedUp()
                    print(count)
                    count = count + 1
                    sleep(2)

        feedList = ['Alfalfa', 'Birdsfoot', 'Bromegrass', 'Cheatgrass', 'Elephant Grass', 'Grain', 'Grass',
                    'Meadow',
                    'Oat',
                    'Orchardgrass', 'Prairie', 'Rye', 'Sanfoin', 'Sorghum', 'Sudangrass', 'Vetch', 'Wheat',
                    'Wheatgrass']
        feedChoice = random.choice(feedList)

        # try:
        #     browser.find_element_by_xpath('//input[@aria-describedby="textfield-1055-errorEl"]').send_keys(feedChoice)
        # except:
        #     browser.find_element_by_xpath('//input[@aria-describedby="textfield-1054-errorEl"]').send_keys(feedChoice)

        print(feedChoice)
        findFeed(feedChoice, find)

        sleep(2)
        if feedChoice == 'Alfalfa':
            choiceList = ['Cubes', 'Dehydrated 17% CP', 'Whole', 'Hay Early Bloom', 'Hay Mid Bloom',
                          'Hay Full Bloom',
                          'Hay Mature', 'Seed Screening', 'Silage', 'Silage Wilted', 'Leaf Meal', 'Stems']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Birdsfoot':
            choiceList = ['Trefoil Fresh', 'Trefoil Hay']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Bromegrass':
            choiceList = ['Fresh Immature', 'Hay', 'Haylage']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == "Cheatgrass":
            choiceList = ['Fresh Immature']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Elephant Grass':
            choiceList = ['Hay, Chopped']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Grain':
            choiceList = ['Screenings', 'Dust']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Grass':
            choiceList = ['Hay', 'Silage']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Meadow':
            choiceList = ['Hay']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Oat':
            choiceList = ['Silage', 'Straw', 'Hay', 'Grain', 'Grain, Steam Flaked', 'Groats', 'Middlings',
                          'Mill Byproduct', 'Hulls']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Orchardgrass':
            choiceList = ['Fresh Early Bloom', 'Hay']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Prairie':
            choiceList = ['Hay']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Rye':
            choiceList = ['Grass Hay', 'Grass Silage', 'Grain', 'Straw']
            choice = random.choice(choiceList)
            print(choice)
            sleep(2)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Sanfoin':
            choiceList = ['Hay']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Sorghum':
            choiceList = ['Stover', 'Silage', 'Grain (Milo), Ground', 'Grain (Milo), Flaked']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Sudangrass':
            choiceList = ['Fresh Immature', 'Hay', 'Silage']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Vetch':
            choiceList = ['Hay']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Wheat':
            choiceList = ['Fresh Pasture', 'Hay', 'Silage', 'Straw', 'Straw Ammoniated', 'Grain, Steam Flaked',
                          'Grain', 'Grain Hard', 'Grain Soft', 'Bran', 'Middlings', 'Mill Runs', 'Shorts']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        if feedChoice == 'Wheatgrass':
            choiceList = ['Crested Fresh Early Bloom', 'Crested Hay', 'Crested Fresh Full Bloom']
            choice = random.choice(choiceList)
            print(choice)
            browser.find_element_by_xpath('//div[text()="' + choice + '"]').click()

        browser.find_element_by_xpath('//span[text()="Add >>"]').click()
        browser.find_element_by_xpath('//span[text()="Confirm Selections"]').click()

        setAnimalCharacterstics("feedList0-bodyEl", 100)

        browser.find_element_by_xpath('//span[text()="Done - Save feed types for current selected date"]').click()
        sleep(3)
        browser.find_element_by_xpath('//span[text()="Ok"]').click()
        sleep(1)

        try:
            browser.find_element_by_xpath(
                '//span[text()="Done - Save feed types for current selected date"]').click()
            sleep(1)
            browser.find_element_by_xpath('//span[text()="Ok"]').click()
            sleep(1)
        except:
            pass
        if AnimalCategory == "Heifer":
            ws.cell(i, 151).value = feedChoice
            ws.cell(i, 152).value = "100"
        elif AnimalCategory == "Steer":
            ws.cell(i, 183).value = feedChoice
            ws.cell(i, 184).value = "100"
        elif AnimalCategory == "Dairy":
            ws.cell(i, 321).value = feedChoice
            ws.cell(i, 322).value = "100"

        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)


    def manureSystemTypesDairy(animalCategory):
        solidLiiquidSeperator = ["SolidLiquid_0-errorEl", "SolidLiquid_1-errorEl"]
        solidLiiquidSeperatorChoice = random.choice(solidLiiquidSeperator)
        browser.find_element_by_xpath('//input[@aria-describedby="' + solidLiiquidSeperatorChoice + '"]').click()
        print(solidLiiquidSeperatorChoice)
        if solidLiiquidSeperatorChoice == "SolidLiquid_1-errorEl":

            browser.find_element_by_xpath('//input[@aria-describedby="manureTreatment-errorEl"]').click()
            sleep(2)
            storageMethodList = ['Temporary Stack and Long-Term Stockpile', 'Composting', 'Aerobic lagoon',
                                 'Anaerobic Digester with Biogas Utilization or Methane Capture',
                                 'Thermochemical Conversion(Pyrolysis, Incineration, Gasification)',
                                 'Constructed wetland',
                                 'Daily Spread', 'Deposited on Pasture/Range/Paddock', 'Removed Offsite',
                                 'Land Applied']
            storageMethodChoice = random.choice(storageMethodList)
            browser.find_element_by_xpath('//input[@aria-describedby="manureTreatment-errorEl"]').click()
            browser.find_element_by_xpath('//input[@aria-describedby="manureTreatment-errorEl"]').click()

            try:
                list = browser.find_element_by_id("boundlist-1165-listEl")
            except:
                try:
                    list = browser.find_element_by_id("boundlist-1074-listEl")
                except:
                    try:
                        list = browser.find_element_by_id("boundlist-1144-listEl")
                    except:
                        list = browser.find_element_by_id("boundlist-1237-listEl")

            option = list.find_element_by_xpath('//li[text()="' + storageMethodChoice + '"]')
            actions = ActionChains(browser)
            actions.click(option)
            actions.perform()
            if animalCategory == "Dairy":
                ws.cell(i, 329).value = "No"
                ws.cell(i, 330).value = storageMethodChoice
        else:
            seperarorTypeList = ['Stationary inclined screen', 'Vibrating screen', 'Rotating screen',
                                 'In-channel flighted conveyor screen',
                                 'Belt press', 'Roller press', 'Screw press', 'Other']
            seperarorTypeChoice = random.choice(seperarorTypeList)
            browser.find_element_by_xpath('//input[@aria-describedby="separatorType-errorEl"]').click()
            browser.find_element_by_xpath('//li[text()="' + seperarorTypeChoice + '"]').click()

            percentage = random.randrange(0, 100, 1)
            browser.find_element_by_xpath('//input[@aria-describedby="solidRemoved-errorEl"]').clear()
            browser.find_element_by_xpath('//input[@aria-describedby="solidRemoved-errorEl"]').send_keys(percentage)

            solidTreatmentList = ['Temporary Stack and Long-Term Stockpile', 'Composting',
                                  'Thermochemical Conversion(Pyrolysis, Incineration, Gasification)',
                                  'Daily Spread',
                                  'Deposited on Pasture/Range/Paddock', 'Removed Offsite', 'Land Applied']
            solidTreatmentChoice = random.choice(solidTreatmentList)
            browser.find_element_by_xpath('//input[@aria-describedby="solidTreatment-errorEl"]').click()
            sleep(2)
            browser.find_element_by_xpath('//li[text()="' + solidTreatmentChoice + '"]').click()

            liquidTreatmentList = ['Aerobic lagoon',
                                   'Anaerobic Digester with Biogas Utilization or Methane Capture',
                                   'Thermochemical Conversion(Pyrolysis, Incineration, Gasification)',
                                   'Constructed wetland', 'Daily Spread', 'Deposited on Pasture/Range/Paddock',
                                   'Land Applied']
            liquidTreatmentChoice = random.choice(liquidTreatmentList)
            browser.find_element_by_xpath('//input[@aria-describedby="liquidTreatment-errorEl"]').click()
            sleep(2)
            print(liquidTreatmentChoice)
            try:
                browser.find_element_by_xpath(
                    '//div[@id="boundlist-1146-listEl"]/ul//li[text()="' + liquidTreatmentChoice + '"]').click()

            except:
                try:
                    browser.find_element_by_xpath(
                        '//div[@id="boundlist-1148-listEl"]/ul//li[text()="' + liquidTreatmentChoice + '"]').click()

                except:
                    try:
                        browser.find_element_by_xpath(
                            '//div[@id="boundlist-1078-listEl"]/ul//li[text()="' + liquidTreatmentChoice + '"]').click()
                    except:
                        try:
                            browser.find_element_by_xpath(
                                '//div[@id="boundlist-1224-listEl"]/ul//li[text()="' + liquidTreatmentChoice + '"]').click()
                        except:
                            try:
                                browser.find_element_by_xpath(
                                    '//div[@id="boundlist-1169-listEl"]/ul//li[text()="' + liquidTreatmentChoice + '"]').click()
                            except:
                                browser.find_element_by_xpath(
                                    '//div[@id="boundlist-1222-listEl"]/ul//li[text()="' + liquidTreatmentChoice + '"]').click()
            if animalCategory == "Dairy":
                ws.cell(i, 329).value = "Yes"
                ws.cell(i, 331).value = seperarorTypeChoice
                ws.cell(i, 332).value = percentage
                ws.cell(i, 333).value = solidTreatmentChoice
                ws.cell(i, 334).value = liquidTreatmentChoice

        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)
        if solidLiiquidSeperatorChoice == "SolidLiquid_0-errorEl":
            if solidTreatmentChoice == 'Temporary stack and long-term stockpile':
                tempList = ['Uncovered Solid', 'Covered Solid', 'Uncovered Semi-Solid', 'Covered Semi-Solid']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="systemCoverTypeSS-errorEl"]').click()
                sleep(2)
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()
                if animalCategory == "Dairy":
                    ws.cell(i, 335).value = tempChoice
                termList = ['Long Term (more than six months)', 'Short Term (less than six months)']
                termChoice = random.choice(termList)
                browser.find_element_by_xpath('//input[@aria-describedby="digesterType-errorEl"]').click()
                sleep(2)
                browser.find_element_by_xpath('//li[text()="' + termChoice + '"]').click()
                if animalCategory == "Dairy":
                    ws.cell(i, 336).value = termChoice

            if solidTreatmentChoice == 'Composting':
                tempList = ['In Vessel', 'Static Pile', 'Intensive Windrow', 'Passive Windrow']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="compostingMethod-errorEl"]').click()
                sleep(2)
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()
                if animalCategory == "Dairy":
                    ws.cell(i, 339).value = tempChoice

            if liquidTreatmentChoice == 'Thermochemical Conversion(Pyrolysis, Incineration, Gasification)' or liquidTreatmentChoice == 'Constructed wetland' or liquidTreatmentChoice == 'Daily Spread' or liquidTreatmentChoice == 'Deposited on Pasture/Range/Paddock' or liquidTreatmentChoice == 'Removed Offsite' or liquidTreatmentChoice == 'Land Applied':
                pass

            elif liquidTreatmentChoice == 'Aerobic lagoon':
                choice = random.randrange(400, 20000, 1)
                browser.find_element_by_xpath('//input[@aria-describedby="volumeLagoon-errorEl"]').clear()
                browser.find_element_by_xpath('//input[@aria-describedby="volumeLagoon-errorEl"]').send_keys(choice)
                tempList = ['Natural Aeration', 'Forced Aeration']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="aeratedNaturally-errorEl"]').click()
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()
                if animalCategory == "Dairy":
                    ws.cell(i, 337).value = choice
                    ws.cell(i, 338).value = tempChoice

            elif liquidTreatmentChoice == 'Anaerobic Digester with Biogas Utilization or Methane Capture':
                tempList = [
                    'Steel or lined concrete or fiberglass digesters with a gas holding system (egg shaped digesters) and monolithic construction',
                    'Up-flow Anaerobic Sludge Blanket (UASB) type with floating gas holders and no external seal',
                    'Unlined concrete/ferrocement/brick masonry arched type gas holding section and monolithic fixed dome digesters',
                    'Other']
                tempChoice = random.choice(tempList)
                print(tempChoice)
                browser.find_element_by_xpath('//input[@aria-describedby="digesterType-errorEl"]').click()
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()
                if animalCategory == "Dairy":
                    ws.cell(i, 340).value = tempChoice
        else:
            if storageMethodChoice == 'Temporary stack and long-term stockpile':
                tempList = ['Uncovered Solid', 'Covered Solid', 'Uncovered Semi-Solid', 'Covered Semi-Solid']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="systemCoverTypeSS-errorEl"]').click()
                sleep(2)
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()
                termList = ['Long Term (more than six months)', 'Short Term (less than six months)']
                termChoice = random.choice(termList)
                browser.find_element_by_xpath('//input[@aria-describedby="digesterType-errorEl"]').click()
                sleep(2)
                browser.find_element_by_xpath('//li[text()="' + termChoice + '"]').click()

            if storageMethodChoice == 'Composting':
                tempList = ['In Vessel', 'Static Pile', 'Intensive Windrow', 'Passive Windrow']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="compostingMethod-errorEl"]').click()
                sleep(2)
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()

            if storageMethodChoice == 'Thermochemical Conversion(Pyrolysis, Incineration, Gasification)' or storageMethodChoice == 'Constructed wetland' or storageMethodChoice == 'Daily Spread' or storageMethodChoice == 'Deposited on Pasture/Range/Paddock' or storageMethodChoice == 'Removed Offsite' or storageMethodChoice == 'Land Applied':
                pass

            elif storageMethodChoice == 'Aerobic lagoon':
                choice = random.randrange(400, 20000, 1)
                browser.find_element_by_xpath('//input[@aria-describedby="volumeLagoon-errorEl"]').clear()
                browser.find_element_by_xpath('//input[@aria-describedby="volumeLagoon-errorEl"]').send_keys(choice)
                tempList = ['Natural Aeration', 'Forced Aeration']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="aeratedNaturally-errorEl"]').click()
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()


            elif storageMethodChoice == 'Anaerobic Digester with Biogas Utilization or Methane Capture':
                tempList = [
                    'Steel or lined concrete or fiberglass digesters with a gas holding system (egg shaped digesters) and monolithic construction',
                    'Up-flow Anaerobic Sludge Blanket (UASB) type with floating gas holders and no external seal',
                    'Unlined concrete/ferrocement/brick masonry arched type gas holding section and monolithic fixed dome digesters',
                    'Other']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="digesterType-errorEl"]').click()
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()

        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()


    def manureSystemTypes(animalCategory):
        solidLiiquidSeperator = ["SolidLiquid_0-errorEl", "SolidLiquid_1-errorEl"]
        solidLiiquidSeperatorChoice = random.choice(solidLiiquidSeperator)
        browser.find_element_by_xpath('//input[@aria-describedby="' + solidLiiquidSeperatorChoice + '"]').click()
        print(solidLiiquidSeperatorChoice)

        if solidLiiquidSeperatorChoice == "SolidLiquid_1-errorEl":
            browser.find_element_by_xpath('//input[@aria-describedby="manureTreatment-errorEl"]').click()
            sleep(2)
            storageMethodList = ['Temporary Stack and Long-Term Stockpile', 'Composting', 'Aerobic lagoon',
                                 'Anaerobic Digester with Biogas Utilization or Methane Capture',
                                 'Thermochemical Conversion(Pyrolysis, Incineration, Gasification)',
                                 'Constructed wetland',
                                 'Daily Spread', 'Deposited on Pasture/Range/Paddock', 'Removed Offsite',
                                 'Land Applied']
            storageMethodChoice = random.choice(storageMethodList)
            browser.find_element_by_xpath('//input[@aria-describedby="manureTreatment-errorEl"]').click()
            browser.find_element_by_xpath('//input[@aria-describedby="manureTreatment-errorEl"]').click()

            try:
                list = browser.find_element_by_id("boundlist-1165-listEl")
            except:
                try:
                    list = browser.find_element_by_id("boundlist-1074-listEl")
                except:
                    try:
                        list = browser.find_element_by_id("boundlist-1144-listEl")
                    except:
                        try:
                            list = browser.find_element_by_id("boundlist-1237-listEl")
                        except:
                            try:
                                list = browser.find_element_by_id("boundlist-1222-listEl")
                            except:
                                try:
                                    list = browser.find_element_by_id("boundlist-1220-listEl")
                                except:
                                    list = browser.find_element_by_id("boundlist-1216-listEl")


            option = list.find_element_by_xpath('//li[text()="' + storageMethodChoice + '"]')
            actions = ActionChains(browser)
            actions.click(option)
            actions.perform()
            if animalCategory == "Heifer":
                ws.cell(i, 154).value = "No"
                ws.cell(i, 155).value = storageMethodChoice
            elif animalCategory == "Steer":
                ws.cell(i, 186).value = "No"
                ws.cell(i, 187).value = storageMethodChoice
            elif animalCategory == "Feedlot":
                ws.cell(i, 262).value = "No"
                ws.cell(i, 263).value = storageMethodChoice


        else:
            seperarorTypeList = ['Stationary inclined screen', 'Vibrating screen', 'Rotating screen', 'Centrifuge',
                                 'Decanter centrifuge', 'Roller press', 'Pressure filter', 'Other']
            seperarorTypeChoice = random.choice(seperarorTypeList)
            browser.find_element_by_xpath('//input[@aria-describedby="separatorType-errorEl"]').click()
            browser.find_element_by_xpath('//li[text()="' + seperarorTypeChoice + '"]').click()

            percentage = random.randrange(0, 100, 1)
            browser.find_element_by_xpath('//input[@aria-describedby="solidRemoved-errorEl"]').clear()
            browser.find_element_by_xpath('//input[@aria-describedby="solidRemoved-errorEl"]').send_keys(percentage)

            solidTreatmentList = ['Temporary Stack and Long-Term Stockpile', 'Composting',
                                  'Thermochemical Conversion(Pyrolysis, Incineration, Gasification)',
                                  'Daily Spread',
                                  'Deposited on Pasture/Range/Paddock', 'Removed Offsite', 'Land Applied']
            solidTreatmentChoice = random.choice(solidTreatmentList)
            browser.find_element_by_xpath('//input[@aria-describedby="solidTreatment-errorEl"]').click()
            sleep(2)
            browser.find_element_by_xpath('//li[text()="' + solidTreatmentChoice + '"]').click()
            if animalCategory == "Feedlot":
                liquidTreatmentList = ['Aerobic lagoon',
                                       'Anaerobic Digester with Biogas Utilization or Methane Capture',
                                       'Thermochemical Conversion(Pyrolysis, Incineration, Gasification)',
                                       'Constructed wetland', 'Daily Spread', 'Deposited on Pasture/Range/Paddock',
                                       'Land Applied', 'Removed Offsite',
                                       'Anaerobic lagoon, liquid/slurry storage pond, storage tanks',
                                       'Combined Aerobic treatment system']
            else:
                liquidTreatmentList = ['Aerobic lagoon',
                                       'Anaerobic Digester with Biogas Utilization or Methane Capture',
                                       'Thermochemical Conversion(Pyrolysis, Incineration, Gasification)',
                                       'Constructed wetland', 'Daily Spread', 'Deposited on Pasture/Range/Paddock',
                                       'Land Applied', 'Removed Offsite']
            liquidTreatmentChoice = random.choice(liquidTreatmentList)
            browser.find_element_by_xpath('//input[@aria-describedby="liquidTreatment-errorEl"]').click()
            sleep(2)
            print(liquidTreatmentChoice)
            try:
                browser.find_element_by_xpath(
                    '//div[@id="boundlist-1146-listEl"]/ul//li[text()="' + liquidTreatmentChoice + '"]').click()

            except:
                try:
                    browser.find_element_by_xpath(
                        '//div[@id="boundlist-1148-listEl"]/ul//li[text()="' + liquidTreatmentChoice + '"]').click()

                except:
                    try:
                        browser.find_element_by_xpath(
                            '//div[@id="boundlist-1078-listEl"]/ul//li[text()="' + liquidTreatmentChoice + '"]').click()
                    except:
                        try:
                            browser.find_element_by_xpath(
                                '//div[@id="boundlist-1224-listEl"]/ul//li[text()="' + liquidTreatmentChoice + '"]').click()
                        except:
                            try:
                                browser.find_element_by_xpath(
                                    '//div[@id="boundlist-1169-listEl"]/ul//li[text()="' + liquidTreatmentChoice + '"]').click()
                            except:
                                browser.find_element_by_xpath(
                                    '//div[@id="boundlist-1222-listEl"]/ul//li[text()="' + liquidTreatmentChoice + '"]').click()
            if animalCategory == "Heifer":
                ws.cell(i, 154).value = "Yes"
                ws.cell(i, 156).value = seperarorTypeChoice
                ws.cell(i, 157).value = percentage
                ws.cell(i, 158).value = solidTreatmentChoice
                ws.cell(i, 159).value = liquidTreatmentChoice
            elif animalCategory == "Steer":
                ws.cell(i, 186).value = "Yes"
                ws.cell(i, 188).value = seperarorTypeChoice
                ws.cell(i, 189).value = percentage
                ws.cell(i, 190).value = solidTreatmentChoice
                ws.cell(i, 191).value = liquidTreatmentChoice
            elif animalCategory == "Feedlot":
                ws.cell(i, 262).value = "Yes"
                ws.cell(i, 264).value = seperarorTypeChoice
                ws.cell(i, 265).value = percentage
                ws.cell(i, 266).value = solidTreatmentChoice
                ws.cell(i, 267).value = liquidTreatmentChoice

        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()


        sleep(5)
        if solidLiiquidSeperatorChoice == "SolidLiquid_0-errorEl":
            if solidTreatmentChoice == 'Temporary stack and long-term stockpile':
                tempList = ['Uncovered Solid', 'Covered Solid', 'Uncovered Semi-Solid', 'Covered Semi-Solid']
                tempChoice = random.choice(tempList)
                if animalCategory == "Heifer":
                    ws.cell(i, 160).value = tempChoice
                elif animalCategory == "Steer":
                    ws.cell(i, 192).value = tempChoice
                elif animalCategory == "Feedlot":
                    ws.cell(i, 268).value = tempChoice

                browser.find_element_by_xpath('//input[@aria-describedby="systemCoverTypeSS-errorEl"]').click()
                sleep(2)
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()
                termList = ['Long Term (more than six months)', 'Short Term (less than six months)']
                termChoice = random.choice(termList)
                browser.find_element_by_xpath('//input[@aria-describedby="digesterType-errorEl"]').click()
                sleep(2)
                browser.find_element_by_xpath('//li[text()="' + termChoice + '"]').click()
                if animalCategory == "Heifer":
                    ws.cell(i, 161).value = termChoice
                if animalCategory == "Steer":
                    ws.cell(i, 193).value = termChoice
                elif animalCategory == "Feedlot":
                    ws.cell(i, 269).value = termChoice
                elif animalCategory == "Dairy":
                    ws.cell(i, 192).value = termChoice

            if solidTreatmentChoice == 'Composting':
                tempList = ['In Vessel', 'Static Pile', 'Intensive Windrow', 'Passive Windrow']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="compostingMethod-errorEl"]').click()
                sleep(2)
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()
                if animalCategory == "Heifer":
                    ws.cell(i, 164).value = tempChoice
                elif animalCategory == "Steer":
                    ws.cell(i, 196).value = tempChoice
                elif animalCategory == "Feedlot":
                    ws.cell(i, 272).value = tempChoice

            if liquidTreatmentChoice == 'Thermochemical Conversion(Pyrolysis, Incineration, Gasification)' or liquidTreatmentChoice == 'Constructed wetland' or liquidTreatmentChoice == 'Daily Spread' or liquidTreatmentChoice == 'Deposited on Pasture/Range/Paddock' or liquidTreatmentChoice == 'Removed Offsite' or liquidTreatmentChoice == 'Land Applied':
                pass

            elif liquidTreatmentChoice == 'Aerobic lagoon':
                choice = random.randrange(400, 20000, 1)
                browser.find_element_by_xpath('//input[@aria-describedby="volumeLagoon-errorEl"]').clear()
                browser.find_element_by_xpath('//input[@aria-describedby="volumeLagoon-errorEl"]').send_keys(choice)
                tempList = ['Natural Aeration', 'Forced Aeration']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="aeratedNaturally-errorEl"]').click()
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()
                if animalCategory == "Heifer":
                    ws.cell(i, 162).value = choice
                    ws.cell(i, 163).value = tempChoice
                if animalCategory == "Steer":
                    ws.cell(i, 194).value = choice
                    ws.cell(i, 195).value = tempChoice
                if animalCategory == "Feedlot":
                    ws.cell(i, 270).value = choice
                    ws.cell(i, 271).value = tempChoice



            elif liquidTreatmentChoice == 'Anaerobic Digester with Biogas Utilization or Methane Capture':
                tempList = [
                    'Steel or lined concrete or fiberglass digesters with a gas holding system (egg shaped digesters) and monolithic construction',
                    'Up-flow Anaerobic Sludge Blanket (UASB) type with floating gas holders and no external seal',
                    'Unlined concrete/ferrocement/brick masonry arched type gas holding section and monolithic fixed dome digesters',
                    'Other']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="digesterType-errorEl"]').click()
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()
                if animalCategory == "Heifer":
                    ws.cell(i, 165).value = tempChoice
                if animalCategory == "Steer":
                    ws.cell(i, 197).value = tempChoice
                if animalCategory == "Feedlot":
                    ws.cell(i, 273).value = tempChoice

        else:
            if storageMethodChoice == 'Temporary stack and long-term stockpile':
                tempList = ['Uncovered Solid', 'Covered Solid', 'Uncovered Semi-Solid', 'Covered Semi-Solid']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="systemCoverTypeSS-errorEl"]').click()
                sleep(2)
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()
                termList = ['Long Term (more than six months)', 'Short Term (less than six months)']
                termChoice = random.choice(termList)
                browser.find_element_by_xpath('//input[@aria-describedby="digesterType-errorEl"]').click()
                sleep(2)
                browser.find_element_by_xpath('//li[text()="' + termChoice + '"]').click()
                if animalCategory == "Heifer":
                    ws.cell(i, 160).value = tempChoice
                if animalCategory == "Steer":
                    ws.cell(i, 193).value = tempChoice
                elif animalCategory == "Feedlot":
                    ws.cell(i, 268).value = tempChoice
                elif animalCategory == "Dairy":
                    ws.cell(i, 192).value = tempChoice

            if storageMethodChoice == 'Composting':
                tempList = ['In Vessel', 'Static Pile', 'Intensive Windrow', 'Passive Windrow']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="compostingMethod-errorEl"]').click()
                sleep(2)
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()
                if animalCategory == "Heifer":
                    ws.cell(i, 164).value = tempChoice
                elif animalCategory == "Steer":
                    ws.cell(i, 196).value = tempChoice
                elif animalCategory == "Feedlot":
                    ws.cell(i, 272).value = tempChoice
                elif animalCategory == "Dairy":
                    ws.cell(i, 196).value = tempChoice

            if storageMethodChoice == 'Thermochemical Conversion(Pyrolysis, Incineration, Gasification)' or storageMethodChoice == 'Constructed wetland' or storageMethodChoice == 'Daily Spread' or storageMethodChoice == 'Deposited on Pasture/Range/Paddock' or storageMethodChoice == 'Removed Offsite' or storageMethodChoice == 'Land Applied':
                pass

            elif storageMethodChoice == 'Aerobic lagoon':
                choice = random.randrange(400, 20000, 1)
                browser.find_element_by_xpath('//input[@aria-describedby="volumeLagoon-errorEl"]').clear()
                browser.find_element_by_xpath('//input[@aria-describedby="volumeLagoon-errorEl"]').send_keys(choice)
                tempList = ['Natural Aeration', 'Forced Aeration']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="aeratedNaturally-errorEl"]').click()
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()
                if animalCategory == "Heifer":
                    ws.cell(i, 162).value = choice
                    ws.cell(i, 163).value = tempChoice
                if animalCategory == "Steer":
                    ws.cell(i, 194).value = choice
                    ws.cell(i, 195).value = tempChoice
                if animalCategory == "Feedlot":
                    ws.cell(i, 270).value = choice
                    ws.cell(i, 271).value = tempChoice
                if animalCategory == "Dairy":
                    ws.cell(i, 194).value = choice
                    ws.cell(i, 195).value = tempChoice


            elif storageMethodChoice == 'Anaerobic Digester with Biogas Utilization or Methane Capture':
                tempList = [
                    'Steel or lined concrete or fiberglass digesters with a gas holding system (egg shaped digesters) and monolithic construction',
                    'Up-flow Anaerobic Sludge Blanket (UASB) type with floating gas holders and no external seal',
                    'Unlined concrete/ferrocement/brick masonry arched type gas holding section and monolithic fixed dome digesters',
                    'Other']
                tempChoice = random.choice(tempList)
                browser.find_element_by_xpath('//input[@aria-describedby="digesterType-errorEl"]').click()
                browser.find_element_by_xpath('//li[text()="' + tempChoice + '"]').click()
                if animalCategory == "Heifer":
                    ws.cell(i, 165).value = tempChoice
                if animalCategory == "Steer":
                    ws.cell(i, 197).value = tempChoice
                if animalCategory == "Feedlot":
                    ws.cell(i, 273).value = tempChoice
                if animalCategory == "Dairy":
                    ws.cell(i, 197).value = tempChoice

        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()


    def animalFutureManagement():
        sleep(5)
        browser.find_element_by_xpath(
            '//input[@placeholder="for example, Reduced Feed Practices Scenario"]').send_keys(
            "AnimalManagement1")
        browser.find_element_by_xpath('//span[text()="Create new Scenario >>"]').click()
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)
        browser.find_element_by_xpath('//span[text()="Continue to Report >>"]').click()


    def settingAllAnimalCharactersticsForHeifers():
        sleep(2)
        browser.find_element_by_id("herdNames1").clear()
        browser.find_element_by_id("herdNames1").send_keys("Herd 1")
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)

        animalCount = random.randrange(1, 501, 1)
        animalCount = setAnimalCount("JanPopulation-bodyEl", animalCount)
        ws.cell(i, 136).value = animalCount
        # Feb
        animalCount = setAnimalCount("FebPopulation-bodyEl", animalCount)
        ws.cell(i, 137).value = animalCount
        # Mar
        animalCount = setAnimalCount("MarPopulation-bodyEl", animalCount)
        ws.cell(i, 138).value = animalCount
        # Apr
        animalCount = setAnimalCount("AprPopulation-bodyEl", animalCount)
        ws.cell(i, 139).value = animalCount
        # May
        animalCount = setAnimalCount("MayPopulation-bodyEl", animalCount)
        ws.cell(i, 140).value = animalCount
        # Jun
        animalCount = setAnimalCount("JunPopulation-bodyEl", animalCount)
        ws.cell(i, 141).value = animalCount
        # Jul
        animalCount = setAnimalCount("JulPopulation-bodyEl", animalCount)
        ws.cell(i, 142).value = animalCount
        # Aug
        animalCount = setAnimalCount("AugPopulation-bodyEl", animalCount)
        ws.cell(i, 143).value = animalCount
        # Sep
        animalCount = setAnimalCount("SepPopulation-bodyEl", animalCount)
        ws.cell(i, 144).value = animalCount
        # Oct
        animalCount = setAnimalCount("OctPopulation-bodyEl", animalCount)
        ws.cell(i, 145).value = animalCount
        # Nov
        animalCount = setAnimalCount("NovPopulation-bodyEl", animalCount)
        ws.cell(i, 146).value = animalCount
        # Dec
        animalCount = setAnimalCount("DecPopulation-bodyEl", animalCount)
        ws.cell(i, 147).value = animalCount

        sleep(5)
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)

        # set Average weight
        setAnimalCharacterstics("bodyWeight-bodyEl", 362)
        ws.cell(i, 148).value = "362"

        # average daily weight gain
        weigthGain = round(random.uniform(0.6, 2.5), 1)
        setAnimalCharacterstics("dailyWeightGain-bodyEl", str(weigthGain))
        ws.cell(i, 149).value = weigthGain

        # average Mature weigth
        setAnimalCharacterstics("matureWeight-bodyEl", 545)
        ws.cell(i, 150).value = "545"

        sleep(2)
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)

        # number of hours
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)

        setTyoesOfFeed("Heifer")

        feedingSituationList = ['Stall', 'Pasture', 'Grazing Large Areas']
        feedingSituationChoice = random.choice(feedingSituationList)
        browser.find_element_by_xpath('//input[@aria-describedby="feedingSituation-errorEl"]').clear()
        browser.find_element_by_xpath('//input[@aria-describedby="feedingSituation-errorEl"]').send_keys(
            feedingSituationChoice)
        ws.cell(i, 153).value = feedingSituationChoice

        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)

        manureSystemTypes("Heifer")


    def settingAllAnimalCharactersticsForSteers():
        sleep(2)
        browser.find_element_by_id("herdNames1").clear()
        browser.find_element_by_id("herdNames1").send_keys("Herd 1")
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)

        animalCount = random.randrange(1, 501, 1)
        animalCount = setAnimalCount("JanPopulation-bodyEl", animalCount)
        ws.cell(i, 168).value = animalCount
        # Feb
        animalCount = setAnimalCount("FebPopulation-bodyEl", animalCount)
        ws.cell(i, 169).value = animalCount
        # Mar
        animalCount = setAnimalCount("MarPopulation-bodyEl", animalCount)
        ws.cell(i, 170).value = animalCount
        # Apr
        animalCount = setAnimalCount("AprPopulation-bodyEl", animalCount)
        ws.cell(i, 171).value = animalCount
        # May
        animalCount = setAnimalCount("MayPopulation-bodyEl", animalCount)
        ws.cell(i, 172).value = animalCount
        # Jun
        animalCount = setAnimalCount("JunPopulation-bodyEl", animalCount)
        ws.cell(i, 173).value = animalCount
        # Jul
        animalCount = setAnimalCount("JulPopulation-bodyEl", animalCount)
        ws.cell(i, 174).value = animalCount
        # Aug
        animalCount = setAnimalCount("AugPopulation-bodyEl", animalCount)
        ws.cell(i, 175).value = animalCount
        # Sep
        animalCount = setAnimalCount("SepPopulation-bodyEl", animalCount)
        ws.cell(i, 176).value = animalCount
        # Oct
        animalCount = setAnimalCount("OctPopulation-bodyEl", animalCount)
        ws.cell(i, 177).value = animalCount
        # Nov
        animalCount = setAnimalCount("NovPopulation-bodyEl", animalCount)
        ws.cell(i, 178).value = animalCount
        # Dec
        animalCount = setAnimalCount("DecPopulation-bodyEl", animalCount)
        ws.cell(i, 179).value = animalCount

        sleep(5)
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)

        # set Average weight
        setAnimalCharacterstics("bodyWeight-bodyEl", 362)
        ws.cell(i, 180).value = "362"

        # average daily weight gain
        weigthGain = round(random.uniform(0.6, 2.5), 1)
        setAnimalCharacterstics("dailyWeightGain-bodyEl", str(weigthGain))
        ws.cell(i, 181).value = weigthGain

        # average Mature weigth
        setAnimalCharacterstics("matureWeight-bodyEl", 545)
        sleep(2)
        ws.cell(i, 182).value = "545"
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)

        # number of hours
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)

        setTyoesOfFeed("Steer")

        feedingSituationList = ['Stall', 'Pasture', 'Grazing Large Areas']
        feedingSituationChoice = random.choice(feedingSituationList)
        browser.find_element_by_xpath('//input[@aria-describedby="feedingSituation-errorEl"]').clear()
        browser.find_element_by_xpath('//input[@aria-describedby="feedingSituation-errorEl"]').send_keys(
            feedingSituationChoice)
        ws.cell(i, 185).value = feedingSituationChoice

        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)

        manureSystemTypes("Steer")


    def settingAllAnimalCharactersticsForLactatingCow():
        sleep(2)
        browser.find_element_by_id("herdNames1").clear()
        browser.find_element_by_id("herdNames1").send_keys("Herd 1")
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)
        # Feedlot and lactating
        roofed = random.randrange(1, 501, 1)
        roofed = setAnimalCount("Janroofedfacility-bodyEl", roofed)
        ws.cell(i, 281).value = roofed
        roofed = setAnimalCount("Febroofedfacility-bodyEl", roofed)
        ws.cell(i, 282).value = roofed
        roofed = setAnimalCount("Marroofedfacility-bodyEl", roofed)
        ws.cell(i, 283).value = roofed
        roofed = setAnimalCount("Aprroofedfacility-bodyEl", roofed)
        ws.cell(i, 284).value = roofed
        roofed = setAnimalCount("Mayroofedfacility-bodyEl", roofed)
        ws.cell(i, 285).value = roofed
        roofed = setAnimalCount("Junroofedfacility-bodyEl", roofed)
        ws.cell(i, 286).value = roofed
        roofed = setAnimalCount("Julroofedfacility-bodyEl", roofed)
        ws.cell(i, 287).value = roofed
        roofed = setAnimalCount("Augroofedfacility-bodyEl", roofed)
        ws.cell(i, 288).value = roofed
        roofed = setAnimalCount("Seproofedfacility-bodyEl", roofed)
        ws.cell(i, 289).value = roofed
        roofed = setAnimalCount("Octroofedfacility-bodyEl", roofed)
        ws.cell(i, 290).value = roofed
        roofed = setAnimalCount("Novroofedfacility-bodyEl", roofed)
        ws.cell(i, 291).value = roofed
        roofed = setAnimalCount("Decroofedfacility-bodyEl", roofed)
        ws.cell(i, 292).value = roofed

        # dry loy
        drylot = random.randrange(1, 501, 1)
        drylot = setAnimalCount("Jandrylot-bodyEl", drylot)
        ws.cell(i, 293).value = drylot
        drylot = setAnimalCount("Febdrylot-bodyEl", drylot)
        ws.cell(i, 294).value = drylot
        drylot = setAnimalCount("Mardrylot-bodyEl", drylot)
        ws.cell(i, 295).value = drylot
        drylot = setAnimalCount("Aprdrylot-bodyEl", drylot)
        ws.cell(i, 296).value = drylot
        drylot = setAnimalCount("Maydrylot-bodyEl", drylot)
        ws.cell(i, 297).value = drylot
        drylot = setAnimalCount("Jundrylot-bodyEl", drylot)
        ws.cell(i, 298).value = drylot
        drylot = setAnimalCount("Juldrylot-bodyEl", drylot)
        ws.cell(i, 299).value = drylot
        drylot = setAnimalCount("Augdrylot-bodyEl", drylot)
        ws.cell(i, 300).value = drylot
        drylot = setAnimalCount("Sepdrylot-bodyEl", drylot)
        ws.cell(i, 301).value = drylot
        drylot = setAnimalCount("Octdrylot-bodyEl", drylot)
        ws.cell(i, 302).value = drylot
        drylot = setAnimalCount("Novdrylot-bodyEl", drylot)
        ws.cell(i, 303).value = drylot
        drylot = setAnimalCount("Decdrylot-bodyEl", drylot)
        ws.cell(i, 304).value = drylot

        # pasture range
        pastureRange = random.randrange(1, 501, 1)
        pastureRange = setAnimalCount("Janpasturerangepaddock-bodyEl", pastureRange)
        ws.cell(i, 305).value = pastureRange
        pastureRange = setAnimalCount("Febpasturerangepaddock-bodyEl", pastureRange)
        ws.cell(i, 306).value = pastureRange
        pastureRange = setAnimalCount("Marpasturerangepaddock-bodyEl", pastureRange)
        ws.cell(i, 307).value = pastureRange
        pastureRange = setAnimalCount("Aprpasturerangepaddock-bodyEl", pastureRange)
        ws.cell(i, 308).value = pastureRange
        pastureRange = setAnimalCount("Maypasturerangepaddock-bodyEl", pastureRange)
        ws.cell(i, 309).value = pastureRange
        pastureRange = setAnimalCount("Junpasturerangepaddock-bodyEl", pastureRange)
        ws.cell(i, 310).value = pastureRange
        pastureRange = setAnimalCount("Julpasturerangepaddock-bodyEl", pastureRange)
        ws.cell(i, 311).value = pastureRange
        pastureRange = setAnimalCount("Augpasturerangepaddock-bodyEl", pastureRange)
        ws.cell(i, 312).value = pastureRange
        pastureRange = setAnimalCount("Seppasturerangepaddock-bodyEl", pastureRange)
        ws.cell(i, 313).value = pastureRange
        pastureRange = setAnimalCount("Octpasturerangepaddock-bodyEl", pastureRange)
        ws.cell(i, 315).value = pastureRange
        pastureRange = setAnimalCount("Decpasturerangepaddock-bodyEl", pastureRange)
        ws.cell(i, 316).value = pastureRange

        # daily feed intake
        setAnimalCharacterstics("dailyFeedIntake-bodyEl", 0)
        ws.cell(i, 317).value = "0"
        # live body weight
        liveWeight = random.randrange(362, 545, 1)
        setAnimalCharacterstics("avgLiveWeight-bodyEl", liveWeight)
        ws.cell(i, 318).value = liveWeight
        # days in milk
        setAnimalCharacterstics("daysInMilk-bodyEl", 305)
        ws.cell(i, 319).value = "305"

        # milk production
        setAnimalCharacterstics("milkProduction-bodyEl", str(28.6))
        ws.cell(i, 320).value = "28.6"

        try:
            browser.find_element_by_xpath('//span[text()="Ok"]').click()
            sleep(2)
        except:
            pass

        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)

        try:
            browser.find_element_by_xpath('//span[text()="Ok"]').click()
            sleep(2)
            browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
            sleep(5)
        except:
            pass

        setTyoesOfFeed("Dairy")

        housingTypeList = ['manurestorage_12-errorEl', 'manurestorage_11-errorEl', 'manurestorage_10-errorEl']
        housingTypeChoice = random.choice(housingTypeList)
        browser.find_element_by_xpath('//input[@aria-describedby="' + housingTypeChoice + '"]').click()
        if housingTypeChoice == 'manurestorage_10-errorEl':
            ws.cell(i, 323).value = "Pit Storage"
            pitStorageList = ['pitstorage_10-errorEl', 'pitstorage_11-errorEl']
            pitStorageChoice = random.choice(pitStorageList)
            if pitStorageChoice == "pitstorage_10-errorEl":
                ws.cell(i, 324).value = "Deep"
            else:
                ws.cell(i, 324).value = "Shallow"
            browser.find_element_by_xpath('//input[@aria-describedby="' + pitStorageChoice + '"]').click()
            days = random.randrange(1, 7, 1)
            ws.cell(i, 325).value = days
            browser.find_element_by_xpath('//input[@aria-describedby="NumDaysInHousing1-errorEl"]').send_keys(days)

        elif housingTypeChoice == 'manurestorage_11-errorEl':
            ws.cell(i, 323).value = "Bedded Pack"
            beddedPackList = ['bedpack_10-errorEl', 'bedpack_11-errorEl']
            beddedPackChoice = random.choice(beddedPackList)
            browser.find_element_by_xpath('//input[@aria-describedby="' + beddedPackChoice + '"]').click()
            if beddedPackChoice == 'bedpack_10-errorEl':
                ws.cell(i, 326).value = "Active Mix"
            else:
                ws.cell(i, 326).value = "No Mix"
            days = random.randrange(1, 7, 1)
            browser.find_element_by_xpath('//input[@aria-describedby="NumDaysInHousing1-errorEl"]').send_keys(days)
            ws.cell(i, 327).value = days

        elif housingTypeChoice == 'manurestorage_12-errorEl':
            ws.cell(i, 323).value = "Flushed Or Scrapped"
            areaOfBarn = round(random.uniform(float(0.4), float(2.8)), 1)
            browser.find_element_by_xpath('//input[@aria-describedby="AreaOfBarn-errorEl"]').send_keys(
                str(areaOfBarn))
            ws.cell(i, 328).value = areaOfBarn

        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)

        manureSystemTypesDairy("Dairy")


    def settingAllAnimalCharactersticsForFeedlot():
        sleep(2)
        browser.find_element_by_id("herdNames1").clear()
        browser.find_element_by_id("herdNames1").send_keys("Herd 1")
        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)
        # only feedolt

        # heifers
        heifers = random.randrange(0, 500, 1)
        heifers = setAnimalCount("Janheifers-bodyEl", heifers)
        ws.cell(i, 200).value = heifers
        heifers = setAnimalCount("Febheifers-bodyEl", heifers)
        ws.cell(i, 201).value = heifers
        heifers = setAnimalCount("Marheifers-bodyEl", heifers)
        ws.cell(i, 202).value = heifers
        heifers = setAnimalCount("Aprheifers-bodyEl", heifers)
        ws.cell(i, 203).value = heifers
        heifers = setAnimalCount("Mayheifers-bodyEl", heifers)
        ws.cell(i, 204).value = heifers
        heifers = setAnimalCount("Junheifers-bodyEl", heifers)
        ws.cell(i, 205).value = heifers
        heifers = setAnimalCount("Julheifers-bodyEl", heifers)
        ws.cell(i, 206).value = heifers
        heifers = setAnimalCount("Augheifers-bodyEl", heifers)
        ws.cell(i, 207).value = heifers
        heifers = setAnimalCount("Sepheifers-bodyEl", heifers)
        ws.cell(i, 208).value = heifers
        heifers = setAnimalCount("Octheifers-bodyEl", heifers)
        ws.cell(i, 209).value = heifers
        heifers = setAnimalCount("Novheifers-bodyEl", heifers)
        ws.cell(i, 210).value = heifers
        heifers = setAnimalCount("Decheifers-bodyEl", heifers)
        ws.cell(i, 211).value = heifers

        # steers
        steers = random.randrange(0, 500, 1)
        steers = setAnimalCount("Jansteers-bodyEl", steers)
        ws.cell(i, 212).value = steers
        steers = setAnimalCount("Febsteers-bodyEl", steers)
        ws.cell(i, 213).value = steers
        steers = setAnimalCount("Marsteers-bodyEl", steers)
        ws.cell(i, 214).value = steers
        steers = setAnimalCount("Aprsteers-bodyEl", steers)
        ws.cell(i, 215).value = steers
        steers = setAnimalCount("Maysteers-bodyEl", steers)
        ws.cell(i, 216).value = steers
        steers = setAnimalCount("Junsteers-bodyEl", steers)
        ws.cell(i, 217).value = steers
        steers = setAnimalCount("Julsteers-bodyEl", steers)
        ws.cell(i, 218).value = steers
        steers = setAnimalCount("Augsteers-bodyEl", steers)
        ws.cell(i, 219).value = steers
        steers = setAnimalCount("Sepsteers-bodyEl", steers)
        ws.cell(i, 220).value = steers
        steers = setAnimalCount("Octsteers-bodyEl", steers)
        ws.cell(i, 221).value = steers
        steers = setAnimalCount("Novsteers-bodyEl", steers)
        ws.cell(i, 222).value = steers
        steers = setAnimalCount("Decsteers-bodyEl", steers)
        ws.cell(i, 223).value = steers

        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)

        # primary breed
        actionChains = ActionChains(browser)
        breedList = ['Angus', 'Brahman', 'Charolais', 'Chianina', 'Gelbvieh', 'Hereford', 'Limousin', 'Main Anjou',
                     'Pinzgauer', 'Red Poll', 'Sahiwal', 'Simmental', 'South Devon', 'Tarentaise']
        breedChoice = random.choice(breedList)
        actionChains.move_to_element_with_offset(browser.find_element_by_id("primaryBreedFeedlot-triggerWrap"), 2,
                                                 2).click().perform()
        browser.find_element_by_xpath('//li[text()="' + breedChoice + '"]')
        ws.cell(i, 224).value = breedChoice
        sleep(2)
        typicalMatureweight = browser.find_element_by_xpath(
            '//input[@aria-describedby="typicalMatureWeightField-errorEl"]').text
        ws.cell(i, 225).value = typicalMatureweight

        # heifers weight
        heifersWeihtSpring = round(random.uniform(float(0.6), float(2.5)), 1)
        heifersWeihtSummer = round(random.uniform(float(0.6), float(2.5)), 1)
        heifersWeihtFall = round(random.uniform(float(0.6), float(2.5)), 1)
        heifersWeihtWinter = round(random.uniform(float(0.6), float(2.5)), 1)
        setAnimalCharacterstics("SpringheifersWG-bodyEl", str(heifersWeihtSpring))
        setAnimalCharacterstics("SummerheifersWG-bodyEl", str(heifersWeihtSummer))
        setAnimalCharacterstics("FallheifersWG-bodyEl", str(heifersWeihtFall))
        setAnimalCharacterstics("WinterheifersWG-bodyEl", str(heifersWeihtWinter))
        ws.cell(i, 226).value = heifersWeihtSpring
        ws.cell(i, 227).value = heifersWeihtSummer
        ws.cell(i, 228).value = heifersWeihtFall
        ws.cell(i, 229).value = heifersWeihtWinter

        # Steer Weight
        steersWeihtSpring = round(random.uniform(float(0.6), float(2.5)), 1)
        steersWeihtSummer = round(random.uniform(float(0.6), float(2.5)), 1)
        steersWeihtFall = round(random.uniform(float(0.6), float(2.5)), 1)
        steersWeihtWinter = round(random.uniform(float(0.6), float(2.5)), 1)
        setAnimalCharacterstics("SpringsteersWG-bodyEl", str(steersWeihtSpring))
        setAnimalCharacterstics("SummersteersWG-bodyEl", str(steersWeihtSummer))
        setAnimalCharacterstics("FallsteersWG-bodyEl", str(steersWeihtFall))
        setAnimalCharacterstics("WintersteersWG-bodyEl", str(steersWeihtWinter))
        ws.cell(i, 230).value = steersWeihtSpring
        ws.cell(i, 231).value = steersWeihtSummer
        ws.cell(i, 232).value = steersWeihtFall
        ws.cell(i, 233).value = steersWeihtWinter

        # heifers averahe live weight
        heifersAvgLiveWeight = random.randrange(362, 545, 1)
        heifersAvgLiveWeight = setAnimalWeight("JanheifersLW-bodyEl", heifersAvgLiveWeight)
        ws.cell(i, 234).value = heifersAvgLiveWeight
        heifersAvgLiveWeight = setAnimalWeight("FebheifersLW-bodyEl", heifersAvgLiveWeight)
        ws.cell(i, 235).value = heifersAvgLiveWeight
        heifersAvgLiveWeight = setAnimalWeight("MarheifersLW-bodyEl", heifersAvgLiveWeight)
        ws.cell(i, 236).value = heifersAvgLiveWeight
        heifersAvgLiveWeight = setAnimalWeight("AprheifersLW-bodyEl", heifersAvgLiveWeight)
        ws.cell(i, 237).value = heifersAvgLiveWeight
        heifersAvgLiveWeight = setAnimalWeight("MayheifersLW-bodyEl", heifersAvgLiveWeight)
        ws.cell(i, 238).value = heifersAvgLiveWeight
        heifersAvgLiveWeight = setAnimalWeight("JunheifersLW-bodyEl", heifersAvgLiveWeight)
        ws.cell(i, 239).value = heifersAvgLiveWeight
        heifersAvgLiveWeight = setAnimalWeight("JulheifersLW-bodyEl", heifersAvgLiveWeight)
        ws.cell(i, 240).value = heifersAvgLiveWeight
        heifersAvgLiveWeight = setAnimalWeight("AugheifersLW-bodyEl", heifersAvgLiveWeight)
        ws.cell(i, 241).value = heifersAvgLiveWeight
        heifersAvgLiveWeight = setAnimalWeight("SepheifersLW-bodyEl", heifersAvgLiveWeight)
        ws.cell(i, 242).value = heifersAvgLiveWeight
        heifersAvgLiveWeight = setAnimalWeight("OctheifersLW-bodyEl", heifersAvgLiveWeight)
        ws.cell(i, 243).value = heifersAvgLiveWeight
        heifersAvgLiveWeight = setAnimalWeight("NovheifersLW-bodyEl", heifersAvgLiveWeight)
        ws.cell(i, 244).value = heifersAvgLiveWeight
        heifersAvgLiveWeight = setAnimalWeight("DecheifersLW-bodyEl", heifersAvgLiveWeight)
        ws.cell(i, 245).value = heifersAvgLiveWeight

        # steers avg live weight
        steersAvgLiveweight = random.randrange(362, 545, 1)
        steersAvgLiveweight = setAnimalWeight("JansteersLW-bodyEl", steersAvgLiveweight)
        ws.cell(i, 246).value = steersAvgLiveweight
        steersAvgLiveweight = setAnimalWeight("FebsteersLW-bodyEl", steersAvgLiveweight)
        ws.cell(i, 247).value = steersAvgLiveweight
        steersAvgLiveweight = setAnimalWeight("MarsteersLW-bodyEl", steersAvgLiveweight)
        ws.cell(i, 248).value = steersAvgLiveweight
        steersAvgLiveweight = setAnimalWeight("AprsteersLW-bodyEl", steersAvgLiveweight)
        ws.cell(i, 249).value = steersAvgLiveweight
        steersAvgLiveweight = setAnimalWeight("MaysteersLW-bodyEl", steersAvgLiveweight)
        ws.cell(i, 250).value = steersAvgLiveweight
        steersAvgLiveweight = setAnimalWeight("JunsteersLW-bodyEl", steersAvgLiveweight)
        ws.cell(i, 251).value = steersAvgLiveweight
        steersAvgLiveweight = setAnimalWeight("JulsteersLW-bodyEl", steersAvgLiveweight)
        ws.cell(i, 252).value = steersAvgLiveweight
        steersAvgLiveweight = setAnimalWeight("AugsteersLW-bodyEl", steersAvgLiveweight)
        ws.cell(i, 253).value = steersAvgLiveweight
        steersAvgLiveweight = setAnimalWeight("SepsteersLW-bodyEl", steersAvgLiveweight)
        ws.cell(i, 254).value = steersAvgLiveweight
        steersAvgLiveweight = setAnimalWeight("OctsteersLW-bodyEl", steersAvgLiveweight)
        ws.cell(i, 255).value = steersAvgLiveweight
        steersAvgLiveweight = setAnimalWeight("NovsteersLW-bodyEl", steersAvgLiveweight)
        ws.cell(i, 256).value = steersAvgLiveweight
        steersAvgLiveweight = setAnimalWeight("DecsteersLW-bodyEl", steersAvgLiveweight)
        ws.cell(i, 257).value = steersAvgLiveweight

        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(10)

        ionophores = random.randrange(0, 2, 1)
        if ionophores == 0:
            browser.find_element_by_xpath('//input[@aria-describedby="ionophoretype_0-errorEl"]').click()
            ws.cell(i, 258).value = "Yes"
        else:
            browser.find_element_by_xpath('//input[@aria-describedby="ionophoretype_1-errorEl"]').click()
            ws.cell(i, 258).value = "No"

        fatList = ["fatContentType_0-errorEl", "fatContentType_1-errorEl", "fatContentType_0-errorEl",
                   "fatContentType_0-errorEl"]
        fatChoice = random.choice(fatList)
        browser.find_element_by_xpath('//input[@aria-describedby="' + fatChoice + '"]').click()
        if fatChoice == "fatContentType_0-errorEl":
            ws.cell(i, 259).value = "1% Supplemental Fat"
        elif fatChoice == "fatContentType_1-errorEl":
            ws.cell(i, 259).value = "2% Supplemental Fat"
        elif fatChoice == "fatContentType_2-errorEl":
            ws.cell(i, 259).value = "Four higher added fat content"
        elif fatChoice == "fatContentType_3-errorEl":
            ws.cell(i, 259).value = "No Supplemental fat added"

        grainList = ["grainType_0-errorEl", "grainType_1-errorEl", "grainType_2-errorEl"]
        grainChoice = random.choice(grainList)
        browser.find_element_by_xpath('//input[@aria-describedby="' + grainChoice + '"]').click()
        if grainChoice == "grainType_0-errorEl":
            ws.cell(i, 260).value = "Steam Flaked or High Moisture"
        elif grainChoice == "grainType_1-errorEl":
            ws.cell(i, 260).value = "Unprocessed Or Dry rolled"
        elif grainChoice == "grainType_2-errorEl":
            ws.cell(i, 260).value = "Barley rather than corn or sorghum"

        percentageDietList = ["grainConc_0-errorEl", "grainConc_1-errorEl", "grainConc_2-errorEl"]
        percentageDietChoice = random.choice(percentageDietList)
        actionChains = ActionChains(browser)
        browser.find_element_by_xpath('//input[@aria-describedby="' + percentageDietChoice + '"]').click()

        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)
        if percentageDietChoice == "grainConc_0-errorEl":
            ws.cell(i, 261).value = "More than 60 percent grain"
        elif percentageDietChoice == "grainConc_1-errorEl":
            ws.cell(i, 261).value = "45 to 60 percent grain"
        elif percentageDietChoice == "grainConc_1-errorEl":
            ws.cell(i, 261).value = "Less than 45 percent grain"

        manureSystemTypes("Feedlot")

        browser.find_element_by_xpath('//span[text()="Save & Continue >>"]').click()
        sleep(5)


    if "Beef-Heifer StockersCheck-bodyEl" in animalTypes:
        settingAllAnimalCharactersticsForHeifers()
        animalTypes.remove("Beef-Heifer StockersCheck-bodyEl")

        if "Dairy-Lactating CowsCheck-bodyEl" in animalTypes:
            sleep(2)
            browser.find_element_by_xpath('//span[text()="Continue to Dairy-Lactating Cows >>"]').click()
            settingAllAnimalCharactersticsForLactatingCow()
        if "Beef-Steer StockersCheck-bodyEl" in animalTypes:
            sleep(2)
            browser.find_element_by_xpath('//span[text()="Continue to Beef-Steer Stockers >>"]').click()
            settingAllAnimalCharactersticsForSteers()

        if "Feedlot CattleCheck-bodyEl" in animalTypes:
            sleep(2)
            browser.find_element_by_xpath('//span[text()="Continue to Feedlot Cattle >>"]').click()
            settingAllAnimalCharactersticsForFeedlot()



    elif "Beef-Steer StockersCheck-bodyEl" in animalTypes:
        animalTypes.remove("Beef-Steer StockersCheck-bodyEl")
        settingAllAnimalCharactersticsForSteers()
        if "Dairy-Lactating CowsCheck-bodyEl" in animalTypes:
            sleep(2)
            browser.find_element_by_xpath('//span[text()="Continue to Dairy-Lactating Cows >>"]').click()
            settingAllAnimalCharactersticsForLactatingCow()

        if "Beef-Heifer StockersCheck-bodyEl" in animalTypes:
            sleep(2)
            browser.find_element_by_xpath('//span[text()="Continue to Beef-Heifer Stockers >>"]').click()
            settingAllAnimalCharactersticsForHeifers()

        if "Feedlot CattleCheck-bodyEl" in animalTypes:
            sleep(2)
            browser.find_element_by_xpath('//span[text()="Continue to Feedlot Cattle >>"]').click()
            settingAllAnimalCharactersticsForFeedlot()




    elif "Dairy-Lactating CowsCheck-bodyEl" in animalTypes:
        settingAllAnimalCharactersticsForLactatingCow()

        if "Feedlot CattleCheck-bodyEl" in animalTypes:
            sleep(5)
            browser.find_element_by_xpath('//span[text()="Continue to Feedlot Cattle >>"]').click()
            settingAllAnimalCharactersticsForFeedlot()

    elif "Feedlot CattleCheck-bodyEl" in animalTypes:
        settingAllAnimalCharactersticsForFeedlot()

    animalFutureManagement()
    sleep(2)
    # os.remove("C:/Users/aditi/Downloads/cropOfframp.csv")
    sleep(180)
    try:
        v = browser.find_element_by_xpath('//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1262   "]/div').text
    except:
        v = browser.find_element_by_xpath('//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1029   "]/div').text

    print(v)
    ws.cell(i, 348).value = v

    # browser.find_element_by_xpath('//span[text()="Export Crop Management Data"]').click()
    # sleep(2)
    # browser.find_element_by_xpath('//span[text()="Download Crop Report Averages"]').click()
    # try:
    #     browser.find_element_by_id('tool-1141-toolEl').click()
    # except:
    #     try:
    #         browser.find_element_by_id('tool-1121-toolEl').click()
    #     except:
    #         browser.find_element_by_id('tool-1113-toolEl').click()

    # actionChains = ActionChains(browser)
    # actionChains.move_to_element_with_offset(el , 2, 2).click().perform()
    sleep(3)
    # columns=[]
    # with open("C:/Users/aditi/Downloads/cropOfframp.csv") as f:
    #     reader = csv.reader(f)
    #     reader.next()
    #     for row in reader:
    #         for (i, v) in enumerate(row):
    #             columns[i].append(v)
    # print(columns[0])
    sleep(2)
    browser.find_element_by_xpath('//span[text()="Animal Agriculture"]').click()
    sleep(3)
    try:
        print(browser.find_element_by_xpath('//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1437   "]/div').text)
        j = browser.find_element_by_xpath('//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1437   "]/div').text
        ws.cell(i, 349).value = j
    except:
        try:
            print(browser.find_element_by_xpath(
                '//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1147   "]/div').text)
            j = browser.find_element_by_xpath('//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1147   "]/div').text
            ws.cell(i, 349).value = j
        except:
            try:
                print(browser.find_element_by_xpath('//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1147   "]/div').text)
                j = browser.find_element_by_xpath('//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1147   "]/div').text
                ws.cell(i, 349).value = j
            except:
                try:
                    j=browser.find_element_by_xpath('//td[@class =" x-grid-cell x-grid-cell-gridcolumn-1133   x-grid-cell-last"]/div').text
                    print(j)
                    ws.cell(i, 349).value = j
                    j = browser.find_element_by_xpath('//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1144   x-grid-cell-last"]/div').text
                    print(j)
                    ws.cell(i, 350).value = j
                except:
                    try:
                        print(browser.find_element_by_xpath(
                        '//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1223   "]/div').text)
                        j = browser.find_element_by_xpath(
                        '//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1223   "]/div').text
                        ws.cell(i, 349).value = j
                    except:
                        try:
                            print(browser.find_element_by_xpath(
                            '//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1221   "]/div').text)
                            j = browser.find_element_by_xpath(
                            '//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1221   "]/div').text
                            ws.cell(i, 349).value = j
                        except:
                            print(browser.find_element_by_xpath(
                                '//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1127   "]/div').text)
                            j = browser.find_element_by_xpath(
                                '//td[@class=" x-grid-cell x-grid-cell-gridcolumn-1127   "]/div').text
                            ws.cell(i, 349).value = j




    print("Saved")
    wb.save("D:/Study Material/Research/CometOutput.xlsx")














